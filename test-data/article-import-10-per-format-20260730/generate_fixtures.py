from __future__ import annotations

import csv
import io
import json
import shutil
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
JOURNAL_SLUG = "foundation-model-systems"
DATE_STAMP = "20260730"
ARTICLE_COUNT = 10
CHAPTER_COUNT = 10

COLUMNS = [
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "body_html",
    "html_file",
    "docx_file",
    "markdown_file",
    "cover_image",
    "primary_category_code",
    "primary_category_path",
    "related_category_codes",
    "related_category_paths",
    "notes",
]


def article_title(fmt: str, index: int) -> str:
    return f"Import Test {DATE_STAMP} {fmt.upper()} Article {index:02d}"


def article_slug(fmt: str, index: int) -> str:
    return f"import-test-{DATE_STAMP}-{fmt.lower()}-{index:02d}"


def chapter_text(fmt: str, article_index: int, chapter_index: int) -> str:
    return (
        f"This is chapter {chapter_index} of {article_title(fmt, article_index)}. "
        "It is deterministic test content for verifying article import, preview, "
        "draft creation, Unicode text handling, and repeated idempotent imports. "
        f"中文测试段落：第 {chapter_index} 章，用于验证正文转换和草稿导入。"
    )


def html_body(fmt: str, index: int) -> str:
    sections = [f"<h1>{escape(article_title(fmt, index))}</h1>"]
    for chapter in range(1, CHAPTER_COUNT + 1):
        sections.append(f"<h2>Chapter {chapter}: Import Verification</h2>")
        sections.append(f"<p>{escape(chapter_text(fmt, index, chapter))}</p>")
        sections.append(
            "<ul><li>Expected state: draft</li>"
            "<li>No automatic approval</li><li>No automatic placement or static publish</li></ul>"
        )
    return "".join(sections)


def row_for(
    fmt: str, index: int, *, source_kind: str = "body_html", source_path: str = ""
) -> dict[str, str]:
    row = {column: "" for column in COLUMNS}
    row.update(
        {
            "journal_slug": JOURNAL_SLUG,
            "title": article_title(fmt, index),
            "slug": article_slug(fmt, index),
            "article_type": "ai_article",
            "authors": f"Test Author {index:02d}",
            "ai_co_authors": "AI Author Forum Test Assistant",
            "abstract": f"Automated {fmt.upper()} import fixture containing {CHAPTER_COUNT} chapters.",
            "keywords": f"import test, {fmt.lower()}, fixture, 10 chapters",
            "publication_date": f"2026-07-{min(20 + index, 30):02d}T09:00:00+08:00",
            "notes": "Testing only. Import must remain draft and must not auto-publish.",
        }
    )
    if source_kind == "body_html":
        row["body_html"] = html_body(fmt, index)
    else:
        row[source_kind] = source_path
    return row


def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "articles"
    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in COLUMNS])
    fill = PatternFill("solid", fgColor="1D5E8C")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(column) + 2, 16), 34
        )

    metadata = workbook.create_sheet("_meta")
    metadata.sheet_state = "hidden"
    metadata.append(["template_type", "article_import"])
    metadata.append(["template_version", 2])
    metadata.append(["scope", "global"])
    metadata.append(["journal_slug", ""])
    metadata.append(
        [
            "business_rule",
            "draft_only;no_review_approval;no_placement;no_static_publish",
        ]
    )
    workbook.save(path)


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def markdown_text(fmt: str, index: int, *, include_front_matter: bool = True) -> str:
    chunks: list[str] = []
    if include_front_matter:
        chunks.append(
            "---\n"
            f"title: {article_title(fmt, index)}\n"
            f"slug: {article_slug(fmt, index)}\n"
            f"journal_slug: {JOURNAL_SLUG}\n"
            "article_type: ai_article\n"
            f"authors:\n  - Test Author {index:02d}\n"
            "ai_co_authors:\n  - AI Author Forum Test Assistant\n"
            f"abstract: Markdown import test article with {CHAPTER_COUNT} chapters.\n"
            f"keywords: import test, markdown, fixture, article {index:02d}\n"
            f"publication_date: 2026-07-{min(20 + index, 30):02d}T09:00:00+08:00\n"
            "---\n\n"
        )
    chunks.append(f"# {article_title(fmt, index)}\n\n")
    for chapter in range(1, CHAPTER_COUNT + 1):
        chunks.append(f"## Chapter {chapter}: Import Verification\n\n")
        chunks.append(chapter_text(fmt, index, chapter) + "\n\n")
        chunks.append(
            "- Expected state: draft\n- No automatic approval\n- No automatic placement or static publish\n\n"
        )
    return "".join(chunks)


def paragraph_xml(text: str, style: str | None = None) -> str:
    style_xml = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ""
    return f'<w:p>{style_xml}<w:r><w:t xml:space="preserve">{escape(text)}</w:t></w:r></w:p>'


def create_docx_bytes(fmt: str, index: int) -> bytes:
    title = article_title(fmt, index)
    paragraphs = [paragraph_xml(title, "Heading1")]
    for chapter in range(1, CHAPTER_COUNT + 1):
        paragraphs.append(
            paragraph_xml(f"Chapter {chapter}: Import Verification", "Heading2")
        )
        paragraphs.append(paragraph_xml(chapter_text(fmt, index, chapter)))
        paragraphs.append(
            paragraph_xml(
                "Expected state: draft; no automatic approval, placement, or static publish."
            )
        )

    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(paragraphs)
        + '<w:sectPr><w:pgSz w:w="12240" w:h="15840"/><w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/></w:sectPr>'
        "</w:body></w:document>"
    )
    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:style w:type="paragraph" w:default="1" w:styleId="Normal"><w:name w:val="Normal"/></w:style>
  <w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="Heading 1"/><w:basedOn w:val="Normal"/><w:qFormat/></w:style>
  <w:style w:type="paragraph" w:styleId="Heading2"><w:name w:val="Heading 2"/><w:basedOn w:val="Normal"/><w:qFormat/></w:style>
</w:styles>"""
    core_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>{escape(title)}</dc:title>
  <dc:creator>Test Author {index:02d}</dc:creator>
  <dc:subject>DOCX import fixture</dc:subject>
  <dc:description>DOCX import test article containing {CHAPTER_COUNT} chapters.</dc:description>
  <cp:keywords>import test, docx, fixture</cp:keywords>
  <cp:lastModifiedBy>AI Author Forum Test Generator</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">2026-07-30T01:00:00Z</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">2026-07-30T01:00:00Z</dcterms:modified>
</cp:coreProperties>"""
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
</Types>"""
    root_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
</Relationships>"""
    document_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""

    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("docProps/core.xml", core_xml)
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/styles.xml", styles_xml)
        archive.writestr("word/_rels/document.xml.rels", document_rels)
    return stream.getvalue()


def write_docx(path: Path, fmt: str, index: int) -> None:
    path.write_bytes(create_docx_bytes(fmt, index))


def reset_output() -> None:
    for name in ["01-xlsx", "02-csv", "03-zip", "04-docx", "05-markdown"]:
        target = ROOT / name
        if target.exists():
            shutil.rmtree(target)
        target.mkdir(parents=True)


def main() -> None:
    reset_output()

    xlsx_rows = [row_for("xlsx", index) for index in range(1, ARTICLE_COUNT + 1)]
    write_xlsx(ROOT / "01-xlsx" / "articles-10.xlsx", xlsx_rows)

    csv_rows = [row_for("csv", index) for index in range(1, ARTICLE_COUNT + 1)]
    write_csv(ROOT / "02-csv" / "articles-10.csv", csv_rows)

    package_work = ROOT / ".zip-work"
    if package_work.exists():
        shutil.rmtree(package_work)
    (package_work / "documents").mkdir(parents=True)
    zip_rows = []
    for index in range(1, ARTICLE_COUNT + 1):
        if index <= 5:
            rel = f"documents/zip-article-{index:02d}.docx"
            (package_work / rel).write_bytes(create_docx_bytes("zip", index))
            zip_rows.append(
                row_for("zip", index, source_kind="docx_file", source_path=rel)
            )
        else:
            rel = f"documents/zip-article-{index:02d}.md"
            (package_work / rel).write_text(
                markdown_text("zip", index, include_front_matter=False),
                encoding="utf-8",
            )
            zip_rows.append(
                row_for("zip", index, source_kind="markdown_file", source_path=rel)
            )
    write_xlsx(package_work / "articles.xlsx", zip_rows)
    (package_work / "README.txt").write_text(
        "AI Author Forum ZIP import test package\n"
        "- 10 articles total\n- Articles 01-05 use DOCX\n- Articles 06-10 use Markdown\n"
        "- Each article contains 10 chapters\n- All imported articles must remain drafts\n",
        encoding="utf-8",
    )
    zip_path = ROOT / "03-zip" / "articles-10-mixed-documents.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(package_work.rglob("*")):
            if path.is_file():
                archive.write(path, path.relative_to(package_work).as_posix())
    shutil.rmtree(package_work)

    for index in range(1, ARTICLE_COUNT + 1):
        write_docx(ROOT / "04-docx" / f"docx-article-{index:02d}.docx", "docx", index)
        (ROOT / "05-markdown" / f"markdown-article-{index:02d}.md").write_text(
            markdown_text("markdown", index), encoding="utf-8"
        )

    manifest = {
        "generated_on": "2026-07-30",
        "journal_slug": JOURNAL_SLUG,
        "article_count_per_format": ARTICLE_COUNT,
        "chapter_count_per_article": CHAPTER_COUNT,
        "formats": {
            "xlsx": ["01-xlsx/articles-10.xlsx"],
            "csv": ["02-csv/articles-10.csv"],
            "zip": ["03-zip/articles-10-mixed-documents.zip"],
            "docx": [
                f"04-docx/docx-article-{i:02d}.docx"
                for i in range(1, ARTICLE_COUNT + 1)
            ],
            "markdown": [
                f"05-markdown/markdown-article-{i:02d}.md"
                for i in range(1, ARTICLE_COUNT + 1)
            ],
        },
    }
    (ROOT / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    readme = f"""# 文章导入测试文件（每种格式 10 篇）

生成日期：2026-07-30
默认测试子期刊：`{JOURNAL_SLUG}`

## 文件说明

| 格式 | 文件 | 文章数量 | 使用方式 |
|---|---|---:|---|
| XLSX | `01-xlsx/articles-10.xlsx` | 10 | 全局导入，文件内已填写 `journal_slug` |
| CSV | `02-csv/articles-10.csv` | 10 | 全局导入，编码为 UTF-8-SIG |
| ZIP | `03-zip/articles-10-mixed-documents.zip` | 10 | 5 篇 DOCX + 5 篇 Markdown，由 `articles.xlsx` 引用 |
| DOCX | `04-docx/*.docx` | 10 个单篇文件 | 每个文件单独上传；选择默认子期刊和 `AI Article` 类型 |
| Markdown | `05-markdown/*.md` | 10 个单篇文件 | 每个文件单独上传；选择默认子期刊，元数据在 Front Matter 中 |

每篇测试文章正文都包含 **10 个章节**，因此也可以验证长正文、标题层级和转换结果。

## 后台测试步骤

1. 进入“文章管理 → 一键导入文章”。
2. XLSX、CSV、ZIP 可直接上传并预检。
3. 单篇 DOCX：选择默认子期刊 `{JOURNAL_SLUG}`，文章类型选择 `AI Article`，再逐个上传。
4. 单篇 Markdown：选择默认子期刊 `{JOURNAL_SLUG}`，再逐个上传。
5. 核对预检结果后人工确认。
6. 确认文章仅进入草稿，未自动审核、投放或静态发布。

## 幂等与隔离

- 各格式使用不同 slug，不会互相覆盖。
- 重复导入同一文件应命中同一子期刊 + slug，适合测试更新/幂等逻辑。
- 文件仅用于测试，请勿在生产站点直接确认导入。
"""
    (ROOT / "README.md").write_text(readme, encoding="utf-8")

    print(f"Generated fixtures in {ROOT}")
    print(f"Total files: {sum(1 for p in ROOT.rglob('*') if p.is_file())}")


if __name__ == "__main__":
    main()
