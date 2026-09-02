from io import BytesIO
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.test import TestCase
from openpyxl import Workbook

from ai_author_forum.articles.models import ArticlePage
from ai_author_forum.articles.services import sync_imported_article
from ai_author_forum.journals.models import Journal, StaticArticle
from ai_author_forum.journals.services import (
    get_active_journals,
    get_journal_context,
    import_package,
)
from ai_author_forum.placements.models import ArticlePlacement, LayoutSlot
from ai_author_forum.site_settings.models import AuditAction, AuditLog, AuditStatus
from ai_author_forum.test_helpers import (
    formally_approve_test_article,
    grant_business_super_admin,
)


class ImportServiceTests(TestCase):
    def _build_package(self):
        buffer = BytesIO()
        with ZipFile(buffer, "w") as zf:
            journal_wb = Workbook()
            ws = journal_wb.active
            ws.append(["journal_name", "slug", "az_group", "status"])
            ws.append(["AI Ethics Forum", "ai-ethics-forum", "A", "active"])
            journal_stream = BytesIO()
            journal_wb.save(journal_stream)
            zf.writestr("journals.xlsx", journal_stream.getvalue())

            article_wb = Workbook()
            ws = article_wb.active
            ws.append(["journal_slug", "title", "slug", "article_type", "body_html"])
            ws.append(
                [
                    "ai-ethics-forum",
                    "Responsible Co-authoring",
                    "responsible-co-authoring",
                    "ai_article",
                    "<html><body><h1>Responsible Co-authoring</h1></body></html>",
                ]
            )
            article_stream = BytesIO()
            article_wb.save(article_stream)
            zf.writestr("articles.xlsx", article_stream.getvalue())
        buffer.seek(0)
        buffer.name = "bundle.zip"
        return buffer

    def test_import_package_creates_journal_and_article(self):
        package = self._build_package()
        result = import_package(package, operator=None)

        self.assertEqual(result.journal_created, 1)
        self.assertEqual(result.article_created, 1)
        self.assertTrue(Journal.objects.filter(slug="ai-ethics-forum").exists())
        self.assertTrue(
            StaticArticle.objects.filter(
                slug="responsible-co-authoring",
                journal__slug="ai-ethics-forum",
            ).exists()
        )

        article = StaticArticle.objects.get(slug="responsible-co-authoring")
        self.assertTrue(
            ArticlePage.objects.filter(source_static_article=article).exists()
        )
        canonical = ArticlePage.objects.get(source_static_article=article)
        self.assertEqual(canonical.review_status, ArticlePage.ReviewStatus.DRAFT)
        self.assertFalse(canonical.live)
        self.assertFalse(ArticlePlacement.objects.filter(article=canonical).exists())
        actor = grant_business_super_admin(
            get_user_model().objects.create_superuser(
                username="import-service-admin",
                email="import-service-admin@example.com",
                password="test",
            )
        )
        formally_approve_test_article(canonical, actor=actor)
        self.assertEqual(list(get_active_journals()), [article.journal])
        context = get_journal_context("ai-ethics-forum")
        self.assertEqual(context["journal"], article.journal)
        self.assertEqual(context["articles"], [])

        ArticlePlacement.objects.create(
            article=canonical,
            slot=LayoutSlot.objects.get_or_create(
                code="journal_featured",
                defaults={
                    "title": "Journal featured",
                    "scope": LayoutSlot.Scope.JOURNAL,
                    "max_items": 20,
                },
            )[0],
            target_type=ArticlePlacement.TargetType.JOURNAL,
            target_slug=article.journal.slug,
        )
        context = get_journal_context("ai-ethics-forum")
        self.assertEqual(context["articles"], [canonical])

        import_statuses = list(
            AuditLog.objects.filter(action=AuditAction.IMPORT)
            .order_by("created_at")
            .values_list("status", flat=True)
        )
        self.assertEqual(import_statuses[0], AuditStatus.STARTED)
        self.assertTrue(
            all(status == AuditStatus.SUCCESS for status in import_statuses[1:])
        )

    def test_legacy_publication_status_never_bypasses_moderation(self):
        journal = Journal.objects.create(
            name="Legacy Status Journal",
            slug="legacy-status-journal",
            az_group="L",
        )
        legacy_statuses = ["approved", "placed", "built", "published", "offline"]

        for index, legacy_status in enumerate(legacy_statuses, start=1):
            with self.subTest(legacy_status=legacy_status):
                legacy = StaticArticle.objects.create(
                    journal=journal,
                    title=f"Legacy Status {legacy_status}",
                    slug=f"legacy-status-{index}",
                    review_status=legacy_status,
                    build_version=f"legacy-v{index}",
                )
                canonical = sync_imported_article(legacy)

                self.assertEqual(
                    canonical.review_status, ArticlePage.ReviewStatus.DRAFT
                )
                self.assertEqual(canonical.publication_status, "")
                self.assertFalse(canonical.live)
                self.assertEqual(canonical.build_version, f"legacy-v{index}")
                self.assertEqual(canonical.published_version, "")
                self.assertFalse(
                    ArticlePlacement.objects.filter(article=canonical).exists()
                )

    def test_single_invalid_row_does_not_roll_back_valid_rows(self):
        buffer = BytesIO()
        with ZipFile(buffer, "w") as package:
            journal_workbook = Workbook()
            sheet = journal_workbook.active
            sheet.append(["journal_name", "slug", "az_group", "status"])
            sheet.append(["AI Safety Review", "ai-safety-review", "A", "active"])
            stream = BytesIO()
            journal_workbook.save(stream)
            package.writestr("journals.xlsx", stream.getvalue())

            article_workbook = Workbook()
            sheet = article_workbook.active
            sheet.append(
                [
                    "journal_slug",
                    "title",
                    "slug",
                    "article_type",
                    "body_html",
                    "sort_order",
                ]
            )
            sheet.append(
                [
                    "ai-safety-review",
                    "Valid article",
                    "valid-article",
                    "ai_article",
                    "<p>valid</p>",
                    10,
                ]
            )
            sheet.append(
                [
                    "ai-safety-review",
                    "Invalid article",
                    "invalid-article",
                    "ai_article",
                    "<p>invalid</p>",
                    "bad-order",
                ]
            )
            stream = BytesIO()
            article_workbook.save(stream)
            package.writestr("articles.xlsx", stream.getvalue())
        buffer.seek(0)
        buffer.name = "partial-package.zip"

        result = import_package(buffer, operator=None)

        self.assertEqual(result.article_created, 1)
        self.assertEqual(len(result.article_issues), 1)
        self.assertTrue(StaticArticle.objects.filter(slug="valid-article").exists())
        self.assertFalse(StaticArticle.objects.filter(slug="invalid-article").exists())
        self.assertEqual(result.article_job.rows.count(), 2)
        self.assertEqual(result.article_job.failed_rows, 1)
        self.assertTrue(result.article_job.error_report)

    def test_duplicate_journal_and_article_slugs_update_existing_records(self):
        package = self._build_package()
        first = import_package(package, operator=None)

        replacement = self._build_package_with_article_title(
            "Responsible Co-authoring, revised"
        )
        second = import_package(replacement, operator=None)

        self.assertEqual(first.journal_created, 1)
        self.assertEqual(first.article_created, 1)
        self.assertEqual(second.journal_created, 0)
        self.assertEqual(second.journal_updated, 1)
        self.assertEqual(second.article_created, 0)
        self.assertEqual(second.article_updated, 1)
        self.assertEqual(Journal.objects.count(), 1)
        self.assertEqual(StaticArticle.objects.count(), 1)
        self.assertEqual(ArticlePage.objects.count(), 1)
        self.assertEqual(
            ArticlePage.objects.get(
                source_static_article__slug="responsible-co-authoring"
            ).review_status,
            ArticlePage.ReviewStatus.DRAFT,
        )
        self.assertEqual(
            StaticArticle.objects.get(slug="responsible-co-authoring").title,
            "Responsible Co-authoring, revised",
        )

    def test_missing_article_media_fails_only_the_referencing_row(self):
        buffer = BytesIO()
        with ZipFile(buffer, "w") as package:
            journal_workbook = Workbook()
            sheet = journal_workbook.active
            sheet.append(["journal_name", "slug", "az_group", "status"])
            sheet.append(["AI Safety Review", "ai-safety-review", "A", "active"])
            stream = BytesIO()
            journal_workbook.save(stream)
            package.writestr("journals.xlsx", stream.getvalue())

            article_workbook = Workbook()
            sheet = article_workbook.active
            sheet.append(
                [
                    "journal_slug",
                    "title",
                    "slug",
                    "article_type",
                    "cover_image",
                    "body_html",
                ]
            )
            sheet.append(
                [
                    "ai-safety-review",
                    "Valid article",
                    "valid-article",
                    "ai_article",
                    "",
                    "<p>valid</p>",
                ]
            )
            sheet.append(
                [
                    "ai-safety-review",
                    "Missing media article",
                    "missing-media-article",
                    "ai_article",
                    "missing-cover.png",
                    "<p>invalid</p>",
                ]
            )
            stream = BytesIO()
            article_workbook.save(stream)
            package.writestr("articles.xlsx", stream.getvalue())
        buffer.seek(0)
        buffer.name = "missing-media-package.zip"

        result = import_package(buffer, operator=None)

        self.assertEqual(result.article_created, 1)
        self.assertEqual(len(result.article_issues), 1)
        self.assertIn("Missing cover_image", result.article_issues[0].message)
        self.assertTrue(StaticArticle.objects.filter(slug="valid-article").exists())
        self.assertFalse(
            StaticArticle.objects.filter(slug="missing-media-article").exists()
        )

    def test_zip_path_traversal_is_rejected(self):
        buffer = BytesIO()
        with ZipFile(buffer, "w") as package:
            package.writestr("../evil.txt", "outside")
        buffer.seek(0)
        buffer.name = "unsafe-package.zip"

        with self.assertRaisesMessage(ValidationError, "Unsafe package path"):
            import_package(buffer, operator=None)

    def _build_package_with_article_title(self, title):
        buffer = BytesIO()
        with ZipFile(buffer, "w") as zf:
            journal_wb = Workbook()
            ws = journal_wb.active
            ws.append(["journal_name", "slug", "az_group", "status"])
            ws.append(["AI Ethics Forum Updated", "ai-ethics-forum", "A", "active"])
            journal_stream = BytesIO()
            journal_wb.save(journal_stream)
            zf.writestr("journals.xlsx", journal_stream.getvalue())

            article_wb = Workbook()
            ws = article_wb.active
            ws.append(["journal_slug", "title", "slug", "article_type", "body_html"])
            ws.append(
                [
                    "ai-ethics-forum",
                    title,
                    "responsible-co-authoring",
                    "ai_article",
                    f"<html><body><h1>{title}</h1></body></html>",
                ]
            )
            article_stream = BytesIO()
            article_wb.save(article_stream)
            zf.writestr("articles.xlsx", article_stream.getvalue())
        buffer.seek(0)
        buffer.name = "replacement-bundle.zip"
        return buffer
