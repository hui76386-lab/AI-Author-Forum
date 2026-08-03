from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch
from urllib.parse import parse_qs, urlparse
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from ai_author_forum.journals.models import (
    ArticleImportJob,
    ImportJobStatus,
    ImportRowStatus,
    Journal,
    JournalImportJob,
    StaticArticle,
)
from ai_author_forum.journals.publishing import start_import_publish_process


class JournalImportDashboardTests(TestCase):
    def _build_package(self, *, invalid_article=False, add_blank_row=False):
        buffer = BytesIO()
        with ZipFile(buffer, "w") as zf:
            journal_wb = Workbook()
            ws = journal_wb.active
            ws.append(["journal_name", "slug", "az_group", "status"])
            ws.append(["AI Ethics Forum", "ai-ethics-forum", "A", "active"])
            journal_stream = BytesIO()
            journal_wb.save(journal_stream)
            zf.writestr("journals.xlsx", journal_stream.getvalue())

            article_wb = Workbook()
            ws = article_wb.active
            ws.append(
                [
                    "journal_slug",
                    "title",
                    "slug",
                    "article_type",
                    "body_html",
                    "sort_order",
                ]
            )
            ws.append(
                [
                    "ai-ethics-forum",
                    "Responsible Co-authoring",
                    "responsible-co-authoring",
                    "ai_article",
                    "<html><body><h1>Responsible Co-authoring</h1></body></html>",
                    "not-an-integer" if invalid_article else 10,
                ]
            )
            if add_blank_row:
                ws.append([" ", " ", " ", " ", " ", " "])
            article_stream = BytesIO()
            article_wb.save(article_stream)
            zf.writestr("articles.xlsx", article_stream.getvalue())
        buffer.seek(0)
        buffer.name = "bundle.zip"
        return buffer

    def _login_superuser(self):
        user = get_user_model().objects.create_superuser(
            "admin", "admin@example.com", "password"
        )
        self.client.force_login(user)
        return user

    def _login_importer(self):
        user = get_user_model().objects.create_user("importer", password="password")
        user.is_staff = True
        user.save(update_fields=("is_staff",))
        user.user_permissions.add(
            Permission.objects.get(codename="access_admin"),
            Permission.objects.get(
                content_type__app_label="site_settings", codename="import_journals"
            ),
        )
        self.client.force_login(user)
        return user

    def _upload_preview(self, package=None):
        response = self.client.post(
            reverse("journals_import_dashboard"),
            data={"package": package or self._build_package()},
        )
        self.assertEqual(response.status_code, 302)
        query = parse_qs(urlparse(response.url).query)
        return (
            JournalImportJob.objects.get(pk=query["journal_job"][0]),
            ArticleImportJob.objects.get(pk=query["article_job"][0]),
            response.url,
        )

    def test_template_download_matches_importer_columns(self):
        self._login_superuser()
        response = self.client.get(reverse("journals_import_template"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        with ZipFile(BytesIO(response.content)) as package:
            self.assertIn("journals.xlsx", package.namelist())
            self.assertIn("articles.xlsx", package.namelist())
            journal_book = load_workbook(BytesIO(package.read("journals.xlsx")))
            article_book = load_workbook(BytesIO(package.read("articles.xlsx")))
        self.assertEqual(
            [cell.value for cell in journal_book.active[1]][:4],
            ["journal_name", "journal_name_cn", "slug", "az_group"],
        )
        self.assertEqual(
            [cell.value for cell in article_book.active[1]][:4],
            ["journal_slug", "title", "slug", "article_type"],
        )

    def test_upload_creates_row_level_preview_without_business_writes(self):
        self._login_superuser()
        journal_job, article_job, preview_url = self._upload_preview(
            self._build_package(add_blank_row=True)
        )

        self.assertEqual(journal_job.status, ImportJobStatus.READY)
        self.assertEqual(article_job.status, ImportJobStatus.READY)
        self.assertEqual(journal_job.summary["created"], 1)
        self.assertEqual(article_job.summary["created"], 1)
        self.assertEqual(article_job.summary["skipped"], 1)
        self.assertEqual(journal_job.rows.get().action, "create")
        self.assertEqual(
            set(article_job.rows.values_list("action", flat=True)),
            {"create", "skip"},
        )
        self.assertEqual(Journal.objects.count(), 0)
        self.assertEqual(StaticArticle.objects.count(), 0)

        page = self.client.get(preview_url)
        self.assertContains(page, "逐行预览")
        self.assertContains(page, "确认导入")
        self.assertContains(page, "Responsible Co-authoring")

    def test_confirmed_preview_queues_background_import_and_publish(self):
        user = self._login_superuser()
        journal_job, article_job, _ = self._upload_preview()

        with TemporaryDirectory(dir=Path.cwd()) as temp_dir:
            queue_root = Path(temp_dir) / "queue-root"
            with override_settings(AI_AUTHOR_FORUM_IMPORT_QUEUE_ROOT=str(queue_root)):
                with patch(
                    "ai_author_forum.journals.wagtail_hooks.start_import_publish_process",
                    return_value=SimpleNamespace(pid=1234),
                ) as start_process:
                    response = self.client.post(
                        reverse("journals_import_confirm"),
                        data={
                            "journal_job_id": journal_job.pk,
                            "article_job_id": article_job.pk,
                            "publish_static_site": "on",
                        },
                    )

        self.assertEqual(response.status_code, 302)
        start_process.assert_called_once()
        call_kwargs = start_process.call_args.kwargs
        self.assertTrue(call_kwargs["publish_static_site"])
        self.assertEqual(call_kwargs["operator_id"], user.pk)
        self.assertEqual(call_kwargs["preview_journal_job_id"], journal_job.pk)
        self.assertEqual(call_kwargs["preview_article_job_id"], article_job.pk)
        self.assertTrue(
            str(call_kwargs["package_path"]).startswith(str(queue_root.resolve()))
        )
        journal_job.refresh_from_db()
        article_job.refresh_from_db()
        self.assertEqual(journal_job.status, ImportJobStatus.PENDING)
        self.assertEqual(article_job.status, ImportJobStatus.PENDING)
        self.assertEqual(journal_job.summary["background_pid"], 1234)

    def test_importer_without_publish_permission_cannot_confirm_publish(self):
        self._login_importer()
        journal_job, article_job, _ = self._upload_preview()

        with patch(
            "ai_author_forum.journals.wagtail_hooks.start_import_publish_process"
        ) as start_process:
            response = self.client.post(
                reverse("journals_import_confirm"),
                data={
                    "journal_job_id": journal_job.pk,
                    "article_job_id": article_job.pk,
                    "publish_static_site": "on",
                },
            )

        self.assertEqual(response.status_code, 302)
        start_process.assert_not_called()
        journal_job.refresh_from_db()
        self.assertEqual(journal_job.status, ImportJobStatus.READY)

    def test_invalid_row_is_retained_and_error_report_is_downloadable(self):
        self._login_superuser()
        journal_job, article_job, _ = self._upload_preview(
            self._build_package(invalid_article=True)
        )

        self.assertEqual(journal_job.summary["created"], 1)
        self.assertEqual(article_job.summary["failed"], 1)
        row = article_job.rows.get()
        self.assertEqual(row.status, ImportRowStatus.FAILED)
        self.assertIn("whole number", row.error_message)
        self.assertTrue(article_job.error_report)

        response = self.client.get(
            reverse(
                "journals_import_error_report",
                kwargs={"scope": "articles", "job_id": article_job.pk},
            )
        )
        self.assertEqual(response.status_code, 200)
        body = b"".join(response.streaming_content).decode("utf-8-sig")
        self.assertIn("row_number,external_id,journal_code", body)
        self.assertIn(
            "error_code,error_message,field_name,raw_value,error_type,suggestion,scope,payload",
            body,
        )
        self.assertIn("whole number", body)

    def test_status_endpoint_exposes_static_publish_job(self):
        self._login_superuser()
        journal_job, article_job, _ = self._upload_preview()
        journal_job.status = ImportJobStatus.COMPLETED
        journal_job.summary["static_publish_job"] = {
            "id": 42,
            "status": "succeeded",
            "version": "release-42",
        }
        journal_job.save(update_fields=("status", "summary"))
        article_job.status = ImportJobStatus.COMPLETED
        article_job.save(update_fields=("status",))

        response = self.client.get(
            reverse("journals_import_status"),
            data={"journal_job": journal_job.pk, "article_job": article_job.pk},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["terminal"])
        self.assertEqual(payload["jobs"][0]["static_publish_job"]["id"], 42)
        self.assertIn(
            "/static-publish/jobs/42/", payload["jobs"][0]["static_publish_url"]
        )


class ImportProcessCommandTests(TestCase):
    def test_background_process_uses_unified_publish_and_preview_flags(self):
        package_path = Path("queued-package.zip").resolve()

        with patch("ai_author_forum.journals.publishing.subprocess.Popen") as popen:
            start_import_publish_process(
                package_path=package_path,
                dry_run=False,
                publish_static_site=True,
                operator_id=37,
                preview_journal_job_id=11,
                preview_article_job_id=12,
            )

        args = popen.call_args.args[0]
        self.assertIn("import_journal_package", args)
        self.assertIn("--publish-static-site", args)
        self.assertIn("--operator-id", args)
        self.assertIn("37", args)
        self.assertIn("--preview-journal-job-id", args)
        self.assertIn("11", args)
        self.assertIn("--preview-article-job-id", args)
        self.assertIn("12", args)
        self.assertNotIn("--static-output-dir", args)
        self.assertNotIn("--clear-static-output", args)
