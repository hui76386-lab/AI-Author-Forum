from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

JOURNAL_COLUMNS = [
    "journal_name",
    "journal_name_cn",
    "slug",
    "az_group",
    "status",
    "sort_order",
    "seo_title",
    "seo_description",
    "homepage_intro",
    "cover_image",
    "metrics_image",
    "static_site_path",
    "target_article_count",
    "notes",
]

ARTICLE_COLUMNS = [
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "status",
    "sort_order",
    "is_pinned",
    "cover_image",
    "body_html",
    "html_file",
    "build_version",
    "static_output_path",
    "primary_category_code",
    "primary_category_path",
    "related_category_codes",
    "related_category_paths",
    "main_site_slot",
    "main_site_slot_name",
    "main_site_slot_title",
    "main_site_slot_summary",
    "main_site_slot_order",
    "main_site_slot_pinned",
    "journal_slot",
    "journal_slot_name",
    "journal_slot_title",
    "journal_slot_summary",
    "journal_slot_order",
    "journal_slot_pinned",
    "notes",
]


def _workbook_bytes(columns: list[str], example: list[object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "import"
    sheet.append(columns)
    sheet.append(example)
    fill = PatternFill("solid", fgColor="1D5E8C")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[cell_column(index)].width = min(
            max(len(column) + 2, 14), 32
        )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def cell_column(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def build_import_template_package() -> bytes:
    journal_example = [
        "AI Ethics Forum",
        "人工智能伦理论坛",
        "ai-ethics-forum",
        "A",
        "active",
        10,
        "AI Ethics Forum",
        "Research and commentary on responsible AI.",
        "Journal homepage introduction.",
        "journal-cover.jpg",
        "journal-metrics.png",
        "journals/ai-ethics-forum/index.html",
        100,
        "Optional operator note",
    ]
    article_example = [
        "ai-ethics-forum",
        "Responsible Co-authoring",
        "responsible-co-authoring",
        "ai_article",
        "Example Author",
        "Example AI",
        "Article abstract.",
        "AI ethics, authorship",
        "2026-07-19T09:00:00+08:00",
        "approved",
        10,
        False,
        "article-cover.jpg",
        "<h1>Responsible Co-authoring</h1><p>Article body.</p>",
        "",
        "",
        "",
        "AIEF-RESEARCH-GENAI",
        "Research > Generative AI",
        "AIEF-ETHICS;AIEF-LLM",
        "Society > AI Ethics;Research > LLM",
        "home_featured",
        "Homepage featured",
        "",
        "",
        10,
        True,
        "journal_featured",
        "Journal featured",
        "",
        "",
        10,
        False,
        "Optional operator note",
    ]
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as package:
        package.writestr(
            "journals.xlsx", _workbook_bytes(JOURNAL_COLUMNS, journal_example)
        )
        package.writestr(
            "articles.xlsx", _workbook_bytes(ARTICLE_COLUMNS, article_example)
        )
        package.writestr(
            "README.txt",
            "Upload this zip after editing journals.xlsx and articles.xlsx.\n"
            "Put referenced images under media/. Keep file names unique.\n"
            "You may remove either workbook when importing only journals or articles.\n"
            "Article category fields: primary_category_code, primary_category_path, "
            "related_category_codes, related_category_paths.\n"
            "Use English semicolons (;) between related categories and ` > ` between "
            "path levels. If code and path are both supplied, they must identify the "
            "same existing category. Categories are never created automatically.\n"
            "Import apply only writes StaticArticle records and their category assignments; "
            "ArticlePage conversion and moderation are separate steps.\n",
        )
        package.writestr("media/.keep", "")
    return output.getvalue()
