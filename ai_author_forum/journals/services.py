from __future__ import annotations

import csv
import json
import shutil
import tempfile
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.text import slugify
from openpyxl import load_workbook
from PIL import Image as PILImage

from ai_author_forum.articles.services import get_articles_by_journal
from ai_author_forum.images.models import CustomImage
from ai_author_forum.journals.models import (
    ArticleImportJob,
    ArticleImportRow,
    ArticleReviewStatus,
    ArticleType,
    AssetBindingStatus,
    ImportJobStatus,
    ImportRowStatus,
    Journal,
    JournalAssetBinding,
    JournalCategory,
    JournalImportJob,
    JournalImportRow,
    JournalStatus,
    StaticArticle,
)
from ai_author_forum.site_settings.models import AuditAction, AuditStatus
from ai_author_forum.site_settings.services import record_audit_event
from ai_author_forum.utils.i18n import localized_journal_name
from ai_author_forum.utils.public_i18n import (
    localized_journal_description,
    localized_journal_intro,
    localized_journal_seo_title,
)

from .validators import scan_mapping_for_suspicious_text, truncate_text

REQUIRED_JOURNAL_COLUMNS = {"journal_name", "slug", "az_group"}
REQUIRED_ARTICLE_COLUMNS = {"journal_slug", "title", "slug", "article_type"}


def get_active_journals():
    """Return journals that are eligible for frontend and static publishing."""
    return Journal.objects.filter(status=JournalStatus.ACTIVE).order_by(
        "sort_order", "name", "pk"
    )


def get_journal_context(slug: str, at=None):
    """Return the canonical static publishing context for one active journal."""
    journal = get_active_journals().get(slug=slug)
    article_pages = list(get_articles_by_journal(journal.slug, at=at))
    legacy_articles = list(
        StaticArticle.objects.filter(
            journal=journal,
            canonical_page__isnull=True,
        ).order_by("sort_order", "title", "pk")
    )
    from .category_services import get_category_navigation

    return {
        "journal": journal,
        "journal_display_name": localized_journal_name(journal),
        "journal_display_seo_title": localized_journal_seo_title(journal),
        "journal_display_description": localized_journal_description(journal),
        "journal_display_intro": localized_journal_intro(journal),
        "articles": article_pages,
        "article_pages": article_pages,
        "legacy_articles": legacy_articles,
        "category_navigation": get_category_navigation(journal=journal),
        "seo": {
            "title": localized_journal_seo_title(journal),
            "description": localized_journal_description(journal),
        },
    }


@dataclass
class ImportIssue:
    row_no: int
    scope: str
    message: str
    payload: dict = field(default_factory=dict)
    error_code: str = "IMPORT_VALIDATION_ERROR"
    field_name: str = ""
    raw_value: str = ""
    error_type: str = "validation"
    suggestion: str = "请核对原始文件后修正"


@dataclass(frozen=True)
class ResolvedArticleImportCategories:
    primary: JournalCategory | None
    related: tuple[JournalCategory, ...]
    provided: bool
    report_payload: dict


class ArticleImportCategoryError(ValidationError):
    def __init__(self, code: str, message: str, *, report_payload: dict):
        self.code = code
        self.report_payload = report_payload
        super().__init__(message, code=code)


@dataclass
class PackageImportResult:
    package_name: str
    journal_job: JournalImportJob | None = None
    article_job: ArticleImportJob | None = None
    journal_created: int = 0
    journal_updated: int = 0
    article_created: int = 0
    article_updated: int = 0
    journal_issues: list[ImportIssue] = field(default_factory=list)
    article_issues: list[ImportIssue] = field(default_factory=list)


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return _to_text(value).lower() in {"1", "true", "yes", "y", "on"}


def _parse_int_value(value: object, *, default: int, field_name: str) -> int:
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError({field_name: "Enter a whole number."}) from exc


def _parse_datetime_value(value: object):
    if not value:
        return None
    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value
    parsed = parse_datetime(str(value))
    if parsed and timezone.is_naive(parsed):
        return timezone.make_aware(parsed)
    return parsed


def _extract_package(source_file) -> tuple[str, Path, Path, bytes, str]:
    source_name = getattr(source_file, "name", "import-package.zip")
    package_name = Path(source_name).stem
    package_bytes = source_file.read()

    tmpdir = Path(tempfile.mkdtemp(prefix="ai-author-forum-import-"))
    try:
        source_path = tmpdir / Path(source_name).name
        source_path.write_bytes(package_bytes)

        extract_root = tmpdir / "extract"
        extract_root.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(source_path) as zf:
            root = extract_root.resolve()
            for member in zf.infolist():
                target = (extract_root / member.filename).resolve()
                if target != root and root not in target.parents:
                    raise ValidationError(f"Unsafe package path: {member.filename}")
            zf.extractall(extract_root)
        return package_name, tmpdir, extract_root, package_bytes, source_name
    except Exception:
        shutil.rmtree(tmpdir, ignore_errors=True)
        raise


def _read_workbook(workbook_path: Path) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_normalize_header(value) for value in rows[0]]
    return [
        {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        for row in rows[1:]
    ], headers


def _read_csv(
    csv_path: Path, *, csv_encoding: str = "auto"
) -> tuple[list[dict], list[str]]:
    encoding = "gb18030" if csv_encoding == "gb18030" else "utf-8-sig"
    try:
        with csv_path.open("r", encoding=encoding, newline="") as source:
            reader = csv.DictReader(source)
            headers = [_normalize_header(value) for value in (reader.fieldnames or [])]
            rows = []
            for source_row in reader:
                normalized = {}
                for source_header, value in source_row.items():
                    header = _normalize_header(source_header)
                    if header:
                        normalized[header] = value
                rows.append(normalized)
            return rows, headers
    except UnicodeDecodeError as exc:
        if csv_encoding == "auto":
            raise ValidationError(
                "CSV 无法按 UTF-8/UTF-8-SIG 解码，请核对源文件；如确为国标编码，请显式选择 GB18030。"
            ) from exc
        raise ValidationError(
            "CSV 无法按 GB18030 解码，请核对文件编码和文件内容。"
        ) from exc


def _read_tabular(
    path: Path, *, csv_encoding: str = "auto"
) -> tuple[list[dict], list[str]]:
    if path.suffix.lower() == ".csv":
        return _read_csv(path, csv_encoding=csv_encoding)
    return _read_workbook(path)


def _annotate_suspicious_rows(rows: list[dict]) -> list[dict]:
    findings = []
    for row_no, source_row in enumerate(rows, start=2):
        issues = scan_mapping_for_suspicious_text(source_row)
        if not issues:
            continue
        source_row["suspicious_text_issues"] = issues
        for issue in issues:
            findings.append({"row_no": row_no, **issue})
    return findings


def _find_file(extract_root: Path, *candidates: str) -> Path | None:
    for candidate in candidates:
        candidate_path = extract_root / candidate
        if candidate_path.exists():
            return candidate_path
    return None


def _existing_image_by_title(title: str):
    return CustomImage.objects.filter(title=title).first()


def _create_or_reuse_image(package_name: str, relative_path: str, file_path: Path):
    title = f"{package_name}:{relative_path}"
    existing = _existing_image_by_title(title)
    if existing:
        return existing

    image_bytes = file_path.read_bytes()
    with PILImage.open(BytesIO(image_bytes)) as pil_image:
        width, height = pil_image.size

    image = CustomImage(
        title=title,
        file=ContentFile(image_bytes, name=file_path.name),
        width=width,
        height=height,
    )
    image.description = relative_path
    image.save()
    return image


def _load_assets(
    extract_root: Path, package_name: str, *, dry_run: bool = False
) -> dict[str, CustomImage | None]:
    media_root = extract_root / "media"
    if not media_root.exists():
        return {}

    asset_map: dict[str, CustomImage | None] = {}
    allowed_suffixes = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}
    for path in media_root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in allowed_suffixes:
            continue
        relative_path = path.relative_to(media_root).as_posix()
        image = (
            None
            if dry_run
            else _create_or_reuse_image(package_name, relative_path, path)
        )
        asset_map[relative_path] = image
        asset_map[path.name] = image
    return asset_map


def _resolve_image(reference: object, asset_map: dict[str, CustomImage | None]):
    if reference in (None, ""):
        return None
    text = _to_text(reference)
    if text.isdigit():
        return CustomImage.objects.filter(pk=int(text)).first()
    if text in asset_map:
        return asset_map[text]
    normalized = text.lstrip("/")
    if normalized in asset_map:
        return asset_map[normalized]
    if (
        normalized.startswith("media/")
        and normalized.removeprefix("media/") in asset_map
    ):
        return asset_map[normalized.removeprefix("media/")]
    return _existing_image_by_title(text)


def _asset_reference_exists(
    reference: object, asset_map: dict[str, CustomImage | None]
) -> bool:
    if reference in (None, ""):
        return True
    text = _to_text(reference)
    normalized = text.lstrip("/")
    package_keys = {text, normalized}
    if normalized.startswith("media/"):
        package_keys.add(normalized.removeprefix("media/"))
    return (
        any(key in asset_map for key in package_keys)
        or _resolve_image(reference, asset_map) is not None
    )


def _default_static_path(journal_slug: str, article_slug: str) -> str:
    del journal_slug  # Kept in the compatibility signature for existing callers.
    return f"/articles/{article_slug}/index.html"


def _save_html_source(
    article: StaticArticle,
    package_name: str,
    source_text: str,
    filename: str | None = None,
):
    if not source_text:
        return
    filename = filename or f"{article.slug}.html"
    storage_path = (
        f"journal-articles/html/{package_name}/{article.journal.slug}/{filename}"
    )
    default_storage.save(storage_path, ContentFile(source_text.encode("utf-8")))
    article.html_source.name = storage_path
    article.source_html_path = storage_path


def _report_value(value):
    if value is None:
        return ""
    if isinstance(value, list | tuple | set):
        return ";".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _article_error_payload(raw: dict, **updates) -> dict:
    payload = {
        "external_id": _to_text(raw.get("external_id") or raw.get("slug")),
        "journal_code": _to_text(raw.get("journal_code") or raw.get("journal_slug")),
        "input_primary_category_code": _to_text(raw.get("primary_category_code")),
        "input_primary_category_path": _to_text(raw.get("primary_category_path")),
        "resolved_primary_category_id": "",
        "input_related_categories": {
            "codes": _to_text(raw.get("related_category_codes")),
            "paths": _to_text(raw.get("related_category_paths")),
        },
        "resolved_related_category_ids": [],
    }
    payload.update(updates)
    return payload


def _build_error_report(issues: Iterable[ImportIssue]) -> bytes:
    buffer = StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=[
            "row_number",
            "external_id",
            "journal_code",
            "input_primary_category_code",
            "input_primary_category_path",
            "resolved_primary_category_id",
            "input_related_categories",
            "resolved_related_category_ids",
            "status",
            "error_code",
            "error_message",
            "field_name",
            "raw_value",
            "error_type",
            "suggestion",
            "scope",
            "payload",
        ],
    )
    writer.writeheader()
    for issue in issues:
        payload = issue.payload or {}
        writer.writerow(
            {
                "row_number": issue.row_no,
                "external_id": _report_value(payload.get("external_id")),
                "journal_code": _report_value(payload.get("journal_code")),
                "input_primary_category_code": _report_value(
                    payload.get("input_primary_category_code")
                ),
                "input_primary_category_path": _report_value(
                    payload.get("input_primary_category_path")
                ),
                "resolved_primary_category_id": _report_value(
                    payload.get("resolved_primary_category_id")
                ),
                "input_related_categories": _report_value(
                    payload.get("input_related_categories")
                ),
                "resolved_related_category_ids": _report_value(
                    payload.get("resolved_related_category_ids")
                ),
                "status": "failed",
                "error_code": issue.error_code,
                "error_message": issue.message,
                "field_name": issue.field_name,
                "raw_value": truncate_text(issue.raw_value),
                "error_type": issue.error_type,
                "suggestion": issue.suggestion,
                "scope": issue.scope,
                "payload": _report_value(payload),
            }
        )
    return buffer.getvalue().encode("utf-8-sig")


def _bind_asset(
    journal: Journal,
    asset_type: str,
    image: CustomImage,
    source_name: str,
    used_by: str,
):
    JournalAssetBinding.objects.update_or_create(
        journal=journal,
        asset_type=asset_type,
        source_name=source_name,
        defaults={
            "image": image,
            "used_by": used_by,
            "status": AssetBindingStatus.BOUND,
        },
    )


def import_journal_rows(
    job: JournalImportJob,
    rows: list[dict],
    *,
    asset_map: dict[str, CustomImage | None],
    dry_run: bool = False,
) -> tuple[int, int, list[ImportIssue]]:
    issues: list[ImportIssue] = []
    journal_rows = []

    for row_no, row in enumerate(rows, start=2):
        raw = dict(row)
        if not any(_to_text(value) for value in raw.values()):
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.SKIPPED,
                    action="skip",
                )
            )
            continue
        missing = REQUIRED_JOURNAL_COLUMNS - raw.keys()
        if missing:
            message = f"Missing columns: {', '.join(sorted(missing))}"
            issues.append(ImportIssue(row_no, "journal", message, raw))
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.FAILED,
                    error_message=message,
                )
            )
            continue

        slug = slugify(_to_text(raw.get("slug")), allow_unicode=True)
        name = _to_text(raw.get("journal_name"))
        if not slug or not name:
            message = "journal_name and slug are required"
            issues.append(ImportIssue(row_no, "journal", message, raw))
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.FAILED,
                    error_message=message,
                )
            )
            continue

        journal = Journal.objects.filter(slug=slug).first() or Journal(slug=slug)
        created = journal.pk is None
        journal.name = name
        journal.name_cn = _to_text(raw.get("journal_name_cn"))
        journal.az_group = _to_text(raw.get("az_group")).upper() or "#"
        journal.status = _to_text(raw.get("status")).lower() or journal.status
        try:
            journal.sort_order = _parse_int_value(
                raw.get("sort_order"),
                default=journal.sort_order or 0,
                field_name="sort_order",
            )
            journal.target_article_count = _parse_int_value(
                raw.get("target_article_count"),
                default=journal.target_article_count or 100,
                field_name="target_article_count",
            )
        except ValidationError as exc:
            message = str(exc)
            issues.append(ImportIssue(row_no, "journal", message, raw))
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.FAILED,
                    error_message=message,
                )
            )
            continue
        journal.seo_title = _to_text(raw.get("seo_title"))
        journal.seo_description = _to_text(raw.get("seo_description"))
        journal.homepage_intro = _to_text(raw.get("homepage_intro"))
        journal.static_site_path = _to_text(raw.get("static_site_path"))
        journal.notes = _to_text(raw.get("notes"))

        cover_ref = raw.get("cover_image")
        metrics_ref = raw.get("metrics_image")
        resolved_cover = _resolve_image(cover_ref, asset_map)
        resolved_metrics = _resolve_image(metrics_ref, asset_map)
        if resolved_cover is not None or not dry_run:
            journal.cover_image = resolved_cover
        if resolved_metrics is not None or not dry_run:
            journal.metrics_image = resolved_metrics
        if cover_ref and not _asset_reference_exists(cover_ref, asset_map):
            message = f"Missing cover_image: {cover_ref}"
            issues.append(ImportIssue(row_no, "journal", message, raw))
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.FAILED,
                    error_message=message,
                )
            )
            continue
        if metrics_ref and not _asset_reference_exists(metrics_ref, asset_map):
            message = f"Missing metrics_image: {metrics_ref}"
            issues.append(ImportIssue(row_no, "journal", message, raw))
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.FAILED,
                    error_message=message,
                )
            )
            continue

        if dry_run:
            try:
                journal.full_clean()
            except ValidationError as exc:
                message = str(exc)
                issues.append(ImportIssue(row_no, "journal", message, raw))
                journal_rows.append(
                    JournalImportRow(
                        job=job,
                        row_no=row_no,
                        raw_data=raw,
                        status=ImportRowStatus.FAILED,
                        error_message=message,
                    )
                )
                continue
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.SUCCESS,
                    action="create" if created else "update",
                )
            )
            continue

        try:
            journal.full_clean()
            journal.save()
            from ai_author_forum.site_settings.navigation import (
                ensure_navigation_for_journal,
            )

            ensure_navigation_for_journal(journal, actor=job.operator)
        except ValidationError as exc:
            message = str(exc)
            issues.append(ImportIssue(row_no, "journal", message, raw))
            journal_rows.append(
                JournalImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.FAILED,
                    error_message=message,
                )
            )
            continue
        if journal.cover_image and cover_ref:
            _bind_asset(
                journal,
                "cover",
                journal.cover_image,
                _to_text(cover_ref),
                "journal cover",
            )
        if journal.metrics_image and metrics_ref:
            _bind_asset(
                journal,
                "metrics",
                journal.metrics_image,
                _to_text(metrics_ref),
                "journal metrics",
            )
        journal_rows.append(
            JournalImportRow(
                job=job,
                row_no=row_no,
                raw_data=raw,
                status=ImportRowStatus.SUCCESS,
                action="create" if created else "update",
                journal=journal,
            )
        )

    if dry_run:
        JournalImportRow.objects.bulk_create(journal_rows)
        created_count = sum(
            1
            for row in journal_rows
            if row.status == ImportRowStatus.SUCCESS and row.action == "create"
        )
        updated_count = sum(
            1
            for row in journal_rows
            if row.status == ImportRowStatus.SUCCESS and row.action == "update"
        )
        return created_count, updated_count, issues

    JournalImportRow.objects.bulk_create(journal_rows)
    created_count = sum(
        1
        for row in journal_rows
        if row.status == ImportRowStatus.SUCCESS and row.action == "create"
    )
    updated_count = sum(
        1
        for row in journal_rows
        if row.status == ImportRowStatus.SUCCESS and row.action == "update"
    )
    return created_count, updated_count, issues


def _split_category_values(value: object) -> list[str]:
    text = _to_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(";") if item.strip()]


def _category_error_message(exc: ValidationError) -> str:
    messages = getattr(exc, "messages", None)
    return "; ".join(messages) if messages else str(exc)


def _resolve_article_import_categories(
    *, journal: Journal, raw: dict
) -> ResolvedArticleImportCategories:
    from ai_author_forum.journals.category_services import (
        CategoryError,
        resolve_category,
    )

    category_fields = {
        "primary_category_code",
        "primary_category_path",
        "related_category_codes",
        "related_category_paths",
    }
    provided = bool(category_fields.intersection(raw))
    report = _article_error_payload(raw)
    primary_codes = _split_category_values(raw.get("primary_category_code"))
    primary_paths = _split_category_values(raw.get("primary_category_path"))
    related_codes = _split_category_values(raw.get("related_category_codes"))
    related_paths = _split_category_values(raw.get("related_category_paths"))

    def fail(code: str, message: str):
        raise ArticleImportCategoryError(
            code,
            message,
            report_payload=dict(report),
        )

    if len(primary_codes) > 1 or len(primary_paths) > 1:
        fail(
            "ARTICLE_MULTIPLE_PRIMARY_CATEGORIES",
            "Only one primary category may be supplied.",
        )
    if related_codes and related_paths and len(related_codes) != len(related_paths):
        fail(
            "CATEGORY_RELATED_PAIR_COUNT_MISMATCH",
            "related_category_codes and related_category_paths must contain the same number of values.",
        )
    related_count = max(len(related_codes), len(related_paths))
    if related_count > 10:
        fail(
            "ARTICLE_TOO_MANY_RELATED_CATEGORIES",
            "At most 10 related categories are allowed.",
        )
    if not journal.pk and any(
        (primary_codes, primary_paths, related_codes, related_paths)
    ):
        fail(
            "CATEGORY_NOT_FOUND",
            "Categories cannot be resolved for a journal that has not been applied yet.",
        )

    def resolve_pair(code: str = "", path: str = ""):
        try:
            return resolve_category(
                journal=journal,
                code=code or None,
                full_path=path or None,
            )
        except CategoryError as exc:
            fail(exc.code, _category_error_message(exc))

    primary = None
    if primary_codes or primary_paths:
        resolved = resolve_pair(
            primary_codes[0] if primary_codes else "",
            primary_paths[0] if primary_paths else "",
        )
        primary = resolved.category
        report["resolved_primary_category_id"] = primary.pk

    related = []
    for index in range(related_count):
        resolved = resolve_pair(
            related_codes[index] if related_codes else "",
            related_paths[index] if related_paths else "",
        )
        related.append(resolved.category)
        report["resolved_related_category_ids"] = [item.pk for item in related]

    category_ids = ([primary.pk] if primary else []) + [item.pk for item in related]
    if len(category_ids) != len(set(category_ids)):
        fail(
            "ARTICLE_DUPLICATE_CATEGORY",
            "A category may appear only once and cannot be both primary and related.",
        )
    if related and primary is None:
        fail(
            "ARTICLE_PRIMARY_CATEGORY_REQUIRED",
            "A primary category is required when related categories are supplied.",
        )

    return ResolvedArticleImportCategories(
        primary=primary,
        related=tuple(related),
        provided=provided,
        report_payload=report,
    )


def _validate_static_article_category_selection(
    *, article: StaticArticle, categories: ResolvedArticleImportCategories
):
    from ai_author_forum.articles.category_services import (
        validate_article_category_revision,
    )
    from ai_author_forum.articles.models import ArticlePage

    payload = []
    if categories.primary:
        payload.append({"category_id": categories.primary.pk, "is_primary": True})
    payload.extend(
        {"category_id": category.pk, "is_primary": False}
        for category in categories.related
    )
    validate_article_category_revision(
        article=ArticlePage(primary_journal=article.journal),
        revision_content={"category_assignments": payload},
        action="submit" if payload else "draft",
    )


def import_article_rows(
    job: ArticleImportJob,
    rows: list[dict],
    *,
    package_name: str,
    extract_root: Path,
    asset_map: dict[str, CustomImage | None],
    dry_run: bool = False,
    available_journal_slugs: set[str] | None = None,
) -> tuple[int, int, list[ImportIssue]]:
    from ai_author_forum.articles.category_services import (
        assign_static_article_categories,
    )
    from ai_author_forum.articles.services import sync_imported_article

    issues: list[ImportIssue] = []
    article_rows = []
    created_count = 0
    updated_count = 0
    existing_articles = {
        (article.journal_id, article.slug): article
        for article in StaticArticle.objects.select_related("journal").all()
    }

    def append_failure(row_no, raw, message, *, error_code, payload=None):
        issue_payload = payload or _article_error_payload(raw)
        issues.append(
            ImportIssue(
                row_no,
                "article",
                message,
                issue_payload,
                error_code=error_code,
            )
        )
        article_rows.append(
            ArticleImportRow(
                job=job,
                row_no=row_no,
                raw_data=raw,
                status=ImportRowStatus.FAILED,
                error_message=message,
            )
        )

    available_journal_slugs = available_journal_slugs or set()
    for row_no, row in enumerate(rows, start=2):
        raw = dict(row)
        if not any(_to_text(value) for value in raw.values()):
            article_rows.append(
                ArticleImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.SKIPPED,
                    action="skip",
                )
            )
            continue
        missing = REQUIRED_ARTICLE_COLUMNS - raw.keys()
        if missing:
            message = f"Missing columns: {', '.join(sorted(missing))}"
            append_failure(
                row_no,
                raw,
                message,
                error_code="IMPORT_MISSING_COLUMNS",
            )
            continue

        journal_slug = slugify(_to_text(raw.get("journal_slug")), allow_unicode=True)
        journal = Journal.objects.filter(slug=journal_slug).first()
        preview_only_journal = (
            dry_run and journal is None and journal_slug in available_journal_slugs
        )
        if preview_only_journal:
            journal = Journal(slug=journal_slug, name=journal_slug, az_group="#")
        if not journal:
            message = f"Unknown journal_slug: {journal_slug}"
            append_failure(
                row_no,
                raw,
                message,
                error_code="JOURNAL_NOT_FOUND",
            )
            continue

        slug = slugify(_to_text(raw.get("slug")), allow_unicode=True)
        title = _to_text(raw.get("title"))
        existing = existing_articles.get((journal.pk, slug)) if journal.pk else None
        article = existing or StaticArticle(journal=journal, slug=slug)
        created = article.pk is None
        article.title = title
        article.article_type = (
            _to_text(raw.get("article_type")).lower() or ArticleType.AI_ARTICLE
        )
        article.authors = _to_text(raw.get("authors"))
        article.ai_co_authors = _to_text(raw.get("ai_co_authors"))
        article.abstract = _to_text(raw.get("abstract"))
        article.keywords = _to_text(raw.get("keywords"))
        article.publication_date = _parse_datetime_value(raw.get("publication_date"))
        review_status = _to_text(raw.get("status")).lower()
        if review_status and review_status not in ArticleReviewStatus.values:
            message = f"Invalid status: {review_status}"
            append_failure(
                row_no,
                raw,
                message,
                error_code="ARTICLE_STATUS_INVALID",
            )
            continue
        article.review_status = review_status or article.review_status
        try:
            article.sort_order = _parse_int_value(
                raw.get("sort_order"),
                default=article.sort_order or 0,
                field_name="sort_order",
            )
        except ValidationError as exc:
            message = _category_error_message(exc)
            append_failure(
                row_no,
                raw,
                message,
                error_code=getattr(exc, "code", "ARTICLE_SORT_ORDER_INVALID"),
            )
            continue
        article.is_pinned = _parse_bool(raw.get("is_pinned"))
        article.build_version = _to_text(raw.get("build_version"))
        article.notes = _to_text(raw.get("notes"))
        article.static_output_path = _default_static_path(journal.slug, slug)

        try:
            categories = _resolve_article_import_categories(journal=journal, raw=raw)
            if categories.provided:
                _validate_static_article_category_selection(
                    article=article,
                    categories=categories,
                )
        except (ArticleImportCategoryError, ValidationError) as exc:
            message = _category_error_message(exc)
            append_failure(
                row_no,
                raw,
                message,
                error_code=getattr(exc, "code", "ARTICLE_CATEGORY_INVALID"),
                payload=getattr(exc, "report_payload", _article_error_payload(raw)),
            )
            continue

        cover_ref = raw.get("cover_image")
        resolved_cover = _resolve_image(cover_ref, asset_map)
        if resolved_cover is not None or not dry_run:
            article.cover_image = resolved_cover
        if cover_ref and not _asset_reference_exists(cover_ref, asset_map):
            message = f"Missing cover_image: {cover_ref}"
            append_failure(
                row_no,
                raw,
                message,
                error_code="ARTICLE_COVER_IMAGE_NOT_FOUND",
                payload=categories.report_payload,
            )
            continue

        html_source_text = None
        html_source_name = None
        if html_text := _to_text(raw.get("body_html")):
            html_source_text = html_text
            html_source_name = f"{slug}.html"
        elif html_path := _to_text(raw.get("html_file") or raw.get("html_path")):
            candidate = extract_root / html_path.lstrip("/")
            if not candidate.exists():
                candidate = Path(html_path)
            if not candidate.exists():
                message = f"Missing html_file: {html_path}"
                append_failure(
                    row_no,
                    raw,
                    message,
                    error_code="ARTICLE_HTML_FILE_NOT_FOUND",
                    payload=categories.report_payload,
                )
                continue
            html_source_text = candidate.read_text(encoding="utf-8")
            html_source_name = candidate.name

        if dry_run:
            try:
                article.full_clean(
                    exclude={"journal"} if preview_only_journal else None
                )
            except ValidationError as exc:
                message = _category_error_message(exc)
                append_failure(
                    row_no,
                    raw,
                    message,
                    error_code=getattr(exc, "code", "ARTICLE_VALIDATION_ERROR"),
                    payload=categories.report_payload,
                )
                continue
            article_rows.append(
                ArticleImportRow(
                    job=job,
                    row_no=row_no,
                    raw_data=raw,
                    status=ImportRowStatus.SUCCESS,
                    action="create" if created else "update",
                )
            )
            continue

        try:
            with transaction.atomic():
                if html_source_text:
                    _save_html_source(
                        article,
                        package_name,
                        html_source_text,
                        filename=html_source_name,
                    )
                article.full_clean()
                article.save()
                if categories.provided:
                    assign_static_article_categories(
                        static_article=article,
                        primary=categories.primary,
                        related=categories.related,
                    )
                sync_imported_article(
                    article,
                    source_html=html_source_text or "",
                    owner=job.operator,
                )
        except Exception as exc:
            message = _category_error_message(exc)
            append_failure(
                row_no,
                raw,
                message,
                error_code=getattr(exc, "code", "ARTICLE_VALIDATION_ERROR"),
                payload=categories.report_payload,
            )
            if existing is not None:
                existing.refresh_from_db()
            continue

        existing_articles[(journal.pk, slug)] = article
        article_rows.append(
            ArticleImportRow(
                job=job,
                row_no=row_no,
                raw_data=raw,
                status=ImportRowStatus.SUCCESS,
                action="create" if created else "update",
                article=article,
            )
        )
        created_count += int(created)
        updated_count += int(not created)

    if dry_run:
        ArticleImportRow.objects.bulk_create(article_rows)
        created_count = sum(
            1
            for row in article_rows
            if row.status == ImportRowStatus.SUCCESS and row.action == "create"
        )
        updated_count = sum(
            1
            for row in article_rows
            if row.status == ImportRowStatus.SUCCESS and row.action == "update"
        )
        return created_count, updated_count, issues

    ArticleImportRow.objects.bulk_create(article_rows)
    return created_count, updated_count, issues


def _write_error_report(job, issues: list[ImportIssue]):
    if not issues:
        return
    filename = f"{Path(job.source_file.name).stem}-errors.csv"
    job.error_report.save(
        filename, ContentFile(_build_error_report(issues)), save=False
    )


def import_package(
    source_file,
    *,
    operator=None,
    dry_run: bool = False,
    allow_suspicious_text: bool = False,
    csv_encoding: str = "auto",
) -> PackageImportResult:
    source_name = getattr(source_file, "name", "import-package.zip")
    target_label = Path(source_name).name
    try:
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.STARTED,
            actor=operator,
            target_type="ImportPackage",
            target_label=target_label,
            message="Journal package import started",
            metadata={
                "dry_run": dry_run,
                "allow_suspicious_text": allow_suspicious_text,
                "csv_encoding": csv_encoding,
            },
        )
        result = _import_package(
            source_file,
            operator=operator,
            dry_run=dry_run,
            allow_suspicious_text=allow_suspicious_text,
            csv_encoding=csv_encoding,
        )
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.SUCCESS,
            actor=operator,
            target=result.journal_job or result.article_job,
            target_type="ImportPackage",
            target_label=target_label,
            message="Journal package import completed",
            metadata={
                "dry_run": dry_run,
                "journal_created": result.journal_created,
                "journal_updated": result.journal_updated,
                "article_created": result.article_created,
                "article_updated": result.article_updated,
                "journal_issues": len(result.journal_issues),
                "article_issues": len(result.article_issues),
                "allow_suspicious_text": allow_suspicious_text,
                "csv_encoding": csv_encoding,
            },
        )
        return result
    except Exception as exc:
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.FAILURE,
            actor=operator,
            target_type="ImportPackage",
            target_label=target_label,
            message="Journal package import failed",
            metadata={"dry_run": dry_run, "error": str(exc)},
        )
        raise


def _import_package(
    source_file,
    *,
    operator=None,
    dry_run: bool = False,
    allow_suspicious_text: bool = False,
    csv_encoding: str = "auto",
) -> PackageImportResult:
    tmpdir = None
    try:
        package_name, tmpdir, extract_root, package_bytes, source_name = (
            _extract_package(source_file)
        )
        source_name = Path(source_name).name
        asset_map = _load_assets(extract_root, package_name, dry_run=dry_run)
        result = PackageImportResult(package_name=package_name)
        available_journal_slugs = set(Journal.objects.values_list("slug", flat=True))

        journals_path = _find_file(
            extract_root, "journals.xlsx", "journal.xlsx", "journals.csv", "journal.csv"
        )
        if journals_path:
            rows, _headers = _read_tabular(journals_path, csv_encoding=csv_encoding)
            suspicious = _annotate_suspicious_rows(rows)
            if suspicious and not dry_run and not allow_suspicious_text:
                raise ValidationError(
                    "预览中发现可疑文本；必须由超级管理员确认按原文导入并填写理由。"
                )
            journal_job = JournalImportJob.objects.create(
                source_file=ContentFile(package_bytes, name=source_name),
                package_name=package_name,
                status=ImportJobStatus.VALIDATING,
                total_rows=len(rows),
                operator=operator,
            )
            result.journal_job = journal_job
            created, updated, issues = import_journal_rows(
                journal_job,
                rows,
                asset_map=asset_map,
                dry_run=dry_run,
            )
            skipped = journal_job.rows.filter(status=ImportRowStatus.SKIPPED).count()
            journal_job.success_rows = created + updated
            journal_job.failed_rows = len(issues)
            journal_job.status = (
                ImportJobStatus.READY if dry_run else ImportJobStatus.COMPLETED
            )
            journal_job.started_at = journal_job.started_at or timezone.now()
            journal_job.finished_at = timezone.now()
            journal_job.summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "failed": len(issues),
                "dry_run": dry_run,
                "csv_encoding": csv_encoding,
                "suspicious_text_count": len(suspicious),
                "suspicious_rows": suspicious,
            }
            suspicious_report = [
                ImportIssue(
                    row_no=item["row_no"],
                    scope="journals",
                    message=item["message"],
                    error_code=f"SUSPICIOUS_{item['rule'].upper()}",
                    field_name=item["field_name"],
                    raw_value=item["raw_value"],
                    error_type=item["severity"],
                    suggestion=item["suggestion"],
                )
                for item in suspicious
            ]
            _write_error_report(journal_job, [*issues, *suspicious_report])
            journal_job.save()
            result.journal_created = created
            result.journal_updated = updated
            result.journal_issues = issues
            available_journal_slugs.update(
                slugify(_to_text(row.raw_data.get("slug")), allow_unicode=True)
                for row in journal_job.rows.filter(status=ImportRowStatus.SUCCESS)
            )

        articles_path = _find_file(
            extract_root, "articles.xlsx", "article.xlsx", "articles.csv", "article.csv"
        )
        if articles_path:
            rows, _headers = _read_tabular(articles_path, csv_encoding=csv_encoding)
            suspicious = _annotate_suspicious_rows(rows)
            if suspicious and not dry_run and not allow_suspicious_text:
                raise ValidationError(
                    "预览中发现可疑文本；必须由超级管理员确认按原文导入并填写理由。"
                )
            article_job = ArticleImportJob.objects.create(
                source_file=ContentFile(package_bytes, name=source_name),
                package_name=package_name,
                status=ImportJobStatus.VALIDATING,
                total_rows=len(rows),
                operator=operator,
            )
            result.article_job = article_job
            created, updated, issues = import_article_rows(
                article_job,
                rows,
                package_name=package_name,
                extract_root=extract_root,
                asset_map=asset_map,
                dry_run=dry_run,
                available_journal_slugs=available_journal_slugs,
            )
            skipped = article_job.rows.filter(status=ImportRowStatus.SKIPPED).count()
            article_job.success_rows = created + updated
            article_job.failed_rows = len(issues)
            article_job.status = (
                ImportJobStatus.READY if dry_run else ImportJobStatus.COMPLETED
            )
            article_job.started_at = article_job.started_at or timezone.now()
            article_job.finished_at = timezone.now()
            article_job.summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "failed": len(issues),
                "dry_run": dry_run,
                "csv_encoding": csv_encoding,
                "suspicious_text_count": len(suspicious),
                "suspicious_rows": suspicious,
            }
            suspicious_report = [
                ImportIssue(
                    row_no=item["row_no"],
                    scope="articles",
                    message=item["message"],
                    error_code=f"SUSPICIOUS_{item['rule'].upper()}",
                    field_name=item["field_name"],
                    raw_value=item["raw_value"],
                    error_type=item["severity"],
                    suggestion=item["suggestion"],
                )
                for item in suspicious
            ]
            _write_error_report(article_job, [*issues, *suspicious_report])
            article_job.save()
            result.article_created = created
            result.article_updated = updated
            result.article_issues = issues

        if result.journal_job is None and result.article_job is None:
            raise ValidationError(
                "The package must contain journals.xlsx and/or articles.xlsx."
            )
        return result
    finally:
        if tmpdir is not None:
            shutil.rmtree(tmpdir, ignore_errors=True)


# Dynamic category service contract (journals.services.categories).
import sys as _sys  # noqa: E402

from . import category_services as categories  # noqa: E402

_sys.modules[__name__ + ".categories"] = categories
