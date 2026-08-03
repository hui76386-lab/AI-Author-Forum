from __future__ import annotations

import csv
import zipfile
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARTICLE_IMPORT_TEMPLATE_TYPE = "article_import"
ARTICLE_IMPORT_TEMPLATE_VERSION = 2

GLOBAL_COLUMNS = [
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "body_html",
    "html_file",
    "docx_file",
    "markdown_file",
    "cover_image",
    "primary_category_code",
    "primary_category_path",
    "related_category_codes",
    "related_category_paths",
    "notes",
]
JOURNAL_COLUMNS = [column for column in GLOBAL_COLUMNS if column != "journal_slug"]


def _example_row(*, scope: str, journal_slug: str = "") -> list[str]:
    values = {
        "journal_slug": journal_slug or "ai-ethics-forum",
        "title": "Responsible Co-authoring",
        "slug": "responsible-co-authoring",
        "article_type": "ai_article",
        "authors": "Example Author",
        "ai_co_authors": "Example AI",
        "abstract": "Article abstract.",
        "keywords": "AI ethics, authorship",
        "publication_date": "2026-07-28T09:00:00+08:00",
        "body_html": "<h2>Responsible Co-authoring</h2><p>Article body.</p>",
        "html_file": "",
        "docx_file": "",
        "markdown_file": "",
        "cover_image": "",
        "primary_category_code": "",
        "primary_category_path": "",
        "related_category_codes": "",
        "related_category_paths": "",
        "notes": "Imported articles always enter draft review status.",
    }
    columns = GLOBAL_COLUMNS if scope == "global" else JOURNAL_COLUMNS
    return [values[column] for column in columns]


def build_article_import_xlsx(*, scope: str, journal_slug: str = "") -> bytes:
    columns = GLOBAL_COLUMNS if scope == "global" else JOURNAL_COLUMNS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "articles"
    sheet.append(columns)
    sheet.append(_example_row(scope=scope, journal_slug=journal_slug))
    fill = PatternFill("solid", fgColor="1D5E8C")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(column) + 2, 16), 34
        )

    metadata = workbook.create_sheet("_meta")
    metadata.sheet_state = "hidden"
    metadata.append(["template_type", ARTICLE_IMPORT_TEMPLATE_TYPE])
    metadata.append(["template_version", ARTICLE_IMPORT_TEMPLATE_VERSION])
    metadata.append(["scope", scope])
    metadata.append(["journal_slug", journal_slug if scope == "journal" else ""])
    metadata.append(
        [
            "business_rule",
            "draft_only;no_review_approval;no_placement;no_static_publish",
        ]
    )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_article_import_csv(*, scope: str, journal_slug: str = "") -> bytes:
    columns = GLOBAL_COLUMNS if scope == "global" else JOURNAL_COLUMNS
    stream = StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(columns)
    writer.writerow(_example_row(scope=scope, journal_slug=journal_slug))
    return stream.getvalue().encode("utf-8-sig")


def build_article_document_import_zip(*, scope: str, journal_slug: str = "") -> bytes:
    """Build the documented ZIP example for batch Markdown/DOCX imports."""
    workbook = load_workbook(
        BytesIO(build_article_import_xlsx(scope=scope, journal_slug=journal_slug))
    )
    sheet = workbook["articles"]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    sheet.cell(2, headers["body_html"], "")
    sheet.cell(2, headers["html_file"], "")
    sheet.cell(2, headers["docx_file"], "")
    sheet.cell(2, headers["markdown_file"], "documents/example.md")
    workbook_stream = BytesIO()
    workbook.save(workbook_stream)

    example_markdown = """---
title: Responsible Co-authoring
slug: responsible-co-authoring
article_type: ai_article
authors:
  - Example Author
ai_co_authors:
  - Example AI
---

# Responsible Co-authoring

This Markdown document is converted during preview and imported only as a draft.

## Example section

- Review the converted content.
- Confirm the draft import manually.
"""
    readme = """AI Author Forum 文章文档批量导入示例包

1. 每行正文必须在 body_html、html_file、docx_file、markdown_file 中严格四选一。
2. DOCX/Markdown 导入后仍然只生成草稿，不会自动审核、投放或静态发布。
3. Markdown 本地图片必须放在本 ZIP 内，并使用相对于 Markdown 文件的安全路径引用。
4. DOCX 不支持宏、ActiveX、OLE/嵌入对象、外部模板或外部图片。
5. 上传包最大 50 MB；每个任务最多 200 个唯一文档；外层 ZIP 与 DOCX 内层包共享 250 MB 逻辑解压预算。

目录说明：
- articles.xlsx：版本 2 清单。
- documents/example.md：Markdown 示例。
- media/README.txt：图片目录规则。
"""
    media_readme = """把 Markdown 引用的本地图片放在此目录。
例如 documents/example.md 可使用 ../media/figure-1.png。
禁止绝对路径、file://、路径穿越到 ZIP 外部或远程图片 URL。
"""

    stream = BytesIO()
    with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("articles.xlsx", workbook_stream.getvalue())
        archive.writestr("documents/example.md", example_markdown.encode("utf-8"))
        archive.writestr("media/README.txt", media_readme.encode("utf-8"))
        archive.writestr("README.txt", readme.encode("utf-8"))
    return stream.getvalue()
