from __future__ import annotations

import csv
import hashlib
import io
import re
import stat
import tempfile
import warnings
import zipfile
from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path, PurePosixPath
from urllib.parse import urlsplit

from bs4 import BeautifulSoup, Comment
from django.conf import settings
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.core.validators import validate_slug
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import load_workbook
from PIL import Image as PILImage

from ai_author_forum.images.models import CustomImage
from ai_author_forum.journals.models import (
    ArticleImportJob,
    ArticleImportRow,
    ArticleImportScope,
    ArticleReviewStatus,
    ArticleType,
    ImportJobStatus,
    ImportRowStatus,
    Journal,
    JournalStatus,
    StaticArticle,
)
from ai_author_forum.journals.services import (
    _resolve_article_import_categories,
    _validate_static_article_category_selection,
)
from ai_author_forum.journals.validators import scan_mapping_for_suspicious_text
from ai_author_forum.site_settings.models import AuditAction, AuditStatus
from ai_author_forum.site_settings.services import record_audit_event

from .category_services import assign_static_article_categories
from .document_importers import (
    FORBIDDEN_METADATA_KEYS,
    DocumentImportError,
    ImportExtractionBudget,
    convert_article_document,
)
from .import_permissions import (
    can_import_articles,
    can_override_suspicious_article_text,
    can_view_article_import_job,
)
from .import_templates import (
    ARTICLE_IMPORT_TEMPLATE_TYPE,
    ARTICLE_IMPORT_TEMPLATE_VERSION,
)
from .models import ArticlePage
from .services import sync_imported_article

MAX_ROWS = 15_000
MAX_SOURCE_SIZE = 50 * 1024 * 1024
MAX_ZIP_FILES = 500
MAX_ZIP_MEMBER_SIZE = 25 * 1024 * 1024
MAX_ZIP_TOTAL_SIZE = 250 * 1024 * 1024
MAX_IMAGE_FILE_SIZE = 10 * 1024 * 1024
MAX_IMAGE_WIDTH = 12_000
MAX_IMAGE_HEIGHT = 12_000
MAX_IMAGE_PIXELS = 50_000_000
STATIC_ARTICLE_QUERY_BATCH_SIZE = 500
ALLOWED_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
ALLOWED_IMAGE_FORMATS = {"JPEG", "PNG", "GIF", "WEBP"}
DANGEROUS_TAGS = {
    "script",
    "style",
    "object",
    "embed",
    "iframe",
    "base",
    "form",
    "svg",
    "math",
}
DANGEROUS_SCHEMES = {"javascript", "vbscript", "data", "file"}
ALLOWED_HTML_TAGS = {
    "article",
    "section",
    "div",
    "span",
    "h1",
    "h2",
    "h3",
    "h4",
    "h5",
    "h6",
    "p",
    "br",
    "hr",
    "ul",
    "ol",
    "li",
    "blockquote",
    "pre",
    "code",
    "strong",
    "b",
    "em",
    "i",
    "u",
    "s",
    "sup",
    "sub",
    "a",
    "img",
    "figure",
    "figcaption",
    "table",
    "caption",
    "colgroup",
    "col",
    "thead",
    "tbody",
    "tfoot",
    "tr",
    "th",
    "td",
}
GLOBAL_HTML_ATTRIBUTES = {"class", "id", "title", "lang", "dir"}
HTML_ATTRIBUTES_BY_TAG = {
    "a": {"href", "target", "rel"},
    "img": {"src", "alt", "title", "width", "height", "loading"},
    "ol": {"start", "reversed", "type"},
    "li": {"value"},
    "col": {"span"},
    "colgroup": {"span"},
    "th": {"colspan", "rowspan", "scope", "headers"},
    "td": {"colspan", "rowspan", "headers"},
}
ALLOWED_LINK_SCHEMES = {"http", "https", "mailto"}
ERROR_FIELD_BY_CODE = {
    "ARTICLE_REQUIRED_FIELD_MISSING": "required_fields",
    "ARTICLE_JOURNAL_NOT_FOUND": "journal_slug",
    "ARTICLE_JOURNAL_SCOPE_MISMATCH": "journal_slug",
    "ARTICLE_SLUG_INVALID": "slug",
    "ARTICLE_TYPE_INVALID": "article_type",
    "ARTICLE_BODY_MISSING": "body_html/html_file",
    "ARTICLE_HTML_UNSAFE": "body_html/html_file",
    "ARTICLE_HTML_FILE_NOT_FOUND": "html_file",
    "ARTICLE_COVER_IMAGE_NOT_FOUND": "cover_image/body_html",
    "ARTICLE_CATEGORY_INVALID": "primary_category/related_categories",
    "ARTICLE_MULTIPLE_PRIMARY_CATEGORIES": "primary_category_code/primary_category_path",
    "CATEGORY_RELATED_PAIR_COUNT_MISMATCH": "related_category_codes/related_category_paths",
    "ARTICLE_DUPLICATE_CATEGORY": "related_categories",
    "ARTICLE_PRIMARY_CATEGORY_REQUIRED": "primary_category_code/primary_category_path",
    "ARTICLE_TOO_MANY_RELATED_CATEGORIES": "related_categories",
    "CATEGORY_CROSS_JOURNAL": "primary_category/related_categories",
    "CATEGORY_INACTIVE": "primary_category/related_categories",
    "CATEGORY_NOT_FOUND": "primary_category/related_categories",
    "ARTICLE_TEXT_SUSPICIOUS": "row",
    "ARTICLE_DUPLICATE_IN_FILE": "slug",
    "ARTICLE_PUBLICATION_DATE_INVALID": "publication_date",
    "ARTICLE_DOCUMENT_FORMAT_UNSUPPORTED": "source_file",
    "ARTICLE_DOCUMENT_MIME_MISMATCH": "source_file",
    "ARTICLE_DOCUMENT_SOURCE_CONFLICT": "body source",
    "ARTICLE_DOCUMENT_FILE_NOT_FOUND": "docx_file/markdown_file",
    "ARTICLE_DOCUMENT_METADATA_MISSING": "metadata",
    "ARTICLE_DOCUMENT_SLUG_REQUIRED": "slug",
    "ARTICLE_DOCUMENT_BODY_EMPTY": "body",
    "ARTICLE_DOCUMENT_CONVERSION_FAILED": "document",
    "ARTICLE_DOCUMENT_HTML_TOO_LARGE": "body",
    "ARTICLE_DOCX_INVALID_PACKAGE": "docx_file",
    "ARTICLE_DOCX_ENCRYPTED": "docx_file",
    "ARTICLE_DOCX_MACRO_UNSAFE": "docx_file",
    "ARTICLE_DOCX_EMBEDDED_OBJECT_UNSAFE": "docx_file",
    "ARTICLE_DOCX_EXTERNAL_RELATIONSHIP_UNSAFE": "docx_file",
    "ARTICLE_DOCX_LIMIT_EXCEEDED": "docx_file",
    "ARTICLE_MARKDOWN_ENCODING_INVALID": "markdown_file",
    "ARTICLE_MARKDOWN_FRONT_MATTER_INVALID": "front_matter",
    "ARTICLE_MARKDOWN_LOCAL_IMAGE_REQUIRES_ZIP": "body image",
    "ARTICLE_DOCUMENT_IMAGE_UNSAFE": "body image",
}
ERROR_SUGGESTION_BY_CODE = {
    "ARTICLE_REQUIRED_FIELD_MISSING": "补齐错误信息中列出的必填字段后重新上传。",
    "ARTICLE_JOURNAL_NOT_FOUND": "核对子期刊 slug，或在全局模式选择有效的默认子期刊。",
    "ARTICLE_JOURNAL_SCOPE_MISMATCH": "删除 journal_slug，或改为当前锁定子期刊的 slug。",
    "ARTICLE_SLUG_INVALID": "仅使用字母、数字、连字符或下划线。",
    "ARTICLE_TYPE_INVALID": "使用模板中受控的 article_type 值。",
    "ARTICLE_BODY_MISSING": "在 body_html 或 html_file 中提供且仅需提供一个正文来源。",
    "ARTICLE_HTML_UNSAFE": "移除脚本、事件属性、危险 URL、外部图片或不支持的 HTML 标签。",
    "ARTICLE_HTML_FILE_NOT_FOUND": "确认 HTML 文件位于 ZIP 内，路径为相对路径且使用 UTF-8。",
    "ARTICLE_COVER_IMAGE_NOT_FOUND": "确认图片存在于 ZIP 内或引用已有素材，并满足格式、大小和尺寸限制。",
    "ARTICLE_CATEGORY_INVALID": "使用当前主属子期刊中已启用的栏目编码或路径。",
    "ARTICLE_MULTIPLE_PRIMARY_CATEGORIES": "主栏目最多填写一个编码或一个路径。",
    "CATEGORY_RELATED_PAIR_COUNT_MISMATCH": "相关栏目编码与路径数量必须一致。",
    "ARTICLE_DUPLICATE_CATEGORY": "删除重复的相关栏目编码或路径。",
    "ARTICLE_PRIMARY_CATEGORY_REQUIRED": "提交审核前必须选择一个有效主栏目。",
    "ARTICLE_TOO_MANY_RELATED_CATEGORIES": "相关栏目最多填写 10 个。",
    "CATEGORY_CROSS_JOURNAL": "栏目必须属于文章主属子期刊。",
    "CATEGORY_INACTIVE": "改用当前子期刊中已启用的栏目。",
    "CATEGORY_NOT_FOUND": "核对栏目编码或完整路径。",
    "ARTICLE_TEXT_SUSPICIOUS": "核对原始文件编码和异常字符；全局管理员强制处理时必须填写理由。",
    "ARTICLE_DUPLICATE_IN_FILE": "同一文件内保留唯一的 journal_slug + slug 行。",
    "ARTICLE_PUBLICATION_DATE_INVALID": "使用 ISO 日期或日期时间格式。",
}
REQUIRED_COLUMNS = {"title", "slug", "article_type", "authors"}
KNOWN_COLUMNS = {
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "body_html",
    "html_file",
    "cover_image",
    "primary_category_code",
    "primary_category_path",
    "related_category_codes",
    "related_category_paths",
    "notes",
    "docx_file",
    "markdown_file",
}


@dataclass(frozen=True)
class ArticleImportContext:
    scope: str
    target_journal_id: int | None = None
    default_journal_id: int | None = None
    csv_encoding: str = "auto"
    document_defaults: dict | None = None


class ArticleImportValidationError(ValidationError):
    def __init__(self, message: str, *, code: str = "ARTICLE_IMPORT_INVALID"):
        super().__init__(message, code=code)
        self.code = code


@dataclass
class ParsedPackage:
    rows: list[dict]
    root: Path
    package_name: str
    template_version: int
    template_warning: str = ""
    source_format: str = ""
    parser_version: str = ""


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _message(exc: Exception) -> str:
    messages = getattr(exc, "messages", None)
    return "; ".join(messages) if messages else str(exc)


def _read_source(source_file) -> tuple[bytes, str]:
    name = Path(
        getattr(source_file, "name", "article-import.zip") or "article-import.zip"
    ).name
    if hasattr(source_file, "seek"):
        source_file.seek(0)
    data = source_file.read(MAX_SOURCE_SIZE + 1)
    if hasattr(source_file, "seek"):
        source_file.seek(0)
    if len(data) > MAX_SOURCE_SIZE:
        raise ArticleImportValidationError(
            "导入文件不能超过 50 MB。", code="ARTICLE_SOURCE_TOO_LARGE"
        )
    return data, name


def _job_source_bytes(job: ArticleImportJob) -> bytes:
    with job.source_file.open("rb") as source:
        return source.read()


def _active_journal_lookup() -> tuple[dict[int, Journal], dict[str, Journal]]:
    journals = list(Journal.objects.filter(status=JournalStatus.ACTIVE))
    return (
        {journal.pk: journal for journal in journals},
        {journal.slug: journal for journal in journals},
    )


def _candidate_article_keys(
    rows: list[dict],
    context: ArticleImportContext,
    journals_by_id: dict[int, Journal],
    journals_by_slug: dict[str, Journal],
) -> tuple[set[int], set[str]]:
    journal_ids: set[int] = set()
    slugs: set[str] = set()
    for raw in rows:
        slug = _text(raw.get("slug"))
        if not slug:
            continue
        if context.scope == ArticleImportScope.JOURNAL:
            journal = journals_by_id.get(context.target_journal_id)
        else:
            journal_slug = _text(raw.get("journal_slug"))
            journal = (
                journals_by_slug.get(journal_slug)
                if journal_slug
                else journals_by_id.get(context.default_journal_id)
            )
        if journal is not None:
            journal_ids.add(journal.pk)
            slugs.add(slug)
    return journal_ids, slugs


def _existing_static_article_map(
    rows: list[dict],
    context: ArticleImportContext,
    journals_by_id: dict[int, Journal],
    journals_by_slug: dict[str, Journal],
) -> dict[tuple[int, str], StaticArticle]:
    journal_ids, slugs = _candidate_article_keys(
        rows, context, journals_by_id, journals_by_slug
    )
    if not journal_ids or not slugs:
        return {}
    ordered_slugs = sorted(slugs)
    existing = {}
    for offset in range(0, len(ordered_slugs), STATIC_ARTICLE_QUERY_BATCH_SIZE):
        batch = ordered_slugs[offset : offset + STATIC_ARTICLE_QUERY_BATCH_SIZE]
        queryset = StaticArticle.objects.filter(
            journal_id__in=journal_ids, slug__in=batch
        ).select_related("journal")
        existing.update(
            {(article.journal_id, article.slug): article for article in queryset}
        )
    return existing


def _safe_zip_name(info: zipfile.ZipInfo) -> str:
    raw = info.filename.replace("\\", "/")
    path = PurePosixPath(raw)
    if (
        not raw
        or raw.startswith(("/", "\\"))
        or path.is_absolute()
        or ".." in path.parts
    ):
        raise ArticleImportValidationError(
            f"ZIP 包含不安全路径：{info.filename}", code="ARTICLE_ZIP_PATH_UNSAFE"
        )
    if re.match(r"^[A-Za-z]:", raw):
        raise ArticleImportValidationError(
            f"ZIP 包含盘符路径：{info.filename}", code="ARTICLE_ZIP_PATH_UNSAFE"
        )
    mode = info.external_attr >> 16
    if mode and stat.S_ISLNK(mode):
        raise ArticleImportValidationError(
            f"ZIP 不允许符号链接：{info.filename}", code="ARTICLE_ZIP_SYMLINK"
        )
    return path.as_posix()


def _extract_zip(
    data: bytes, root: Path, budget: ImportExtractionBudget | None = None
) -> Path:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ArticleImportValidationError("ZIP 文件损坏或格式不正确。") from exc
    infos = archive.infolist()
    if len(infos) > MAX_ZIP_FILES:
        raise ArticleImportValidationError(
            "ZIP 文件数量超过 500 个。", code="ARTICLE_ZIP_TOO_MANY_FILES"
        )
    total = 0
    seen_names: set[str] = set()
    for info in infos:
        name = _safe_zip_name(info)
        if name in seen_names:
            raise ArticleImportValidationError(
                "ZIP 包含重复文件路径。", code="ARTICLE_ZIP_PATH_UNSAFE"
            )
        seen_names.add(name)
        if info.is_dir():
            continue
        if info.file_size > MAX_ZIP_MEMBER_SIZE:
            raise ArticleImportValidationError(
                f"ZIP 单个文件超过 25 MB：{name}", code="ARTICLE_ZIP_MEMBER_TOO_LARGE"
            )
        total += info.file_size
        if budget is not None:
            try:
                budget.consume(bytes_=info.file_size)
            except DocumentImportError as exc:
                raise ArticleImportValidationError(
                    str(exc), code="ARTICLE_ZIP_TOO_LARGE"
                ) from exc
        if total > MAX_ZIP_TOTAL_SIZE:
            raise ArticleImportValidationError(
                "ZIP 解压总大小超过 250 MB。", code="ARTICLE_ZIP_TOO_LARGE"
            )
        target = (root / name).resolve()
        try:
            target.relative_to(root.resolve())
        except ValueError as exc:
            raise ArticleImportValidationError(
                "ZIP 路径穿越被阻止。", code="ARTICLE_ZIP_PATH_UNSAFE"
            ) from exc
        target.parent.mkdir(parents=True, exist_ok=True)
        with archive.open(info) as source, target.open("wb") as output:
            output.write(source.read())
    archive.close()
    return root


def _decode_csv(data: bytes, encoding: str) -> str:
    if encoding == "gb18030":
        return data.decode("gb18030")
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ArticleImportValidationError(
            "CSV 不是 UTF-8/UTF-8-SIG；如源文件为中文旧编码，请显式选择 GB18030。",
            code="ARTICLE_CSV_ENCODING_INVALID",
        ) from exc


def _read_csv(path: Path, encoding: str) -> list[dict]:
    text = _decode_csv(path.read_bytes(), encoding)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ArticleImportValidationError("CSV 缺少表头。")
    rows = [{str(k or "").strip(): v for k, v in row.items()} for row in reader]
    if len(rows) > MAX_ROWS:
        raise ArticleImportValidationError(
            "文章行数超过 15,000 行上限。", code="ARTICLE_ROW_LIMIT"
        )
    return rows


def _xlsx_metadata(workbook) -> tuple[int, dict[str, str]]:
    if "_meta" not in workbook.sheetnames:
        return ARTICLE_IMPORT_TEMPLATE_VERSION, {}
    meta = {}
    for key, value, *_ in workbook["_meta"].iter_rows(values_only=True):
        if key:
            meta[_text(key)] = _text(value)
    if (
        meta.get("template_type")
        and meta["template_type"] != ARTICLE_IMPORT_TEMPLATE_TYPE
    ):
        raise ArticleImportValidationError(
            "Excel 模板类型不是 article_import。", code="ARTICLE_TEMPLATE_TYPE_INVALID"
        )
    try:
        version = int(meta.get("template_version") or ARTICLE_IMPORT_TEMPLATE_VERSION)
    except ValueError as exc:
        raise ArticleImportValidationError(
            "Excel 模板版本无效。", code="ARTICLE_TEMPLATE_VERSION_INVALID"
        ) from exc
    if version < 0:
        raise ArticleImportValidationError(
            "Excel 模板版本无效。", code="ARTICLE_TEMPLATE_VERSION_INVALID"
        )
    if version > ARTICLE_IMPORT_TEMPLATE_VERSION:
        raise ArticleImportValidationError(
            "模板版本高于当前系统支持版本，请下载当前模板。",
            code="ARTICLE_TEMPLATE_VERSION_UNSUPPORTED",
        )
    return version, meta


def _read_xlsx(path: Path, context: ArticleImportContext) -> tuple[list[dict], int]:
    try:
        workbook = load_workbook(
            path, read_only=False, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise ArticleImportValidationError("Excel 文件无法读取。") from exc
    version, meta = _xlsx_metadata(workbook)
    if context.scope == ArticleImportScope.JOURNAL and meta.get("scope") == "journal":
        target = Journal.objects.filter(pk=context.target_journal_id).first()
        if meta.get("journal_slug") and target and meta["journal_slug"] != target.slug:
            raise ArticleImportValidationError(
                "模板锁定的子期刊与当前导入目标不一致。",
                code="ARTICLE_JOURNAL_SCOPE_MISMATCH",
            )
    sheet = (
        workbook["articles"] if "articles" in workbook.sheetnames else workbook.active
    )
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                raise ArticleImportValidationError(
                    f"Excel 单元格 {cell.coordinate} 含公式，文章导入不允许公式。",
                    code="ARTICLE_XLSX_FORMULA",
                )
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        raise ArticleImportValidationError("Excel 缺少表头。")
    headers = [_text(value) for value in values[0]]
    rows = [dict(zip(headers, row, strict=False)) for row in values[1:]]
    if len(rows) > MAX_ROWS:
        raise ArticleImportValidationError(
            "文章行数超过 15,000 行上限。", code="ARTICLE_ROW_LIMIT"
        )
    return rows, version


def _find_article_table(root: Path) -> Path:
    forbidden = [
        p
        for p in root.rglob("*")
        if p.is_file() and p.name.lower() in {"journals.xlsx", "journals.csv"}
    ]
    if forbidden:
        raise ArticleImportValidationError(
            "文章专用导入包不能包含 journals.xlsx/csv。",
            code="ARTICLE_PACKAGE_CONTAINS_JOURNALS",
        )
    candidates = [
        p
        for p in root.rglob("*")
        if p.is_file() and p.name.lower() in {"articles.xlsx", "articles.csv"}
    ]
    if len(candidates) != 1:
        raise ArticleImportValidationError(
            "ZIP 必须且只能包含一个 articles.xlsx 或 articles.csv。"
        )
    return candidates[0]


def _load_package(
    data: bytes,
    name: str,
    context: ArticleImportContext,
    root: Path,
    budget: ImportExtractionBudget | None = None,
) -> ParsedPackage:
    suffix = Path(name).suffix.lower()
    version = ARTICLE_IMPORT_TEMPLATE_VERSION
    budget = budget or ImportExtractionBudget(
        max_total_bytes=getattr(
            settings, "ARTICLE_IMPORT_MAX_LOGICAL_EXTRACTED_SIZE", MAX_ZIP_TOTAL_SIZE
        ),
        max_nested_members=getattr(
            settings, "ARTICLE_IMPORT_MAX_NESTED_MEMBERS_PER_JOB", 20_000
        ),
    )
    source_format = suffix.lstrip(".")
    parser_version = "article-import-v2"
    documents_enabled = getattr(settings, "ARTICLE_DOCUMENT_IMPORT_ENABLED", True)
    if suffix == ".zip":
        _extract_zip(data, root, budget)
        table = _find_article_table(root)
        if table.suffix.lower() == ".xlsx":
            rows, version = _read_xlsx(table, context)
        else:
            rows = _read_csv(table, context.csv_encoding)
        document_references = {
            _text(row.get("docx_file")) or _text(row.get("markdown_file"))
            for row in rows
            if _text(row.get("docx_file")) or _text(row.get("markdown_file"))
        }
        if document_references and not documents_enabled:
            raise ArticleImportValidationError(
                "DOCX/Markdown 文档导入功能当前已关闭。",
                code="ARTICLE_DOCUMENT_FORMAT_UNSUPPORTED",
            )
        if len(document_references) > getattr(
            settings, "ARTICLE_IMPORT_MAX_DOCUMENTS_PER_JOB", 200
        ):
            raise ArticleImportValidationError(
                "ZIP 中引用的 DOCX/Markdown 文档数量超过 200 个。",
                code="ARTICLE_DOCX_LIMIT_EXCEEDED",
            )
    elif suffix in {".xlsx", ".csv"}:
        table = root / f"articles{suffix}"
        table.write_bytes(data)
        if suffix == ".xlsx":
            rows, version = _read_xlsx(table, context)
        else:
            rows = _read_csv(table, context.csv_encoding)
    elif suffix in {".docx", ".md", ".markdown"}:
        if not documents_enabled:
            raise ArticleImportValidationError(
                "DOCX/Markdown 文档导入功能当前已关闭。",
                code="ARTICLE_DOCUMENT_FORMAT_UNSUPPORTED",
            )
        safe_name = Path(name).name
        document_path = root / safe_name
        document_path.write_bytes(data)
        rows = [
            {
                "docx_file": safe_name if suffix == ".docx" else "",
                "markdown_file": safe_name if suffix in {".md", ".markdown"} else "",
                "_direct_document": True,
                **(context.document_defaults or {}),
            }
        ]
        version = 0
        source_format = "docx" if suffix == ".docx" else "markdown"
        parser_version = "document-import-v1"
    else:
        raise ArticleImportValidationError(
            "仅支持 XLSX、CSV、ZIP、DOCX、MD 或 Markdown 文件。",
            code="ARTICLE_DOCUMENT_FORMAT_UNSUPPORTED",
        )
    warning = ""
    if version < ARTICLE_IMPORT_TEMPLATE_VERSION and suffix in {
        ".xlsx",
        ".csv",
        ".zip",
    }:
        warning = (
            f"当前模板版本为 {version}，低于最新版本。"
            "仍按兼容规则读取，建议下载最新模板。"
        )
    return ParsedPackage(
        rows=rows,
        root=root,
        package_name=Path(name).stem,
        template_version=version,
        template_warning=warning,
        source_format=source_format,
        parser_version=parser_version,
    )


def _safe_package_file(root: Path, reference: str, *, error_code: str) -> Path:
    raw = reference.replace("\\", "/")
    path = PurePosixPath(raw)
    if (
        not raw
        or raw.startswith("/")
        or path.is_absolute()
        or ".." in path.parts
        or re.match(r"^[A-Za-z]:", raw)
    ):
        raise ArticleImportValidationError(
            f"文件路径不安全：{reference}", code=error_code
        )
    candidate = (root / path.as_posix()).resolve()
    try:
        candidate.relative_to(root.resolve())
    except ValueError as exc:
        raise ArticleImportValidationError(
            f"文件路径不安全：{reference}", code=error_code
        ) from exc
    if not candidate.is_file():
        raise ArticleImportValidationError(f"文件不存在：{reference}", code=error_code)
    return candidate


def _normalized_url(value) -> tuple[str, str]:
    raw = " ".join(value) if isinstance(value, list) else str(value or "")
    raw = raw.strip()
    compact = re.sub(r"[\x00-\x20\x7f]+", "", raw)
    return raw, urlsplit(compact).scheme.lower()


def _validate_html(
    html: str, *, root: Path, preview: bool, image_cache: dict[str, CustomImage]
) -> tuple[str, list[CustomImage]]:
    soup = BeautifulSoup(html, "html.parser")
    images = []

    for tag in soup.find_all(True):
        name = tag.name.lower()
        if name in DANGEROUS_TAGS:
            raise ArticleImportValidationError(
                f"HTML 包含危险标签 <{name}>。", code="ARTICLE_HTML_UNSAFE"
            )
        if name == "meta" and _text(tag.get("http-equiv")).lower() == "refresh":
            raise ArticleImportValidationError(
                "HTML 包含 meta refresh 跳转。", code="ARTICLE_HTML_UNSAFE"
            )
        for attr, value in list(tag.attrs.items()):
            lower = attr.lower()
            if lower.startswith("on") or lower in {"srcdoc", "style", "xlink:href"}:
                raise ArticleImportValidationError(
                    f"HTML 包含危险属性 {attr}。", code="ARTICLE_HTML_UNSAFE"
                )
            if lower in {"href", "src", "action", "formaction"}:
                raw, scheme = _normalized_url(value)
                if scheme in DANGEROUS_SCHEMES or raw.startswith("//"):
                    raise ArticleImportValidationError(
                        "HTML 包含危险 URL scheme。", code="ARTICLE_HTML_UNSAFE"
                    )

    if soup.head:
        soup.head.decompose()
    for comment in soup.find_all(string=lambda value: isinstance(value, Comment)):
        comment.extract()

    effective_root = soup.body or soup
    for tag in effective_root.find_all(True):
        name = tag.name.lower()
        if name not in ALLOWED_HTML_TAGS:
            raise ArticleImportValidationError(
                f"HTML 包含不支持的标签 <{name}>。", code="ARTICLE_HTML_UNSAFE"
            )
        allowed_attributes = GLOBAL_HTML_ATTRIBUTES | HTML_ATTRIBUTES_BY_TAG.get(
            name, set()
        )
        for attr in list(tag.attrs):
            if attr.lower() not in allowed_attributes:
                del tag.attrs[attr]

        if name == "a":
            href, scheme = _normalized_url(tag.get("href"))
            if scheme and scheme not in ALLOWED_LINK_SCHEMES:
                raise ArticleImportValidationError(
                    "HTML 链接使用了不允许的 URL scheme。",
                    code="ARTICLE_HTML_UNSAFE",
                )
            if href.startswith("//"):
                raise ArticleImportValidationError(
                    "HTML 链接必须显式使用 HTTP 或 HTTPS。",
                    code="ARTICLE_HTML_UNSAFE",
                )
            if tag.get("target") == "_blank":
                rel = set(tag.get("rel") or [])
                rel.update({"noopener", "noreferrer"})
                tag["rel"] = sorted(rel)

        if name == "img":
            src = _text(tag.get("src"))
            parsed = urlsplit(re.sub(r"[\x00-\x20\x7f]+", "", src))
            if not src or parsed.scheme or parsed.netloc or src.startswith("//"):
                raise ArticleImportValidationError(
                    "HTML 图片必须引用导入包内素材。",
                    code="ARTICLE_HTML_UNSAFE",
                )
            image_path = _safe_package_file(
                root, src, error_code="ARTICLE_COVER_IMAGE_NOT_FOUND"
            )
            image = _materialize_image(image_path, preview=preview, cache=image_cache)
            if image is not None:
                tag["src"] = image.file.url
                images.append(image)

    if soup.body:
        clean_html = "".join(str(child) for child in soup.body.contents)
    else:
        clean_html = str(soup)
    return clean_html, images


def _materialize_image(
    path: Path, *, preview: bool, cache: dict[str, CustomImage]
) -> CustomImage | None:
    if path.suffix.lower() not in ALLOWED_IMAGE_SUFFIXES:
        raise ArticleImportValidationError(
            f"图片格式不支持：{path.name}", code="ARTICLE_COVER_IMAGE_NOT_FOUND"
        )
    if path.stat().st_size > MAX_IMAGE_FILE_SIZE:
        raise ArticleImportValidationError(
            f"图片超过 10 MB：{path.name}", code="ARTICLE_COVER_IMAGE_NOT_FOUND"
        )
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", PILImage.DecompressionBombWarning)
            with PILImage.open(path) as image:
                width, height = image.size
                image_format = (image.format or "").upper()
                if image_format not in ALLOWED_IMAGE_FORMATS:
                    raise ValueError("unsupported image format")
                if width <= 0 or height <= 0:
                    raise ValueError("invalid image dimensions")
                if width > MAX_IMAGE_WIDTH or height > MAX_IMAGE_HEIGHT:
                    raise ValueError("image dimensions exceed limit")
                if width * height > MAX_IMAGE_PIXELS:
                    raise ValueError("image pixel count exceeds limit")
                image.verify()
    except Exception as exc:
        raise ArticleImportValidationError(
            f"图片文件无效或尺寸超限：{path.name}",
            code="ARTICLE_COVER_IMAGE_NOT_FOUND",
        ) from exc
    raw = path.read_bytes()
    key = hashlib.sha256(raw).hexdigest()
    if preview:
        return None
    if key in cache:
        return cache[key]
    image = CustomImage(
        title=path.stem[:255],
        file=ContentFile(raw, name=path.name),
    )
    image.save()
    cache[key] = image
    return image


def _resolve_existing_image(reference: str) -> CustomImage | None:
    if not reference:
        return None
    if reference.isdigit():
        return CustomImage.objects.filter(pk=int(reference)).first()
    return CustomImage.objects.filter(title=reference).order_by("pk").first()


def _parse_publication_date(value):
    text = _text(value)
    if not text:
        return None
    parsed = parse_datetime(text)
    if parsed:
        return parsed
    date_value = parse_date(text)
    if date_value:
        return date_value
    raise ArticleImportValidationError(
        "publication_date 格式无效。", code="ARTICLE_PUBLICATION_DATE_INVALID"
    )


def _normalized_document_text(value) -> str:
    if isinstance(value, list):
        return ", ".join(_text(item) for item in value if _text(item))
    return _text(value)


def _safe_infer_slug(value: str) -> str:
    text = _text(value)
    if not text or not text.isascii():
        return ""
    slug = re.sub(r"\s+", "-", text.lower())
    slug = re.sub(r"[^a-z0-9_-]+", "-", slug)
    return re.sub(r"-+", "-", slug).strip("-")


def _apply_document_title_rule(html: str, title: str, warnings: list[dict]) -> str:
    soup = BeautifulSoup(html, "html.parser")
    first_h1 = soup.find("h1")
    if first_h1 is not None:
        normalized_h1 = " ".join(first_h1.get_text(" ", strip=True).split()).casefold()
        normalized_title = " ".join(_text(title).split()).casefold()
        if normalized_h1 == normalized_title:
            first_h1.decompose()
            warnings.append(
                {
                    "code": "ARTICLE_DOCUMENT_DUPLICATE_TITLE_REMOVED",
                    "message": "正文首个 H1 与文章标题重复，已从正文中移除。",
                }
            )
        else:
            warnings.append(
                {
                    "code": "ARTICLE_DOCUMENT_HEADING_TITLE_MISMATCH",
                    "message": "正文首个 H1 与文章标题不一致，已保留并生成警告。",
                }
            )
    return str(soup.body.decode_contents() if soup.body else soup)


def _normalized_row(
    raw: dict,
    context: ArticleImportContext,
    root: Path,
    *,
    preview: bool,
    seen: set[tuple[int, str]],
    image_cache: dict[str, CustomImage],
    allow_suspicious_text: bool,
    journals_by_id: dict[int, Journal],
    journals_by_slug: dict[str, Journal],
    document_cache: dict[str, object] | None = None,
    extraction_budget: ImportExtractionBudget | None = None,
) -> dict:
    row = {
        str(key or "").strip(): value for key, value in raw.items() if key is not None
    }
    direct_document = bool(row.pop("_direct_document", False))
    document_cache = document_cache if document_cache is not None else {}
    extraction_budget = extraction_budget or ImportExtractionBudget(
        max_total_bytes=getattr(
            settings, "ARTICLE_IMPORT_MAX_LOGICAL_EXTRACTED_SIZE", MAX_ZIP_TOTAL_SIZE
        ),
        max_nested_members=getattr(
            settings, "ARTICLE_IMPORT_MAX_NESTED_MEMBERS_PER_JOB", 20_000
        ),
    )
    warnings_payload: list[dict] = []
    forbidden_keys = {
        "status",
        "approved",
        "published",
        "main_site_slot",
        "journal_slot",
        "is_pinned",
        "static_output_path",
        "build_version",
        "placement",
    }.intersection(row)
    for key in forbidden_keys:
        row.pop(key, None)
        warnings_payload.append(
            {
                "code": "ARTICLE_DOCUMENT_FORBIDDEN_METADATA_IGNORED",
                "message": f"已忽略禁止元数据字段：{key}",
            }
        )

    source_fields = ["body_html", "html_file", "docx_file", "markdown_file"]
    provided_sources = [field for field in source_fields if _text(row.get(field))]
    if len(provided_sources) != 1:
        raise ArticleImportValidationError(
            "正文必须在 body_html、html_file、docx_file、markdown_file 中严格四选一。",
            code=(
                "ARTICLE_DOCUMENT_SOURCE_CONFLICT"
                if provided_sources
                else "ARTICLE_BODY_MISSING"
            ),
        )

    source_format = "html"
    source_path = ""
    source_sha256 = ""
    converter_name = ""
    converter_version = ""
    conversion_warnings: list[dict] = []
    conversion_statistics: dict = {}
    document_metadata: dict = {}
    body_html = ""
    if provided_sources[0] == "body_html":
        body_html = _text(row.get("body_html"))
        source_sha256 = hashlib.sha256(body_html.encode("utf-8")).hexdigest()
    elif provided_sources[0] == "html_file":
        source_path = _text(row.get("html_file"))
        source = _safe_package_file(
            root, source_path, error_code="ARTICLE_HTML_FILE_NOT_FOUND"
        )
        try:
            body_html = source.read_text(encoding="utf-8")
        except UnicodeDecodeError as exc:
            raise ArticleImportValidationError(
                "HTML 文件必须使用 UTF-8 编码。", code="ARTICLE_HTML_FILE_NOT_FOUND"
            ) from exc
        source_format = "html"
        source_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
    else:
        source_path = _text(row.get(provided_sources[0]))
        source = _safe_package_file(
            root,
            source_path,
            error_code="ARTICLE_DOCUMENT_FILE_NOT_FOUND",
        )
        cache_key = (
            f"{source.resolve()}:{hashlib.sha256(source.read_bytes()).hexdigest()}"
        )
        converted = document_cache.get(cache_key)
        if converted is None:
            generated_root = root / "_converted_assets"
            converted = convert_article_document(
                source,
                package_root=root,
                generated_root=generated_root,
                budget=extraction_budget,
                direct_upload=direct_document,
                source_bytes_counted=not direct_document,
            )
            document_cache[cache_key] = converted
        source_format = converted.source_format
        source_sha256 = converted.source_sha256
        converter_name = converted.converter_name
        converter_version = converted.converter_version
        body_html = converted.html
        document_metadata = dict(converted.metadata or {})
        conversion_warnings.extend(
            {
                "code": warning.code,
                "message": warning.message,
                "source_path": warning.source_path,
                "element": warning.element,
            }
            for warning in converted.warnings
        )
        conversion_statistics = dict(converted.statistics or {})
        warnings_payload.extend(conversion_warnings)

    defaults = dict(context.document_defaults or {})
    merged = dict(row)
    metadata_sources: dict[str, str] = {}
    for key, value in document_metadata.items():
        if direct_document and key == "journal_slug":
            if _normalized_document_text(value):
                warnings_payload.append(
                    {
                        "code": "ARTICLE_DOCUMENT_METADATA_OVERRIDDEN",
                        "message": "直接上传文档中的 journal_slug 已忽略，目标子期刊由服务端上下文决定。",
                    }
                )
            continue
        if key in FORBIDDEN_METADATA_KEYS:
            warnings_payload.append(
                {
                    "code": "ARTICLE_DOCUMENT_FORBIDDEN_METADATA_IGNORED",
                    "message": f"已忽略禁止元数据字段：{key}",
                }
            )
            continue
        normalized_value = _normalized_document_text(value)
        if normalized_value and not _text(merged.get(key)):
            merged[key] = normalized_value
            metadata_sources[key] = "document"
        elif (
            normalized_value
            and _text(merged.get(key))
            and _text(merged.get(key)) != normalized_value
        ):
            warnings_payload.append(
                {
                    "code": "ARTICLE_DOCUMENT_METADATA_OVERRIDDEN",
                    "message": f"清单值覆盖了文档元数据字段：{key}",
                }
            )

    for key, value in defaults.items():
        if _text(value):
            if _text(merged.get(key)) and _text(merged.get(key)) != _text(value):
                warnings_payload.append(
                    {
                        "code": "ARTICLE_DOCUMENT_METADATA_OVERRIDDEN",
                        "message": f"上传表单值覆盖了文档或清单字段：{key}",
                    }
                )
            merged[key] = value
            metadata_sources[key] = "form"

    is_document_source = provided_sources[0] in {"docx_file", "markdown_file"}
    title = _text(merged.get("title"))
    if is_document_source and not title:
        soup = BeautifulSoup(body_html, "html.parser")
        heading = soup.find("h1")
        title = _text(heading.get_text(" ", strip=True) if heading else "")
        if not title:
            title = _text(soup.get_text(" ", strip=True))[:255]
        if not title:
            title = Path(source_path or "article").stem
        warnings_payload.append(
            {
                "code": "ARTICLE_DOCUMENT_TITLE_INFERRED",
                "message": "标题由文档内容或文件名推断。",
            }
        )
    title = title[:255]

    slug = _text(merged.get("slug"))
    slug_source = "explicit"
    if is_document_source and not slug:
        slug = _safe_infer_slug(document_metadata.get("slug", ""))
        slug_source = "document"
    if is_document_source and not slug:
        slug = _safe_infer_slug(Path(source_path).stem)
        slug_source = "filename"
    if is_document_source and not slug:
        slug = _safe_infer_slug(title)
        slug_source = "title"
    if is_document_source and not slug:
        raise ArticleImportValidationError(
            "无法安全推断 slug，请显式填写。", code="ARTICLE_DOCUMENT_SLUG_REQUIRED"
        )
    if is_document_source and slug_source != "explicit":
        warnings_payload.append(
            {
                "code": "ARTICLE_DOCUMENT_SLUG_INFERRED",
                "message": f"slug 由{slug_source}自动推断。",
            }
        )

    merged["title"] = title
    merged["slug"] = slug
    title_rule_warnings: list[dict] = []
    clean_input = (
        _apply_document_title_rule(body_html, title, title_rule_warnings)
        if is_document_source
        else body_html
    )
    warnings_payload.extend(title_rule_warnings)

    required_fields = [
        field for field in REQUIRED_COLUMNS if not _text(merged.get(field))
    ]
    if required_fields:
        code = (
            "ARTICLE_DOCUMENT_METADATA_MISSING"
            if is_document_source
            else "ARTICLE_REQUIRED_FIELD_MISSING"
        )
        raise ArticleImportValidationError(
            f"缺少必填字段：{', '.join(sorted(required_fields))}", code=code
        )

    journal_slug = _text(merged.get("journal_slug"))
    if context.scope == ArticleImportScope.JOURNAL:
        journal = journals_by_id.get(context.target_journal_id)
        if not journal:
            raise ArticleImportValidationError(
                "锁定的目标子期刊不存在或已停用。", code="ARTICLE_JOURNAL_NOT_FOUND"
            )
        if journal_slug and journal_slug != journal.slug:
            raise ArticleImportValidationError(
                "清单或文档中的子期刊与服务端锁定范围不一致。",
                code="ARTICLE_JOURNAL_SCOPE_MISMATCH",
            )
    else:
        journal = journals_by_slug.get(journal_slug) if journal_slug else None
        if journal is None and context.default_journal_id:
            journal = journals_by_id.get(context.default_journal_id)
        if not journal:
            raise ArticleImportValidationError(
                "无法确定目标子期刊，请填写 journal_slug 或选择默认子期刊。",
                code="ARTICLE_JOURNAL_NOT_FOUND",
            )

    try:
        validate_slug(slug)
    except ValidationError as exc:
        raise ArticleImportValidationError(
            "slug 只能包含字母、数字、连字符或下划线。", code="ARTICLE_SLUG_INVALID"
        ) from exc
    key = (journal.pk, slug)
    if key in seen:
        raise ArticleImportValidationError(
            "同一导入文件中存在重复的子期刊 + slug。", code="ARTICLE_DUPLICATE_IN_FILE"
        )
    seen.add(key)
    article_type = _text(merged.get("article_type")).lower()
    if article_type not in ArticleType.values:
        raise ArticleImportValidationError(
            "article_type 不在允许范围内。", code="ARTICLE_TYPE_INVALID"
        )
    authors = _normalized_document_text(merged.get("authors"))
    if len(title) > 255 or len(authors) > 255:
        raise ArticleImportValidationError(
            "标题或作者字段超过 255 个字符。", code="ARTICLE_REQUIRED_FIELD_MISSING"
        )
    normalized_row = dict(merged)
    normalized_row["title"] = title
    normalized_row["slug"] = slug
    normalized_row["journal_slug"] = journal.slug
    normalized_row["authors"] = authors
    normalized_row["article_type"] = article_type
    suspicious = scan_mapping_for_suspicious_text(normalized_row)
    if suspicious and not allow_suspicious_text:
        raise ArticleImportValidationError(
            "检测到可疑文本，必须修正或由有权限用户强制处理。",
            code="ARTICLE_TEXT_SUSPICIOUS",
        )
    clean_html, body_images = _validate_html(
        clean_input, root=root, preview=preview, image_cache=image_cache
    )
    cover = None
    cover_ref = _text(normalized_row.get("cover_image"))
    if cover_ref:
        cover = _resolve_existing_image(cover_ref)
        if cover is None:
            cover_path = _safe_package_file(
                root, cover_ref, error_code="ARTICLE_COVER_IMAGE_NOT_FOUND"
            )
            cover = _materialize_image(cover_path, preview=preview, cache=image_cache)
    shadow = StaticArticle(journal=journal, slug=slug)
    categories = _resolve_article_import_categories(journal=journal, raw=normalized_row)
    if categories.provided:
        _validate_static_article_category_selection(
            article=shadow, categories=categories
        )
    return {
        "raw": normalized_row,
        "journal": journal,
        "slug": slug,
        "title": title,
        "article_type": article_type,
        "authors": authors,
        "ai_co_authors": _normalized_document_text(normalized_row.get("ai_co_authors")),
        "abstract": _normalized_document_text(normalized_row.get("abstract")),
        "keywords": _normalized_document_text(normalized_row.get("keywords")),
        "publication_date": _parse_publication_date(
            normalized_row.get("publication_date")
        ),
        "notes": _text(normalized_row.get("notes")),
        "html": body_html,
        "clean_html": clean_html,
        "cover": cover,
        "categories": categories,
        "suspicious": suspicious,
        "body_images": body_images,
        "source_path": source_path,
        "source_format": source_format,
        "source_sha256": source_sha256,
        "converter_name": converter_name,
        "converter_version": converter_version,
        "conversion_warnings": warnings_payload,
        "conversion_statistics": conversion_statistics,
        "metadata_sources": metadata_sources,
    }


def _row_failure(job, row_no: int, raw: dict, exc: Exception) -> ArticleImportRow:
    error_code = getattr(exc, "code", "ARTICLE_VALIDATION_ERROR")
    error_field = ERROR_FIELD_BY_CODE.get(error_code, "row")
    if error_code == "ARTICLE_REQUIRED_FIELD_MISSING":
        missing = [field for field in REQUIRED_COLUMNS if not _text(raw.get(field))]
        error_field = ",".join(sorted(missing)) or error_field
    suggestion = ERROR_SUGGESTION_BY_CODE.get(
        error_code, "请根据错误信息核对该行数据后重新上传。"
    )
    return ArticleImportRow(
        job=job,
        row_no=row_no,
        raw_data=raw,
        status=ImportRowStatus.FAILED,
        action="fail",
        error_code=error_code,
        error_field=error_field,
        error_message=_message(exc),
        normalized_data={
            "suggestion": suggestion,
            "source_format": (
                "docx"
                if _text(raw.get("docx_file"))
                else "markdown" if _text(raw.get("markdown_file")) else ""
            ),
            "source_path": _text(raw.get("docx_file") or raw.get("markdown_file")),
            "conversion_warning_codes": [],
        },
        source_path=_text(raw.get("docx_file") or raw.get("markdown_file")),
        source_format=(
            "docx"
            if _text(raw.get("docx_file"))
            else "markdown" if _text(raw.get("markdown_file")) else ""
        ),
    )


def _summary(rows: list[ArticleImportRow], *, suspicious_count: int = 0) -> dict:
    return {
        "total_rows": len(rows),
        "created_rows": sum(
            row.action == "create" and row.status != ImportRowStatus.FAILED
            for row in rows
        ),
        "updated_rows": sum(
            row.action == "update" and row.status != ImportRowStatus.FAILED
            for row in rows
        ),
        "skipped_rows": sum(row.status == ImportRowStatus.SKIPPED for row in rows),
        "failed_rows": sum(row.status == ImportRowStatus.FAILED for row in rows),
        "suspicious_text_count": suspicious_count,
    }


def _write_error_report(job: ArticleImportJob, rows: list[ArticleImportRow]) -> None:
    failed = [row for row in rows if row.status == ImportRowStatus.FAILED]
    if not failed:
        job.error_report = ""
        return
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(
        [
            "row_no",
            "source_format",
            "source_path",
            "converter",
            "conversion_warning_codes",
            "error_code",
            "error_field",
            "error_message",
            "suggestion",
            "title",
            "slug",
            "journal_slug",
        ]
    )
    for row in failed:
        writer.writerow(
            [
                row.row_no,
                (row.normalized_data or {}).get("source_format", row.source_format),
                (row.normalized_data or {}).get("source_path", row.source_path),
                (row.normalized_data or {}).get("converter", ""),
                ",".join(
                    (row.normalized_data or {}).get("conversion_warning_codes", [])
                ),
                row.error_code,
                row.error_field,
                row.error_message,
                (row.normalized_data or {}).get("suggestion", ""),
                row.raw_data.get("title", ""),
                row.raw_data.get("slug", ""),
                row.raw_data.get("journal_slug", ""),
            ]
        )
    job.error_report.save(
        f"article-import-{job.pk}-errors.csv",
        ContentFile(stream.getvalue().encode("utf-8-sig")),
        save=False,
    )


def _validate_article_import_context(context: ArticleImportContext) -> None:
    if context.scope not in ArticleImportScope.values:
        raise ArticleImportValidationError("导入范围无效。")
    if context.scope == ArticleImportScope.JOURNAL and not context.target_journal_id:
        raise ArticleImportValidationError("本刊模式必须锁定目标子期刊。")
    if (
        context.scope == ArticleImportScope.JOURNAL
        and not Journal.objects.filter(
            pk=context.target_journal_id, status=JournalStatus.ACTIVE
        ).exists()
    ):
        raise ArticleImportValidationError(
            "锁定的目标子期刊不存在或已停用。",
            code="ARTICLE_JOURNAL_NOT_FOUND",
        )


def _declared_source_format(name: str) -> str:
    suffix = Path(name).suffix.lower()
    if suffix in {".md", ".markdown"}:
        return "markdown"
    return suffix.lstrip(".")


def create_article_import_preview_job(
    source_file, *, context: ArticleImportContext, operator
) -> ArticleImportJob:
    """Persist an immutable pending preview job without converting the upload."""

    if not can_import_articles(operator):
        raise PermissionDenied
    _validate_article_import_context(context)
    data, name = _read_source(source_file)
    suffix = Path(name).suffix.lower()
    if suffix not in {".xlsx", ".csv", ".zip", ".docx", ".md", ".markdown"}:
        raise ArticleImportValidationError(
            "仅支持 XLSX、CSV、ZIP、DOCX、MD 或 Markdown 文件。",
            code="ARTICLE_DOCUMENT_FORMAT_UNSUPPORTED",
        )
    if suffix in {".docx", ".md", ".markdown"}:
        if not getattr(settings, "ARTICLE_DOCUMENT_IMPORT_ENABLED", True):
            raise ArticleImportValidationError(
                "DOCX/Markdown 文档导入功能当前已关闭。",
                code="ARTICLE_DOCUMENT_FORMAT_UNSUPPORTED",
            )
        if (
            context.scope == ArticleImportScope.GLOBAL
            and not context.default_journal_id
        ):
            raise ArticleImportValidationError(
                "全局模式直接上传 DOCX/Markdown 时必须选择默认子期刊。",
                code="ARTICLE_JOURNAL_NOT_FOUND",
            )

    digest = hashlib.sha256(data).hexdigest()
    concurrent_limit = int(
        getattr(settings, "ARTICLE_IMPORT_MAX_CONCURRENT_PREVIEWS_PER_USER", 2)
    )
    with transaction.atomic():
        operator.__class__._default_manager.select_for_update().get(pk=operator.pk)
        active_count = ArticleImportJob.objects.filter(
            operator=operator,
            confirmed_at__isnull=True,
            status__in=[ImportJobStatus.PENDING, ImportJobStatus.VALIDATING],
        ).count()
        if active_count >= concurrent_limit:
            raise ArticleImportValidationError(
                f"同一用户最多同时运行 {concurrent_limit} 个文章预检任务。",
                code="ARTICLE_IMPORT_PREVIEW_CONCURRENCY_LIMIT",
            )
        job = ArticleImportJob.objects.create(
            package_name=name,
            status=ImportJobStatus.PENDING,
            operator=operator,
            import_scope=context.scope,
            target_journal_id=context.target_journal_id,
            source_sha256=digest,
            source_format=_declared_source_format(name),
            template_version=(
                0
                if suffix in {".docx", ".md", ".markdown"}
                else ARTICLE_IMPORT_TEMPLATE_VERSION
            ),
            summary={
                "default_journal_id": context.default_journal_id,
                "csv_encoding": context.csv_encoding,
                "document_defaults": context.document_defaults or {},
                "preview_processed_rows": 0,
                "preview_total_rows": 0,
                "conversion_warning_count": 0,
            },
        )
        job.source_file.save(name, ContentFile(data), save=True)

    record_audit_event(
        action=AuditAction.IMPORT,
        status=AuditStatus.STARTED,
        actor=operator,
        target=job,
        message="已创建文章导入预检任务",
        metadata={
            "job_id": job.pk,
            "scope": context.scope,
            "target_journal_id": context.target_journal_id,
            "source_format": job.source_format,
            "source_sha256": digest,
            "filename": name,
        },
    )
    return job


def _start_article_import_preview(job: ArticleImportJob, *, actor) -> ArticleImportJob:
    rejection: Exception | None = None
    with transaction.atomic():
        locked = (
            ArticleImportJob.objects.select_for_update(of=("self",))
            .select_related("target_journal", "operator")
            .get(pk=job.pk)
        )
        if locked.status != ImportJobStatus.PENDING or locked.confirmed_at is not None:
            rejection = ArticleImportValidationError(
                "预检任务不是可启动的 PENDING 状态。",
                code="ARTICLE_IMPORT_PREVIEW_STATE_INVALID",
            )
        elif actor is None or not can_import_articles(actor):
            rejection = PermissionDenied("当前操作人已无文章导入权限。")
        elif not can_view_article_import_job(actor, locked):
            rejection = PermissionDenied("当前操作人无权执行该文章预检任务。")
        elif locked.import_scope == ArticleImportScope.JOURNAL and (
            locked.target_journal_id is None
            or locked.target_journal.status != JournalStatus.ACTIVE
        ):
            rejection = ArticleImportValidationError(
                "锁定的目标子期刊不存在或已停用。",
                code="ARTICLE_JOURNAL_NOT_FOUND",
            )

        if rejection is None:
            summary = dict(locked.summary or {})
            summary.update(
                {
                    "preview_processed_rows": 0,
                    "preview_total_rows": 0,
                    "conversion_warning_count": 0,
                }
            )
            locked.summary = summary
            locked.status = ImportJobStatus.VALIDATING
            locked.started_at = timezone.now()
            locked.finished_at = None
            locked.notes = ""
            locked.save(
                update_fields=[
                    "summary",
                    "status",
                    "started_at",
                    "finished_at",
                    "notes",
                    "updated_at",
                ]
            )

    if rejection is not None:
        raise rejection
    record_audit_event(
        action=AuditAction.IMPORT,
        status=AuditStatus.STARTED,
        actor=actor,
        target=locked,
        message="开始文章导入预览",
        metadata={
            "job_id": locked.pk,
            "scope": locked.import_scope,
            "target_journal_id": locked.target_journal_id,
            "source_format": locked.source_format,
            "source_sha256": locked.source_sha256,
        },
    )
    return locked


def _update_preview_progress(
    job: ArticleImportJob,
    *,
    processed_rows: int,
    total_rows: int,
    warning_count: int,
) -> None:
    summary = dict(job.summary or {})
    summary.update(
        {
            "preview_processed_rows": processed_rows,
            "preview_total_rows": total_rows,
            "conversion_warning_count": warning_count,
        }
    )
    ArticleImportJob.objects.filter(
        pk=job.pk, status=ImportJobStatus.VALIDATING
    ).update(summary=summary, updated_at=timezone.now())
    job.summary = summary


def _preview_normalized_data(normalized: dict) -> dict:
    statistics = dict(normalized["conversion_statistics"] or {})
    preview_text = BeautifulSoup(normalized["clean_html"], "html.parser").get_text(
        " ", strip=True
    )[:1000]
    warning_codes = [
        item.get("code")
        for item in normalized["conversion_warnings"]
        if item.get("code")
    ]
    return {
        "journal_id": normalized["journal"].pk,
        "journal_slug": normalized["journal"].slug,
        "title": normalized["title"],
        "slug": normalized["slug"],
        "article_type": normalized["article_type"],
        "authors": normalized["authors"],
        "body_sha256": hashlib.sha256(normalized["html"].encode()).hexdigest(),
        "sanitized_sha256": hashlib.sha256(
            normalized["clean_html"].encode()
        ).hexdigest(),
        "source_sha256": normalized["source_sha256"],
        "source_format": normalized["source_format"],
        "source_path": normalized["source_path"],
        "converter_name": normalized["converter_name"],
        "converter_version": normalized["converter_version"],
        "converter": " ".join(
            filter(
                None,
                [normalized["converter_name"], normalized["converter_version"]],
            )
        ),
        "generated_asset_count": statistics.get(
            "image_count", len(normalized["body_images"])
        ),
        "warning_codes": warning_codes,
        "conversion_warning_codes": warning_codes,
        "conversion_statistics": statistics,
        "metadata_sources": normalized["metadata_sources"],
        "body_preview_text": preview_text,
        "body_security_validated": True,
    }


def execute_article_import_preview(
    job: ArticleImportJob,
    *,
    operator=None,
    package_path: Path | None = None,
) -> ArticleImportJob:
    actor = operator or job.operator
    job = _start_article_import_preview(job, actor=actor)
    summary_data = dict(job.summary or {})
    context = ArticleImportContext(
        scope=job.import_scope,
        target_journal_id=job.target_journal_id,
        default_journal_id=summary_data.get("default_journal_id"),
        csv_encoding=summary_data.get("csv_encoding", "auto"),
        document_defaults=summary_data.get("document_defaults") or {},
    )
    try:
        if package_path is None:
            data = _job_source_bytes(job)
        else:
            data = Path(package_path).read_bytes()
        digest = hashlib.sha256(data).hexdigest()
        if digest != job.source_sha256:
            raise ArticleImportValidationError(
                "队列文件哈希与任务记录不一致。",
                code="ARTICLE_IMPORT_HASH_MISMATCH",
            )
        with tempfile.TemporaryDirectory(prefix="article-import-preview-") as temp:
            extraction_budget = ImportExtractionBudget(
                max_total_bytes=getattr(
                    settings,
                    "ARTICLE_IMPORT_MAX_LOGICAL_EXTRACTED_SIZE",
                    MAX_ZIP_TOTAL_SIZE,
                ),
                max_nested_members=getattr(
                    settings, "ARTICLE_IMPORT_MAX_NESTED_MEMBERS_PER_JOB", 20_000
                ),
            )
            package = _load_package(
                data,
                job.package_name,
                context,
                Path(temp),
                budget=extraction_budget,
            )
            job.source_format = package.source_format
            job.parser_version = package.parser_version
            journals_by_id, journals_by_slug = _active_journal_lookup()
            existing_articles = _existing_static_article_map(
                package.rows, context, journals_by_id, journals_by_slug
            )
            seen: set[tuple[int, str]] = set()
            rows: list[ArticleImportRow] = []
            image_cache: dict[str, CustomImage] = {}
            document_cache: dict[str, object] = {}
            suspicious_count = 0
            warning_count = 0
            total_rows = len(package.rows)
            _update_preview_progress(
                job,
                processed_rows=0,
                total_rows=total_rows,
                warning_count=0,
            )
            for index, raw in enumerate(package.rows, start=1):
                row_no = index + 1
                raw = dict(raw)
                if not any(_text(value) for value in raw.values()):
                    rows.append(
                        ArticleImportRow(
                            job=job,
                            row_no=row_no,
                            raw_data=raw,
                            status=ImportRowStatus.SKIPPED,
                            action="skip",
                        )
                    )
                else:
                    try:
                        normalized = _normalized_row(
                            raw,
                            context,
                            package.root,
                            preview=True,
                            seen=seen,
                            image_cache=image_cache,
                            allow_suspicious_text=False,
                            journals_by_id=journals_by_id,
                            journals_by_slug=journals_by_slug,
                            document_cache=document_cache,
                            extraction_budget=extraction_budget,
                        )
                        article_key = (
                            normalized["journal"].pk,
                            normalized["slug"],
                        )
                        action = (
                            "update" if article_key in existing_articles else "create"
                        )
                        warning_count += len(normalized["conversion_warnings"])
                        rows.append(
                            ArticleImportRow(
                                job=job,
                                row_no=row_no,
                                raw_data=raw,
                                status=(
                                    ImportRowStatus.UPDATED
                                    if action == "update"
                                    else ImportRowStatus.SUCCESS
                                ),
                                action=action,
                                source_path=normalized["source_path"],
                                source_format=normalized["source_format"],
                                conversion_warnings=normalized["conversion_warnings"],
                                normalized_data=_preview_normalized_data(normalized),
                            )
                        )
                    except Exception as exc:
                        if getattr(exc, "code", "") == "ARTICLE_TEXT_SUSPICIOUS":
                            suspicious_count += 1
                        rows.append(_row_failure(job, row_no, raw, exc))
                if index == total_rows or index % 25 == 0:
                    _update_preview_progress(
                        job,
                        processed_rows=index,
                        total_rows=total_rows,
                        warning_count=warning_count,
                    )

            summary = _summary(rows, suspicious_count=suspicious_count)
            document_paths = {
                row.source_path
                for row in rows
                if row.source_format in {"docx", "markdown"} and row.source_path
            }
            converter_pairs = sorted(
                {
                    (
                        (row.normalized_data or {}).get("converter_name", ""),
                        (row.normalized_data or {}).get("converter_version", ""),
                    )
                    for row in rows
                    if (row.normalized_data or {}).get("converter_name")
                }
            )
            summary.update(
                {
                    "default_journal_id": context.default_journal_id,
                    "csv_encoding": context.csv_encoding,
                    "document_defaults": context.document_defaults or {},
                    "template_warning": package.template_warning,
                    "source_format": package.source_format,
                    "parser_version": package.parser_version,
                    "preview_processed_rows": total_rows,
                    "preview_total_rows": total_rows,
                    "conversion_warning_count": warning_count,
                    "conversion_warning_codes": sorted(
                        {
                            code
                            for row in rows
                            for code in (row.normalized_data or {}).get(
                                "warning_codes", []
                            )
                        }
                    ),
                    "document_count": len(document_paths),
                    "generated_image_count": sum(
                        int((row.normalized_data or {}).get("generated_asset_count", 0))
                        for row in rows
                    ),
                    "converters": [
                        {"name": name, "version": version}
                        for name, version in converter_pairs
                    ],
                }
            )
            with transaction.atomic():
                locked = ArticleImportJob.objects.select_for_update().get(pk=job.pk)
                if locked.status != ImportJobStatus.VALIDATING:
                    raise ArticleImportValidationError(
                        "预检任务状态已变化，不能写入重复预览结果。",
                        code="ARTICLE_IMPORT_PREVIEW_STATE_INVALID",
                    )
                locked.rows.all().delete()
                for row in rows:
                    row.job = locked
                ArticleImportRow.objects.bulk_create(rows)
                locked.template_version = package.template_version
                locked.source_format = package.source_format
                locked.parser_version = package.parser_version
                locked.total_rows = len(rows)
                locked.success_rows = summary["created_rows"] + summary["updated_rows"]
                locked.failed_rows = summary["failed_rows"]
                locked.summary = summary
                locked.status = ImportJobStatus.READY
                locked.finished_at = timezone.now()
                _write_error_report(locked, rows)
                locked.save()
                job = locked

        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.SUCCESS,
            actor=actor,
            target=job,
            message="文章导入预览完成",
            metadata={
                "job_id": job.pk,
                "scope": context.scope,
                "target_journal_id": context.target_journal_id,
                "source_format": job.source_format,
                "source_sha256": digest,
                "parser_version": job.parser_version,
                **summary,
            },
        )
        forbidden_warning_count = sum(
            1
            for row in rows
            for warning in row.conversion_warnings
            if warning.get("code") == "ARTICLE_DOCUMENT_FORBIDDEN_METADATA_IGNORED"
        )
        if forbidden_warning_count:
            record_audit_event(
                action=AuditAction.IMPORT,
                status=AuditStatus.SUCCESS,
                actor=actor,
                target=job,
                message="文章导入预检已忽略禁止元数据",
                metadata={
                    "job_id": job.pk,
                    "count": forbidden_warning_count,
                    "warning_code": "ARTICLE_DOCUMENT_FORBIDDEN_METADATA_IGNORED",
                },
            )
        return job
    except Exception as exc:
        failure_recorded = False
        with transaction.atomic():
            locked = ArticleImportJob.objects.select_for_update().get(pk=job.pk)
            if locked.status == ImportJobStatus.VALIDATING:
                locked.status = ImportJobStatus.FAILED
                locked.notes = _message(exc)[:1000]
                locked.finished_at = timezone.now()
                locked.save(
                    update_fields=["status", "notes", "finished_at", "updated_at"]
                )
                failure_recorded = True
            job = locked
        if failure_recorded:
            record_audit_event(
                action=AuditAction.IMPORT,
                status=AuditStatus.FAILURE,
                actor=actor,
                target=job,
                message="文章导入预览失败",
                metadata={
                    "job_id": job.pk,
                    "source_format": job.source_format,
                    "source_sha256": job.source_sha256,
                    "error": _message(exc)[:500],
                    "error_code": getattr(
                        exc, "code", "ARTICLE_DOCUMENT_CONVERSION_FAILED"
                    ),
                },
            )
        raise


def preview_article_import(
    source_file, *, context: ArticleImportContext, operator
) -> ArticleImportJob:
    """Compatibility wrapper for synchronous XLSX/CSV preview and tests."""

    job = create_article_import_preview_job(
        source_file, context=context, operator=operator
    )
    return execute_article_import_preview(job, operator=operator)


def fail_stale_article_import_previews(*, operator=None) -> int:
    timeout_seconds = int(
        getattr(settings, "ARTICLE_IMPORT_PREVIEW_TIMEOUT_SECONDS", 600)
    )
    cutoff = timezone.now() - timedelta(seconds=timeout_seconds)
    stale_ids = list(
        ArticleImportJob.objects.filter(
            status=ImportJobStatus.VALIDATING,
            started_at__lt=cutoff,
        ).values_list("pk", flat=True)
    )
    failed = 0
    for job_id in stale_ids:
        with transaction.atomic():
            job = ArticleImportJob.objects.select_for_update().get(pk=job_id)
            if (
                job.status != ImportJobStatus.VALIDATING
                or job.started_at is None
                or job.started_at >= cutoff
            ):
                continue
            job.status = ImportJobStatus.FAILED
            job.notes = "后台预检超过允许时长，已由运维处理标记为失败。"
            job.finished_at = timezone.now()
            job.save(update_fields=["status", "notes", "finished_at", "updated_at"])
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.FAILURE,
            actor=operator,
            target=job,
            message="超时文章预检任务已标记失败",
            metadata={
                "job_id": job.pk,
                "timeout_seconds": timeout_seconds,
                "source_format": job.source_format,
            },
        )
        failed += 1
    return failed


def confirm_article_import(
    preview_job: ArticleImportJob,
    *,
    operator,
    allow_suspicious_text: bool = False,
    override_reason: str = "",
) -> ArticleImportJob:
    if not can_import_articles(operator):
        raise PermissionDenied
    if not can_view_article_import_job(operator, preview_job):
        raise PermissionDenied
    if allow_suspicious_text and not can_override_suspicious_article_text(operator):
        raise PermissionDenied
    if allow_suspicious_text and len(override_reason.strip()) < 8:
        raise ArticleImportValidationError(
            "强制处理可疑文本必须填写至少 8 个字符的理由。"
        )
    with transaction.atomic():
        job = ArticleImportJob.objects.select_for_update().get(pk=preview_job.pk)
        if job.status != ImportJobStatus.READY or job.confirmed_at:
            raise ArticleImportValidationError(
                "任务不是可确认状态，或已被确认。",
                code="ARTICLE_IMPORT_ALREADY_CONFIRMED",
            )
        if hashlib.sha256(_job_source_bytes(job)).hexdigest() != job.source_sha256:
            raise ArticleImportValidationError(
                "源文件哈希与预览时不一致。", code="ARTICLE_IMPORT_HASH_MISMATCH"
            )
        if (
            job.import_scope == ArticleImportScope.JOURNAL
            and not Journal.objects.filter(
                pk=job.target_journal_id, status=JournalStatus.ACTIVE
            ).exists()
        ):
            raise ArticleImportValidationError(
                "锁定的目标子期刊已不可用。", code="ARTICLE_JOURNAL_NOT_FOUND"
            )
        suspicious = int((job.summary or {}).get("suspicious_text_count", 0))
        if suspicious and not allow_suspicious_text:
            raise ArticleImportValidationError(
                "任务包含可疑文本，必须修正或由全局管理员强制处理。",
                code="ARTICLE_TEXT_SUSPICIOUS",
            )
        summary = dict(job.summary or {})
        summary.update(
            {
                "allow_suspicious_text": allow_suspicious_text,
                "override_reason": override_reason.strip(),
            }
        )
        job.summary = summary
        job.confirmed_at = timezone.now()
        job.confirmed_by = operator
        job.status = ImportJobStatus.PENDING
        job.save(
            update_fields=[
                "summary",
                "confirmed_at",
                "confirmed_by",
                "status",
                "updated_at",
            ]
        )
    record_audit_event(
        action=AuditAction.IMPORT,
        status=AuditStatus.STARTED,
        actor=operator,
        target=job,
        message="已确认文章导入",
        metadata={
            "preview_job_id": job.pk,
            "confirmed_by_id": getattr(operator, "pk", None),
            "allow_suspicious_text": allow_suspicious_text,
        },
    )
    if allow_suspicious_text:
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.SUCCESS,
            actor=operator,
            target=job,
            message="强制按原文处理可疑文本",
            metadata={
                "reason": override_reason.strip(),
                "suspicious_text_count": suspicious,
            },
        )
    return job


def fail_pending_article_import_job(
    job: ArticleImportJob,
    *,
    operator=None,
    message: str,
    code: str = "ARTICLE_IMPORT_START_FAILED",
) -> bool:
    """Move a confirmed pending job to its required failed terminal state."""

    with transaction.atomic():
        locked = ArticleImportJob.objects.select_for_update().get(pk=job.pk)
        if locked.status != ImportJobStatus.PENDING:
            return False
        locked.status = ImportJobStatus.FAILED
        locked.notes = message
        locked.finished_at = timezone.now()
        locked.save(update_fields=["status", "notes", "finished_at", "updated_at"])
    record_audit_event(
        action=AuditAction.IMPORT,
        status=AuditStatus.FAILURE,
        actor=operator,
        target=locked,
        message="文章导入启动失败",
        metadata={"error": message[:500], "error_code": code},
    )
    return True


def _start_confirmed_article_import(
    job: ArticleImportJob, *, actor
) -> ArticleImportJob:
    rejection: Exception | None = None
    rejection_code = ""
    mark_failed = False
    with transaction.atomic():
        locked = (
            ArticleImportJob.objects.select_for_update(of=("self",))
            .select_related("operator", "confirmed_by", "target_journal")
            .get(pk=job.pk)
        )
        if locked.status != ImportJobStatus.PENDING:
            rejection = ArticleImportValidationError(
                "文章导入任务不处于待执行状态。",
                code="ARTICLE_IMPORT_STATE_INVALID",
            )
            rejection_code = "ARTICLE_IMPORT_STATE_INVALID"
        elif not locked.confirmed_at:
            rejection = ArticleImportValidationError(
                "文章导入任务尚未确认。",
                code="ARTICLE_IMPORT_NOT_CONFIRMED",
            )
            rejection_code = "ARTICLE_IMPORT_NOT_CONFIRMED"
            mark_failed = True
        elif actor is None or not can_import_articles(actor):
            rejection = PermissionDenied("当前操作人已无文章导入权限。")
            rejection_code = "ARTICLE_IMPORT_PERMISSION_REVOKED"
            mark_failed = True
        elif not can_view_article_import_job(actor, locked):
            rejection = PermissionDenied("当前操作人无权执行该文章导入任务。")
            rejection_code = "ARTICLE_IMPORT_JOB_FORBIDDEN"
            mark_failed = True
        elif locked.import_scope == ArticleImportScope.JOURNAL and (
            locked.target_journal_id is None
            or locked.target_journal.status != JournalStatus.ACTIVE
        ):
            rejection = ArticleImportValidationError(
                "锁定的目标子期刊不存在或已停用。",
                code="ARTICLE_IMPORT_TARGET_JOURNAL_INVALID",
            )
            rejection_code = "ARTICLE_IMPORT_TARGET_JOURNAL_INVALID"
            mark_failed = True

        if rejection is not None and mark_failed:
            locked.status = ImportJobStatus.FAILED
            locked.notes = _message(rejection)
            locked.finished_at = timezone.now()
            locked.save(update_fields=["status", "notes", "finished_at", "updated_at"])
        elif rejection is None:
            locked.status = ImportJobStatus.IMPORTING
            locked.started_at = timezone.now()
            locked.finished_at = None
            locked.notes = ""
            locked.save(
                update_fields=[
                    "status",
                    "started_at",
                    "finished_at",
                    "notes",
                    "updated_at",
                ]
            )

    if rejection is not None:
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.FAILURE,
            actor=actor,
            target=locked,
            message="文章导入执行被拒绝",
            metadata={
                "error": _message(rejection)[:500],
                "error_code": rejection_code,
                "job_status": locked.status,
            },
        )
        raise rejection
    return locked


def execute_confirmed_article_import(
    job: ArticleImportJob, *, operator=None, allow_suspicious_text: bool | None = None
) -> ArticleImportJob:
    actor = operator or job.confirmed_by or job.operator
    job = _start_confirmed_article_import(job, actor=actor)
    summary_data = dict(job.summary or {})
    if allow_suspicious_text is None:
        allow_suspicious_text = bool(summary_data.get("allow_suspicious_text"))
    context = ArticleImportContext(
        scope=job.import_scope,
        target_journal_id=job.target_journal_id,
        default_journal_id=summary_data.get("default_journal_id"),
        csv_encoding=summary_data.get("csv_encoding", "auto"),
        document_defaults=summary_data.get("document_defaults") or {},
    )
    try:
        data = _job_source_bytes(job)
        if hashlib.sha256(data).hexdigest() != job.source_sha256:
            raise ArticleImportValidationError(
                "源文件哈希与预览时不一致。", code="ARTICLE_IMPORT_HASH_MISMATCH"
            )
        with tempfile.TemporaryDirectory(prefix="article-import-execute-") as temp:
            extraction_budget = ImportExtractionBudget(
                max_total_bytes=getattr(
                    settings,
                    "ARTICLE_IMPORT_MAX_LOGICAL_EXTRACTED_SIZE",
                    MAX_ZIP_TOTAL_SIZE,
                ),
                max_nested_members=getattr(
                    settings, "ARTICLE_IMPORT_MAX_NESTED_MEMBERS_PER_JOB", 20_000
                ),
            )
            package = _load_package(
                data, job.package_name, context, Path(temp), budget=extraction_budget
            )
            journals_by_id, journals_by_slug = _active_journal_lookup()
            existing_articles = _existing_static_article_map(
                package.rows, context, journals_by_id, journals_by_slug
            )
            job.rows.all().delete()
            seen: set[tuple[int, str]] = set()
            image_cache: dict[str, CustomImage] = {}
            document_cache: dict[str, object] = {}
            results = []
            suspicious_count = 0
            for row_no, raw in enumerate(package.rows, start=2):
                raw = dict(raw)
                if not any(_text(value) for value in raw.values()):
                    results.append(
                        ArticleImportRow.objects.create(
                            job=job,
                            row_no=row_no,
                            raw_data=raw,
                            status=ImportRowStatus.SKIPPED,
                            action="skip",
                        )
                    )
                    continue
                try:
                    with transaction.atomic():
                        normalized = _normalized_row(
                            raw,
                            context,
                            package.root,
                            preview=False,
                            seen=seen,
                            image_cache=image_cache,
                            allow_suspicious_text=bool(allow_suspicious_text),
                            journals_by_id=journals_by_id,
                            journals_by_slug=journals_by_slug,
                            document_cache=document_cache,
                            extraction_budget=extraction_budget,
                        )
                        article_key = (normalized["journal"].pk, normalized["slug"])
                        article = existing_articles.get(article_key)
                        created = article is None
                        article = article or StaticArticle(
                            journal=normalized["journal"], slug=normalized["slug"]
                        )
                        article.title = normalized["title"]
                        article.article_type = normalized["article_type"]
                        article.authors = normalized["authors"]
                        article.ai_co_authors = normalized["ai_co_authors"]
                        article.abstract = normalized["abstract"] or normalized["title"]
                        article.keywords = normalized["keywords"] or "imported"
                        article.publication_date = normalized["publication_date"]
                        article.cover_image = normalized["cover"]
                        article.notes = normalized["notes"]
                        article.review_status = ArticleReviewStatus.DRAFT
                        article.build_version = ""
                        article.sort_order = 0
                        article.is_pinned = False
                        article.html_source.save(
                            f"{normalized['slug']}.html",
                            ContentFile(normalized["clean_html"].encode("utf-8")),
                            save=False,
                        )
                        article.source_html_path = _text(raw.get("html_file"))
                        article.full_clean()
                        article.save()
                        existing_articles[article_key] = article
                        if normalized["categories"].provided:
                            assign_static_article_categories(
                                static_article=article,
                                primary=normalized["categories"].primary,
                                related=normalized["categories"].related,
                            )
                        page = sync_imported_article(
                            article, source_html=normalized["clean_html"], owner=actor
                        )
                        page.featured_image = normalized["cover"]
                        page.review_status = ArticlePage.ReviewStatus.DRAFT
                        page.publication_status = ""
                        page.build_version = ""
                        page.published_version = ""
                        page.live = False
                        page.has_unpublished_changes = True
                        page.save(
                            clean=False,
                            user=actor,
                            bypass_article_permission_check=True,
                        )
                        page.save_revision(
                            user=actor,
                            changed=True,
                            bypass_article_permission_check=True,
                        )
                        result = ArticleImportRow.objects.create(
                            job=job,
                            row_no=row_no,
                            raw_data=raw,
                            article=article,
                            article_page=page,
                            status=(
                                ImportRowStatus.SUCCESS
                                if created
                                else ImportRowStatus.UPDATED
                            ),
                            action="create" if created else "update",
                            source_path=normalized["source_path"],
                            source_format=normalized["source_format"],
                            conversion_warnings=normalized["conversion_warnings"],
                            normalized_data={
                                "journal_id": normalized["journal"].pk,
                                "journal_slug": normalized["journal"].slug,
                                "body_sha256": hashlib.sha256(
                                    normalized["html"].encode()
                                ).hexdigest(),
                                "sanitized_sha256": hashlib.sha256(
                                    normalized["clean_html"].encode()
                                ).hexdigest(),
                                "source_sha256": normalized["source_sha256"],
                                "source_format": normalized["source_format"],
                                "source_path": normalized["source_path"],
                                "converter_name": normalized["converter_name"],
                                "converter_version": normalized["converter_version"],
                                "generated_asset_count": normalized[
                                    "conversion_statistics"
                                ].get("image_count", len(normalized["body_images"])),
                                "warning_codes": [
                                    item.get("code")
                                    for item in normalized["conversion_warnings"]
                                ],
                                "metadata_sources": normalized["metadata_sources"],
                            },
                        )
                        results.append(result)
                except Exception as exc:
                    if getattr(exc, "code", "") == "ARTICLE_TEXT_SUSPICIOUS":
                        suspicious_count += 1
                    failure = _row_failure(job, row_no, raw, exc)
                    failure.save()
                    results.append(failure)
            final_summary = _summary(results, suspicious_count=suspicious_count)
            final_summary.update(
                {
                    key: summary_data.get(key)
                    for key in (
                        "default_journal_id",
                        "csv_encoding",
                        "document_defaults",
                        "allow_suspicious_text",
                        "override_reason",
                        "template_warning",
                        "source_format",
                        "parser_version",
                        "conversion_warning_count",
                        "conversion_warning_codes",
                        "document_count",
                        "generated_image_count",
                        "converters",
                    )
                }
            )
            job.total_rows = len(results)
            job.success_rows = (
                final_summary["created_rows"] + final_summary["updated_rows"]
            )
            job.failed_rows = final_summary["failed_rows"]
            job.summary = final_summary
            job.status = ImportJobStatus.COMPLETED
            job.finished_at = timezone.now()
            _write_error_report(job, results)
            job.save()
        record_audit_event(
            action=AuditAction.IMPORT,
            status=AuditStatus.SUCCESS,
            actor=actor,
            target=job,
            message="文章导入完成",
            metadata=final_summary,
        )
    except Exception as exc:
        failure_recorded = False
        with transaction.atomic():
            locked = ArticleImportJob.objects.select_for_update().get(pk=job.pk)
            if locked.status == ImportJobStatus.IMPORTING:
                locked.status = ImportJobStatus.FAILED
                locked.notes = _message(exc)
                locked.finished_at = timezone.now()
                locked.save(
                    update_fields=["status", "notes", "finished_at", "updated_at"]
                )
                failure_recorded = True
            job = locked
        if failure_recorded:
            record_audit_event(
                action=AuditAction.IMPORT,
                status=AuditStatus.FAILURE,
                actor=actor,
                target=job,
                message="文章导入失败",
                metadata={
                    "error": _message(exc)[:500],
                    "error_code": getattr(exc, "code", "ARTICLE_IMPORT_FAILED"),
                },
            )
        raise
    return job
