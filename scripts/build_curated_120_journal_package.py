from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from datetime import datetime, timedelta, timezone
from html import escape
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

JOURNAL_HEADERS = [
    "journal_name",
    "journal_name_cn",
    "slug",
    "az_group",
    "status",
    "sort_order",
    "seo_title",
    "seo_description",
    "homepage_intro",
    "cover_image",
    "metrics_image",
    "static_site_path",
    "target_article_count",
    "notes",
]
ARTICLE_HEADERS = [
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "status",
    "sort_order",
    "is_pinned",
    "cover_image",
    "body_html",
    "html_file",
    "build_version",
    "static_output_path",
    "primary_category_code",
    "primary_category_path",
    "related_category_codes",
    "related_category_paths",
    "main_site_slot",
    "main_site_slot_name",
    "main_site_slot_title",
    "main_site_slot_summary",
    "main_site_slot_order",
    "main_site_slot_pinned",
    "journal_slot",
    "journal_slot_name",
    "journal_slot_title",
    "journal_slot_summary",
    "journal_slot_order",
    "journal_slot_pinned",
    "notes",
]

GROUPS = [
    (
        "Core Models and Learning",
        "基础模型与学习",
        [
            (
                "Foundation Model Systems",
                "基础模型系统",
                "scaling and deploying general-purpose models",
                "scaling laws|mixture of experts|context efficiency|model routing",
            ),
            (
                "Machine Learning Theory",
                "机器学习理论",
                "the mathematical foundations of learning algorithms",
                "generalization bounds|optimization geometry|implicit bias|statistical efficiency",
            ),
            (
                "Neural Architecture Research",
                "神经网络架构研究",
                "designing reliable neural architectures",
                "state-space models|sparse attention|modular networks|architecture search",
            ),
            (
                "Representation Learning",
                "表征学习",
                "learning transferable representations from complex data",
                "contrastive objectives|disentanglement|latent spaces|cross-domain transfer",
            ),
            (
                "Self-Supervised Intelligence",
                "自监督智能",
                "learning useful structure without dense labels",
                "masked modeling|predictive coding|pseudo-labeling|pretraining curricula",
            ),
            (
                "Reinforcement Learning Systems",
                "强化学习系统",
                "decision-making through interaction and feedback",
                "offline reinforcement learning|reward design|policy optimization|sim-to-real transfer",
            ),
            (
                "Continual Learning",
                "持续学习",
                "models that adapt without catastrophic forgetting",
                "memory replay|task boundaries|concept drift|lifelong evaluation",
            ),
            (
                "Causal Machine Learning",
                "因果机器学习",
                "causal reasoning for robust prediction and intervention",
                "causal discovery|counterfactual estimation|treatment effects|invariant prediction",
            ),
            (
                "Probabilistic Artificial Intelligence",
                "概率人工智能",
                "uncertainty-aware modeling and inference",
                "Bayesian deep learning|calibration|probabilistic programming|uncertainty decomposition",
            ),
            (
                "Evolutionary Computation",
                "进化计算",
                "population-based search and adaptive optimization",
                "neuroevolution|quality diversity|genetic programming|evolution strategies",
            ),
        ],
    ),
    (
        "Language Knowledge and Multimodal",
        "语言知识与多模态",
        [
            (
                "Language Model Engineering",
                "语言模型工程",
                "engineering dependable language model applications",
                "prompt orchestration|retrieval augmentation|long-context serving|evaluation harnesses",
            ),
            (
                "Natural Language Understanding",
                "自然语言理解",
                "machine understanding of meaning, intent, and discourse",
                "semantic parsing|discourse modeling|pragmatics|reasoning benchmarks",
            ),
            (
                "Multilingual Artificial Intelligence",
                "多语言人工智能",
                "inclusive language technology across scripts and regions",
                "low-resource transfer|cross-lingual alignment|code switching|cultural adaptation",
            ),
            (
                "Speech and Audio Intelligence",
                "语音与音频智能",
                "learning from speech, sound, and acoustic scenes",
                "speech recognition|voice synthesis|audio event detection|speaker modeling",
            ),
            (
                "Knowledge Graph Systems",
                "知识图谱系统",
                "structured knowledge acquisition and reasoning",
                "entity resolution|graph completion|ontology alignment|temporal knowledge graphs",
            ),
            (
                "Semantic Computing",
                "语义计算",
                "computing systems that preserve meaning across data and tasks",
                "semantic interoperability|concept embeddings|metadata quality|reasoning services",
            ),
            (
                "Information Retrieval",
                "信息检索",
                "finding trustworthy information at scale",
                "dense retrieval|hybrid search|ranking calibration|query understanding",
            ),
            (
                "Recommendation Intelligence",
                "推荐智能",
                "responsible personalized ranking and discovery",
                "session recommendation|cold start|diversity objectives|causal recommendation",
            ),
            (
                "Multimodal Understanding",
                "多模态理解",
                "joint reasoning across text, image, audio, and video",
                "vision-language alignment|multimodal grounding|video reasoning|cross-modal retrieval",
            ),
            (
                "Document Intelligence",
                "文档智能",
                "extracting structure and evidence from complex documents",
                "layout understanding|table extraction|document question answering|citation tracing",
            ),
        ],
    ),
    (
        "Agents Robotics and Autonomy",
        "智能体、机器人与自主系统",
        [
            (
                "Agentic Artificial Intelligence Systems",
                "智能体人工智能系统",
                "autonomous systems that plan, use tools, and learn from outcomes",
                "task planning|memory systems|tool selection|agent evaluation",
            ),
            (
                "Tool-Using Models",
                "工具调用模型",
                "models that reliably operate software and external services",
                "API grounding|computer use|tool error recovery|permission boundaries",
            ),
            (
                "Multi-Agent Coordination",
                "多智能体协同",
                "coordination, negotiation, and specialization among agents",
                "role allocation|communication protocols|collective planning|conflict resolution",
            ),
            (
                "Embodied Intelligence",
                "具身智能",
                "intelligence grounded in perception and physical action",
                "sensorimotor learning|world models|affordance discovery|embodied benchmarks",
            ),
            (
                "Robot Learning",
                "机器人学习",
                "data-efficient learning for adaptable robots",
                "imitation learning|grasping policies|robot foundation models|safety constraints",
            ),
            (
                "Autonomous Vehicles",
                "自动驾驶系统",
                "safe perception and decision-making for road autonomy",
                "trajectory prediction|sensor fusion|scenario generation|fallback planning",
            ),
            (
                "Drone Intelligence",
                "无人机智能",
                "aerial autonomy for inspection, mapping, and response",
                "visual navigation|swarm control|energy-aware routing|aerial perception",
            ),
            (
                "Industrial Robotics",
                "工业机器人",
                "flexible automation for factories and warehouses",
                "bin picking|predictive maintenance|motion planning|digital twins",
            ),
            (
                "Human-Robot Collaboration",
                "人机协作机器人",
                "safe and legible collaboration between people and robots",
                "intent recognition|shared control|ergonomic planning|trust calibration",
            ),
            (
                "Spatial Intelligence",
                "空间智能",
                "reasoning about geometry, place, and three-dimensional environments",
                "3D scene graphs|neural mapping|spatial memory|indoor navigation",
            ),
        ],
    ),
    (
        "Trust Safety and Governance",
        "可信、安全与治理",
        [
            (
                "Artificial Intelligence Safety and Alignment",
                "人工智能安全与对齐",
                "keeping advanced systems controllable and aligned with human intent",
                "scalable oversight|reward hacking|alignment evaluations|corrigibility",
            ),
            (
                "Responsible Artificial Intelligence Governance",
                "负责任人工智能治理",
                "governance practices for accountable AI deployment",
                "risk classification|model documentation|impact assessment|governance controls",
            ),
            (
                "Algorithmic Fairness",
                "算法公平",
                "measuring and reducing unfair model outcomes",
                "group fairness|intersectional analysis|bias mitigation|fairness monitoring",
            ),
            (
                "Explainable Artificial Intelligence",
                "可解释人工智能",
                "explanations that support real decisions and audits",
                "feature attribution|concept explanations|counterfactual explanations|human evaluation",
            ),
            (
                "Privacy-Preserving Learning",
                "隐私保护学习",
                "learning from sensitive data with measurable privacy",
                "differential privacy|federated learning|secure aggregation|privacy accounting",
            ),
            (
                "Secure Artificial Intelligence Systems",
                "人工智能系统安全",
                "defending models, data, and pipelines from attack",
                "adversarial robustness|data poisoning|model extraction|supply-chain security",
            ),
            (
                "Model Risk Management",
                "模型风险管理",
                "operational controls for model inventories and lifecycle risk",
                "model validation|risk appetite|change control|performance drift",
            ),
            (
                "Artificial Intelligence Assurance and Audit",
                "人工智能保障与审计",
                "independent evidence for trustworthy AI systems",
                "audit trails|control testing|assurance cases|third-party assessment",
            ),
            (
                "Human-Centered Artificial Intelligence",
                "以人为本的人工智能",
                "AI systems designed around human goals and capabilities",
                "participatory design|human oversight|accessibility|trustworthy interaction",
            ),
            (
                "Digital Rights and Artificial Intelligence",
                "数字权利与人工智能",
                "the impact of AI on rights, agency, and due process",
                "automated decisions|consent|content moderation|procedural fairness",
            ),
        ],
    ),
    (
        "Data Infrastructure and Operations",
        "数据、基础设施与运维",
        [
            (
                "Data-Centric Artificial Intelligence",
                "数据中心人工智能",
                "improving systems through data quality and curation",
                "dataset diagnostics|label quality|data selection|synthetic data",
            ),
            (
                "Machine Learning Operations",
                "机器学习运维",
                "reliable delivery and monitoring of machine learning systems",
                "feature stores|continuous delivery|observability|rollback strategies",
            ),
            (
                "Artificial Intelligence Systems Engineering",
                "人工智能系统工程",
                "end-to-end architecture for production AI services",
                "service reliability|system decomposition|capacity planning|failure isolation",
            ),
            (
                "Efficient Model Inference",
                "高效模型推理",
                "lower-latency and lower-cost model serving",
                "quantization|speculative decoding|batch scheduling|cache management",
            ),
            (
                "Edge Intelligence",
                "边缘智能",
                "learning and inference on constrained devices",
                "tiny models|on-device adaptation|sensor fusion|energy budgeting",
            ),
            (
                "Artificial Intelligence Accelerators",
                "人工智能加速器",
                "hardware and compilers for machine learning workloads",
                "tensor processors|memory bandwidth|kernel fusion|compiler scheduling",
            ),
            (
                "Distributed Training",
                "分布式训练",
                "scalable training across heterogeneous compute clusters",
                "parallelism strategies|checkpointing|network topology|fault tolerance",
            ),
            (
                "Cloud Artificial Intelligence Platforms",
                "云端人工智能平台",
                "multi-tenant platforms for governed AI development",
                "resource isolation|platform APIs|cost controls|model registries",
            ),
            (
                "Open-Source Artificial Intelligence Engineering",
                "开源人工智能工程",
                "sustainable engineering practices for open AI ecosystems",
                "dependency governance|community maintenance|release engineering|security response",
            ),
            (
                "Sustainable Artificial Intelligence Computing",
                "可持续人工智能计算",
                "reducing the environmental cost of AI workloads",
                "energy measurement|carbon-aware scheduling|hardware utilization|efficient training",
            ),
        ],
    ),
    (
        "Life and Health Sciences",
        "生命与健康科学",
        [
            (
                "Medical Imaging Artificial Intelligence",
                "医学影像人工智能",
                "clinical imaging analysis with robust validation",
                "radiology foundation models|segmentation|domain shift|reader studies",
            ),
            (
                "Clinical Decision Intelligence",
                "临床决策智能",
                "evidence-aware support for clinical decisions",
                "risk prediction|clinical pathways|uncertainty communication|workflow integration",
            ),
            (
                "Digital Health Systems",
                "数字健康系统",
                "safe digital tools for continuous care",
                "remote monitoring|digital therapeutics|patient engagement|interoperability",
            ),
            (
                "Computational Biology",
                "计算生物学",
                "computational models of biological systems",
                "protein networks|single-cell analysis|biological simulation|phenotype prediction",
            ),
            (
                "Genomics and Artificial Intelligence",
                "基因组学与人工智能",
                "learning from genomic and multi-omic data",
                "variant interpretation|gene regulation|population genomics|multi-omics integration",
            ),
            (
                "Drug Discovery Intelligence",
                "药物发现智能",
                "AI methods for target discovery and molecular design",
                "molecule generation|binding prediction|target prioritization|ADMET modeling",
            ),
            (
                "Precision Medicine Artificial Intelligence",
                "精准医疗人工智能",
                "personalized prediction and treatment selection",
                "patient stratification|treatment response|biomarker discovery|clinical utility",
            ),
            (
                "Public Health Analytics",
                "公共卫生分析",
                "population-scale intelligence for health planning",
                "outbreak forecasting|health inequity|surveillance systems|resource allocation",
            ),
            (
                "Neuroscience and Artificial Intelligence",
                "神经科学与人工智能",
                "connections between neural computation and machine intelligence",
                "neural decoding|brain-inspired learning|connectomics|cognitive models",
            ),
            (
                "Bioinformatics Systems",
                "生物信息学系统",
                "reproducible infrastructure for biological data analysis",
                "workflow engines|reference databases|sequence analysis|provenance tracking",
            ),
        ],
    ),
]
GROUPS.extend(
    [
        (
            "Physical Sciences and Engineering",
            "物理科学与工程",
            [
                (
                    "Artificial Intelligence for Materials Discovery",
                    "材料发现人工智能",
                    "accelerating the design and validation of new materials",
                    "property prediction|crystal generation|active learning labs|materials databases",
                ),
                (
                    "Computational Chemistry Artificial Intelligence",
                    "计算化学人工智能",
                    "machine learning for molecular simulation and reaction science",
                    "potential energy surfaces|reaction prediction|molecular dynamics|quantum chemistry surrogates",
                ),
                (
                    "Physics-Informed Learning",
                    "物理信息学习",
                    "models constrained by physical laws and scientific structure",
                    "neural operators|conservation laws|inverse problems|scientific priors",
                ),
                (
                    "Scientific Computing Intelligence",
                    "科学计算智能",
                    "AI-enhanced numerical methods and simulation workflows",
                    "surrogate modeling|adaptive meshing|solver acceleration|error estimation",
                ),
                (
                    "Engineering Design Artificial Intelligence",
                    "工程设计人工智能",
                    "computational support for multidisciplinary engineering design",
                    "topology optimization|generative design|requirements reasoning|design verification",
                ),
                (
                    "Manufacturing Intelligence",
                    "制造智能",
                    "data-driven quality, planning, and automation in manufacturing",
                    "process control|defect detection|production scheduling|digital thread",
                ),
                (
                    "Energy Systems Artificial Intelligence",
                    "能源系统人工智能",
                    "intelligent planning and control for modern energy systems",
                    "grid forecasting|demand response|renewable integration|storage optimization",
                ),
                (
                    "Quantum Machine Learning",
                    "量子机器学习",
                    "learning methods at the boundary of quantum and classical computing",
                    "variational circuits|quantum kernels|error mitigation|hybrid optimization",
                ),
                (
                    "Astronomy and Space Artificial Intelligence",
                    "天文与空间人工智能",
                    "AI for observing and operating in space",
                    "transient detection|telescope scheduling|spacecraft autonomy|exoplanet analysis",
                ),
                (
                    "Geoscience Intelligence",
                    "地球科学智能",
                    "learning from geological, seismic, and subsurface data",
                    "seismic interpretation|mineral exploration|subsurface modeling|geohazard assessment",
                ),
            ],
        ),
        (
            "Earth Climate and Sustainability",
            "地球、气候与可持续发展",
            [
                (
                    "Climate Intelligence",
                    "气候智能",
                    "data-driven understanding and prediction of climate systems",
                    "climate emulation|extreme events|downscaling|adaptation analytics",
                ),
                (
                    "Earth Observation Artificial Intelligence",
                    "地球观测人工智能",
                    "learning from satellite and remote-sensing data",
                    "land-cover mapping|change detection|foundation models|sensor harmonization",
                ),
                (
                    "Biodiversity Analytics",
                    "生物多样性分析",
                    "monitoring species, habitats, and ecosystem change",
                    "species recognition|acoustic ecology|habitat modeling|conservation prioritization",
                ),
                (
                    "Ocean Intelligence",
                    "海洋智能",
                    "AI for ocean observation, forecasting, and stewardship",
                    "ocean circulation|marine robotics|plankton imaging|illegal fishing detection",
                ),
                (
                    "Agricultural Artificial Intelligence",
                    "农业人工智能",
                    "precision and resilient decision-making for agriculture",
                    "crop monitoring|yield forecasting|farm robotics|pest detection",
                ),
                (
                    "Food Systems Intelligence",
                    "食品系统智能",
                    "safer and more sustainable food supply networks",
                    "food quality|demand forecasting|cold chains|alternative proteins",
                ),
                (
                    "Water Resources Artificial Intelligence",
                    "水资源人工智能",
                    "planning and monitoring water systems under uncertainty",
                    "flood forecasting|water quality|network leakage|reservoir operations",
                ),
                (
                    "Disaster Risk Intelligence",
                    "灾害风险智能",
                    "anticipating hazards and supporting emergency response",
                    "rapid damage mapping|wildfire prediction|evacuation planning|crisis information",
                ),
                (
                    "Environmental Monitoring",
                    "环境监测智能",
                    "continuous measurement and interpretation of environmental change",
                    "air quality|sensor networks|pollution attribution|ecological indicators",
                ),
                (
                    "Circular Economy Artificial Intelligence",
                    "循环经济人工智能",
                    "AI-enabled reuse, repair, and resource efficiency",
                    "waste sorting|material passports|reverse logistics|lifecycle optimization",
                ),
            ],
        ),
        (
            "Society Policy and Education",
            "社会、政策与教育",
            [
                (
                    "Artificial Intelligence Policy and Regulation",
                    "人工智能政策与监管",
                    "evidence for practical AI policy and regulation",
                    "regulatory sandboxes|standards alignment|policy evaluation|cross-border governance",
                ),
                (
                    "Computational Social Science",
                    "计算社会科学",
                    "computational analysis of social behavior and institutions",
                    "social networks|collective behavior|survey augmentation|computational ethnography",
                ),
                (
                    "Education and Learning Artificial Intelligence",
                    "教育与学习人工智能",
                    "responsible AI for teaching, assessment, and learning support",
                    "intelligent tutoring|learning analytics|assessment validity|teacher copilots",
                ),
                (
                    "Legal Intelligence",
                    "法律智能",
                    "AI systems for legal research and professional workflows",
                    "case retrieval|contract analysis|legal reasoning|access to justice",
                ),
                (
                    "Public Sector Artificial Intelligence",
                    "公共部门人工智能",
                    "accountable AI services for public administration",
                    "benefit delivery|public procurement|case management|service accessibility",
                ),
                (
                    "Urban Intelligence",
                    "城市智能",
                    "data-informed planning and operation of cities",
                    "mobility modeling|urban digital twins|housing analytics|public-space sensing",
                ),
                (
                    "Development and Humanitarian Artificial Intelligence",
                    "发展与人道主义人工智能",
                    "AI for development programs and humanitarian action",
                    "needs assessment|cash assistance|crisis mapping|program evaluation",
                ),
                (
                    "Labor and the Future of Work",
                    "劳动与未来工作",
                    "how AI changes jobs, skills, and organizations",
                    "task exposure|skills transition|workplace augmentation|job quality",
                ),
                (
                    "Social Media Research",
                    "社交媒体研究",
                    "rigorous study of online platforms and digital communities",
                    "misinformation diffusion|online communities|platform governance|content dynamics",
                ),
                (
                    "Science Communication Artificial Intelligence",
                    "科学传播人工智能",
                    "AI tools and evidence for public science communication",
                    "plain-language generation|evidence visualization|audience trust|misinformation response",
                ),
            ],
        ),
        (
            "Economy Industry and Operations",
            "经济、产业与运营",
            [
                (
                    "Financial Intelligence",
                    "金融智能",
                    "machine intelligence for markets, risk, and investment research",
                    "market forecasting|portfolio risk|alternative data|financial language models",
                ),
                (
                    "Financial Technology Artificial Intelligence",
                    "金融科技人工智能",
                    "AI-native financial products and infrastructure",
                    "fraud detection|credit decisioning|payments intelligence|regulatory technology",
                ),
                (
                    "Business Analytics and Artificial Intelligence",
                    "商业分析与人工智能",
                    "decision intelligence for firms and organizations",
                    "causal analytics|executive dashboards|scenario planning|decision automation",
                ),
                (
                    "Operations Research Intelligence",
                    "运筹智能",
                    "learning-enhanced optimization for complex operations",
                    "combinatorial optimization|simulation optimization|stochastic planning|decision policies",
                ),
                (
                    "Supply Chain Artificial Intelligence",
                    "供应链人工智能",
                    "resilient planning across supply networks",
                    "inventory optimization|supplier risk|demand sensing|logistics routing",
                ),
                (
                    "Retail Intelligence",
                    "零售智能",
                    "responsible personalization and store operations",
                    "assortment planning|dynamic pricing|store analytics|customer journeys",
                ),
                (
                    "Marketing Science Artificial Intelligence",
                    "营销科学人工智能",
                    "measurement and optimization of customer engagement",
                    "incrementality testing|media mix modeling|customer lifetime value|creative analytics",
                ),
                (
                    "Insurance Analytics",
                    "保险分析智能",
                    "data-driven underwriting, claims, and risk prevention",
                    "claims automation|catastrophe risk|underwriting models|fraud investigation",
                ),
                (
                    "Cybersecurity Intelligence",
                    "网络安全智能",
                    "AI methods for cyber defense and threat operations",
                    "threat detection|malware analysis|security copilots|attack simulation",
                ),
                (
                    "Telecommunications Artificial Intelligence",
                    "通信人工智能",
                    "intelligent operation of communication networks",
                    "network optimization|traffic forecasting|fault localization|radio resource management",
                ),
            ],
        ),
        (
            "Creativity Media and Humanities",
            "创意、媒体与人文",
            [
                (
                    "Creative Artificial Intelligence and Design",
                    "创意人工智能与设计",
                    "human-AI co-creation in visual and product design",
                    "design ideation|creative control|co-creation workflows|design evaluation",
                ),
                (
                    "Generative Media",
                    "生成式媒体",
                    "responsible generation of image, video, and interactive media",
                    "video generation|image editing|provenance signals|content authenticity",
                ),
                (
                    "Music Intelligence",
                    "音乐智能",
                    "computational understanding and creation of music",
                    "music generation|audio source separation|performance analysis|music recommendation",
                ),
                (
                    "Computational Creativity",
                    "计算创造力",
                    "models and theories of machine-assisted creativity",
                    "creative search|novelty measurement|concept blending|human evaluation",
                ),
                (
                    "Digital Humanities Artificial Intelligence",
                    "数字人文人工智能",
                    "AI methods for historical and cultural scholarship",
                    "historical text mining|archive discovery|cultural analytics|scholarly annotation",
                ),
                (
                    "Cultural Heritage Intelligence",
                    "文化遗产智能",
                    "digital preservation and interpretation of cultural heritage",
                    "artifact reconstruction|heritage imaging|museum knowledge graphs|site monitoring",
                ),
                (
                    "Journalism and Artificial Intelligence",
                    "新闻业与人工智能",
                    "AI-supported reporting, verification, and newsroom practice",
                    "fact checking|investigative search|newsroom workflows|editorial accountability",
                ),
                (
                    "Publishing Technology",
                    "出版技术",
                    "technology for reliable scholarly and professional publishing",
                    "semantic production|content workflows|accessibility|digital preservation",
                ),
                (
                    "Game Intelligence",
                    "游戏智能",
                    "AI for game design, play, and player experience",
                    "procedural content|game agents|player modeling|adaptive difficulty",
                ),
                (
                    "Virtual Worlds and Artificial Intelligence",
                    "虚拟世界与人工智能",
                    "intelligent characters and systems for immersive environments",
                    "interactive agents|world simulation|spatial interaction|virtual economies",
                ),
            ],
        ),
        (
            "Research Practice and Frontier Studies",
            "科研实践与前沿研究",
            [
                (
                    "Artificial Intelligence for Scientific Discovery",
                    "科学发现人工智能",
                    "AI systems that accelerate hypothesis, experiment, and discovery",
                    "hypothesis generation|autonomous laboratories|scientific foundation models|discovery evaluation",
                ),
                (
                    "Automated Research Agents",
                    "自动化科研智能体",
                    "agents that support multi-step research workflows",
                    "experiment planning|tool orchestration|research memory|human supervision",
                ),
                (
                    "Literature Discovery Systems",
                    "文献发现系统",
                    "evidence retrieval and synthesis across scholarly literature",
                    "systematic search|citation networks|evidence synthesis|living reviews",
                ),
                (
                    "Research Integrity and Artificial Intelligence",
                    "科研诚信与人工智能",
                    "responsible use of AI in research and publication",
                    "authorship disclosure|fabrication detection|image integrity|responsibility frameworks",
                ),
                (
                    "Peer Review Intelligence",
                    "同行评审智能",
                    "AI support for fair and rigorous peer review",
                    "reviewer matching|method checking|review quality|bias monitoring",
                ),
                (
                    "Scholarly Knowledge Infrastructure",
                    "学术知识基础设施",
                    "open infrastructure for scholarly records and connections",
                    "persistent identifiers|research graphs|metadata exchange|institutional repositories",
                ),
                (
                    "Open Science Automation",
                    "开放科学自动化",
                    "automation for transparent and accessible research",
                    "data publication|open workflows|license checking|repository integration",
                ),
                (
                    "Reproducible Artificial Intelligence Research",
                    "可复现人工智能研究",
                    "methods and infrastructure for repeatable AI evidence",
                    "experiment tracking|environment capture|statistical reporting|replication studies",
                ),
                (
                    "Benchmarking and Evaluation",
                    "基准测试与评估",
                    "measurement systems that reveal real model capability",
                    "benchmark validity|contamination detection|human baselines|longitudinal evaluation",
                ),
                (
                    "Future Intelligence Studies",
                    "未来智能研究",
                    "interdisciplinary study of emerging forms of intelligence",
                    "capability forecasting|institutional scenarios|human adaptation|long-term governance",
                ),
            ],
        ),
    ]
)

ARTICLE_TYPE_SEQUENCE = ["ai_article", "review", "opinion", "news", "editorial"]
ARTICLE_STATUS_SEQUENCE = ["review", "review", "draft", "review", "draft"]
TITLE_PATTERNS = [
    "Measuring {c1} under real-world constraints",
    "A practical field guide to {c2}",
    "When {c3} meets {c4}: evidence and trade-offs",
    "Operational lessons from {c1} and {c3}",
    "A 2026 research agenda for {focus}",
    "Beyond prototypes: evaluating {c1} in practice",
    "Designing trustworthy {c2} workflows",
    "What {c3} changes about {focus}",
    "Reproducible evidence for {c4}",
    "Five open questions in {focus}",
]
SURNAMES = [
    "Chen",
    "Wang",
    "Li",
    "Zhang",
    "Liu",
    "Yang",
    "Huang",
    "Zhao",
    "Wu",
    "Zhou",
    "Xu",
    "Sun",
    "Ma",
    "Gao",
    "Lin",
    "He",
    "Guo",
    "Tang",
    "Luo",
    "Deng",
    "Smith",
    "Garcia",
    "Patel",
    "Kim",
    "Nguyen",
    "Brown",
    "Martin",
    "Wilson",
    "Taylor",
    "Anderson",
]
GIVEN_NAMES = [
    "Ming",
    "Yue",
    "Jia",
    "Rui",
    "Lan",
    "Tao",
    "Nora",
    "Ethan",
    "Maya",
    "Leo",
    "Sofia",
    "Arjun",
    "Hana",
    "Noah",
    "Amelia",
    "Omar",
    "Elena",
    "Lucas",
    "Iris",
    "Daniel",
]
AI_COAUTHORS = [
    "Atlas Research Assistant",
    "Mosaic Evidence Agent",
    "Orion Analysis Model",
    "Sage Literature Copilot",
    "Verity Methods Assistant",
]


def slugify(text: str) -> str:
    value = text.lower().replace("artificial intelligence", "ai")
    value = value.replace("and", " ")
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return re.sub(r"-+", "-", value)


def flatten_catalog() -> list[dict]:
    records = []
    for group_index, (group_en, group_cn, items) in enumerate(GROUPS, start=1):
        if len(items) != 10:
            raise ValueError(f"{group_en} must contain exactly 10 journals")
        for item_index, (name, name_cn, focus, concept_text) in enumerate(
            items, start=1
        ):
            concepts = [part.strip() for part in concept_text.split("|")]
            if len(concepts) != 4:
                raise ValueError(f"{name} must define exactly four concepts")
            records.append(
                {
                    "group_en": group_en,
                    "group_cn": group_cn,
                    "group_index": group_index,
                    "item_index": item_index,
                    "name": name,
                    "name_cn": name_cn,
                    "focus": focus,
                    "concepts": concepts,
                }
            )
    if len(records) != 120:
        raise ValueError(f"Expected 120 journals, got {len(records)}")
    return records


def pick_font(size: int, bold: bool = False):
    candidates = [
        Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
        Path(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
            if bold
            else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        ),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def palette(
    index: int,
) -> tuple[tuple[int, int, int], tuple[int, int, int], tuple[int, int, int]]:
    hue = (index * 47) % 360

    def hsl(hue_value, saturation, lightness):
        import colorsys

        r, g, b = colorsys.hls_to_rgb(
            hue_value / 360, lightness / 100, saturation / 100
        )
        return tuple(round(v * 255) for v in (r, g, b))

    return (
        hsl(hue, 48, 18),
        hsl((hue + 28) % 360, 65, 52),
        hsl((hue + 180) % 360, 48, 92),
    )


def wrapped_lines(draw, text: str, font, max_width: int) -> list[str]:
    words = text.split()
    lines, current = [], ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def make_cover(record: dict, index: int) -> bytes:
    bg, accent, pale = palette(index)
    image = Image.new("RGB", (900, 1200), bg)
    draw = ImageDraw.Draw(image)
    title_font = pick_font(54, bold=True)
    cn_font = pick_font(34, bold=True)
    small_font = pick_font(24)
    micro_font = pick_font(18)
    draw.rectangle((0, 0, 900, 22), fill=accent)
    draw.text((64, 66), "AI AUTHOR FORUM", font=small_font, fill=pale)
    draw.text((64, 116), record["group_en"].upper(), font=micro_font, fill=accent)
    y = 230
    for line in wrapped_lines(draw, record["name"], title_font, 770):
        draw.text((64, y), line, font=title_font, fill=(255, 255, 255))
        y += 70
    draw.text((64, y + 18), record["name_cn"], font=cn_font, fill=pale)
    y += 105
    draw.line((64, y, 836, y), fill=accent, width=4)
    y += 38
    for concept in record["concepts"]:
        draw.ellipse((67, y + 8, 79, y + 20), fill=accent)
        draw.text((96, y), concept, font=small_font, fill=pale)
        y += 47
    draw.rectangle((64, 1028, 836, 1118), outline=accent, width=3)
    draw.text(
        (91, 1053),
        f"Curated import journal #{index:03d}",
        font=small_font,
        fill=(255, 255, 255),
    )
    stream = BytesIO()
    image.save(stream, format="PNG", optimize=True)
    return stream.getvalue()


def make_metrics(record: dict, index: int) -> bytes:
    bg, accent, pale = palette(index)
    image = Image.new("RGB", (1280, 720), pale)
    draw = ImageDraw.Draw(image)
    title_font = pick_font(42, bold=True)
    label_font = pick_font(24, bold=True)
    value_font = pick_font(42, bold=True)
    body_font = pick_font(22)
    draw.rectangle((0, 0, 1280, 88), fill=bg)
    draw.text(
        (48, 22),
        f"{record['name']} · import profile",
        font=title_font,
        fill=(255, 255, 255),
    )
    seed = int(hashlib.sha256(record["name"].encode("utf-8")).hexdigest()[:8], 16)
    metrics = [
        ("Sample articles", 5),
        ("Editorial themes", 4),
        ("Planned annual papers", 48 + seed % 73),
        ("Review target", f"{7 + seed % 15} days"),
    ]
    for pos, (label, value) in enumerate(metrics):
        x = 48 + (pos % 2) * 610
        y = 140 + (pos // 2) * 220
        draw.rounded_rectangle(
            (x, y, x + 560, y + 166),
            radius=18,
            fill=(255, 255, 255),
            outline=accent,
            width=3,
        )
        draw.text((x + 28, y + 24), label, font=label_font, fill=bg)
        draw.text((x + 28, y + 70), str(value), font=value_font, fill=accent)
    draw.text(
        (48, 630),
        "Synthetic metrics for import acceptance testing; not production statistics.",
        font=body_font,
        fill=bg,
    )
    stream = BytesIO()
    image.save(stream, format="PNG", optimize=True)
    return stream.getvalue()


def make_article_banner(record: dict, index: int) -> bytes:
    bg, accent, pale = palette(index)
    image = Image.new("RGB", (1280, 720), bg)
    draw = ImageDraw.Draw(image)
    title_font = pick_font(48, bold=True)
    body_font = pick_font(27)
    micro_font = pick_font(20)
    draw.polygon([(0, 0), (1280, 0), (1280, 270), (0, 560)], fill=accent)
    draw.text((64, 58), "AI AUTHOR FORUM · JOURNAL ARTICLE", font=micro_font, fill=bg)
    y = 322
    for line in wrapped_lines(draw, record["name"], title_font, 1040):
        draw.text((64, y), line, font=title_font, fill=(255, 255, 255))
        y += 62
    draw.text((64, 628), record["focus"], font=body_font, fill=pale)
    stream = BytesIO()
    image.save(stream, format="PNG", optimize=True)
    return stream.getvalue()


def article_title(record: dict, journal_index: int, article_index: int) -> str:
    c1, c2, c3, c4 = record["concepts"]
    pattern = TITLE_PATTERNS[
        (journal_index * 3 + article_index - 1) % len(TITLE_PATTERNS)
    ]
    return pattern.format(c1=c1, c2=c2, c3=c3, c4=c4, focus=record["focus"])


def article_abstract(record: dict, journal_index: int, article_index: int) -> str:
    c1, c2, c3, c4 = record["concepts"]
    variants = [
        f"This study defines a reproducible evaluation plan for {c1} in {record['focus']}, comparing operational constraints, failure modes, and evidence quality.",
        f"This review maps current practice in {c2}, identifies where reported gains fail to transfer, and proposes a decision checklist for teams working on {record['focus']}.",
        f"Using {c3} and {c4} as contrasting cases, this perspective explains which assumptions matter most and where additional validation is needed.",
        f"This field report turns lessons from {c1} and {c3} into concrete controls, measurement steps, and hand-off criteria for multidisciplinary teams.",
        f"The editorial sets five testable priorities for {record['focus']}, with emphasis on transparent datasets, comparable baselines, and accountable human oversight.",
    ]
    return variants[article_index - 1]


def article_body(record: dict, title: str, abstract: str, article_index: int) -> str:
    concepts = record["concepts"]
    question = [
        f"How should {concepts[0]} be measured before a system is trusted outside a laboratory?",
        f"Which evidence distinguishes durable progress in {concepts[1]} from benchmark-specific improvement?",
        f"What changes when {concepts[2]} is evaluated alongside {concepts[3]}?",
        f"Which operational controls make {concepts[0]} and {concepts[2]} easier to reproduce?",
        f"What research commitments would move {record['focus']} from isolated demonstrations to cumulative evidence?",
    ][article_index - 1]
    method = [
        "We define a task inventory, a baseline matrix, and three stress conditions. Each result is reported with uncertainty, resource use, and a documented failure analysis.",
        "We compare representative methods by assumptions, data requirements, evaluation design, and deployment constraints rather than by a single headline score.",
        "We organize evidence into controlled experiments, field observations, and unresolved counterexamples, then trace how conclusions change across settings.",
        "We use an operational case-study protocol covering inputs, decision rights, monitoring, escalation, rollback, and post-deployment review.",
        "We synthesize open problems into short-term replication work, medium-term shared infrastructure, and long-term governance questions.",
    ][article_index - 1]
    implication = [
        f"Teams should treat {concepts[0]} as a measured capability with explicit operating limits, not as a binary feature.",
        f"Procurement and research reviews should require evidence that {concepts[1]} transfers across datasets, users, and time periods.",
        f"The comparison shows that progress in {concepts[2]} can hide regressions in {concepts[3]}; both need joint reporting.",
        f"Clear ownership and rollback criteria are as important as model quality when operationalizing {concepts[0]}.",
        f"A shared agenda can reduce duplicated experiments and make negative results useful across the {record['name']} community.",
    ][article_index - 1]
    return f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>{escape(title)}</title></head>
<body>
  <p><strong>Journal:</strong> {escape(record['name'])} ({escape(record['name_cn'])})</p>
  <h2>Abstract</h2><p>{escape(abstract)}</p>
  <h2>Research question</h2><p>{escape(question)}</p>
  <h2>Approach</h2><p>{escape(method)}</p>
  <h2>Practical implications</h2><p>{escape(implication)}</p>
  <h2>Reproducibility checklist</h2>
  <ul><li>Publish data and sampling assumptions.</li><li>Report baselines, uncertainty, and failed cases.</li><li>Record compute, software, and human review steps.</li><li>Separate import status from editorial approval and placement.</li></ul>
</body></html>"""


def authors_for(journal_index: int, article_index: int) -> str:
    first = f"{GIVEN_NAMES[(journal_index + article_index) % len(GIVEN_NAMES)]} {SURNAMES[(journal_index * 2 + article_index) % len(SURNAMES)]}"
    second = f"{GIVEN_NAMES[(journal_index * 3 + article_index + 7) % len(GIVEN_NAMES)]} {SURNAMES[(journal_index + article_index * 5 + 11) % len(SURNAMES)]}"
    return f"{first}; {second}"


def make_workbook(headers: list[str], rows: list[dict], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    header_fill = PatternFill("solid", fgColor="17365D")
    alt_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row_index in range(2, ws.max_row + 1):
        if row_index % 2 == 0:
            for cell in ws[row_index]:
                cell.fill = alt_fill
        for cell in ws[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "journal_name": 38,
        "journal_name_cn": 24,
        "slug": 34,
        "seo_title": 42,
        "seo_description": 58,
        "homepage_intro": 64,
        "notes": 46,
        "title": 58,
        "authors": 32,
        "abstract": 72,
        "keywords": 52,
        "body_html": 28,
        "cover_image": 42,
        "metrics_image": 42,
        "static_site_path": 48,
        "static_output_path": 56,
    }
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(
            header, min(max(len(header) + 3, 14), 28)
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    ws.sheet_view.showGridLines = False
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate(output: Path, extracted_dir: Path | None = None) -> dict:
    records = flatten_catalog()
    output.parent.mkdir(parents=True, exist_ok=True)
    generated_at = datetime(2026, 7, 27, 9, 0, tzinfo=timezone(timedelta(hours=8)))
    first_publication = datetime(2025, 1, 6, 9, 0, tzinfo=timezone(timedelta(hours=8)))
    journal_rows: list[dict] = []
    article_rows: list[dict] = []
    assets: dict[str, bytes] = {}

    for journal_index, record in enumerate(records, start=1):
        journal_slug = slugify(record["name"])
        first_letter = next((c for c in record["name"].upper() if "A" <= c <= "Z"), "#")
        c1, c2, c3, c4 = record["concepts"]
        cover_path = f"media/journals/{journal_slug}/cover.png"
        metrics_path = f"media/journals/{journal_slug}/metrics.png"
        banner_path = f"media/journals/{journal_slug}/article-banner.png"
        assets[cover_path] = make_cover(record, journal_index)
        assets[metrics_path] = make_metrics(record, journal_index)
        assets[banner_path] = make_article_banner(record, journal_index)
        journal_rows.append(
            {
                "journal_name": record["name"],
                "journal_name_cn": record["name_cn"],
                "slug": journal_slug,
                "az_group": first_letter,
                "status": "active",
                "sort_order": journal_index,
                "seo_title": f"{record['name']} | AI Author Forum",
                "seo_description": f"Research on {record['focus']}, including {c1}, {c2}, {c3}, and {c4}.",
                "homepage_intro": f"<p><strong>{record['name_cn']}</strong> focuses on {record['focus']}. The test issue includes concrete work on {c1}, {c2}, {c3}, and {c4}.</p>",
                "cover_image": cover_path,
                "metrics_image": metrics_path,
                "static_site_path": f"/journals/{journal_slug}/index.html",
                "target_article_count": 5,
                "notes": f"Curated import fixture; collection={record['group_en']}; generated 2026-07-27; five draft/review articles; synthetic test data only.",
            }
        )
        for article_index in range(1, 6):
            title = article_title(record, journal_index, article_index)
            article_slug = slugify(title)
            abstract = article_abstract(record, journal_index, article_index)
            publication_date = first_publication + timedelta(
                days=(journal_index - 1) * 4 + article_index - 1
            )
            article_rows.append(
                {
                    "journal_slug": journal_slug,
                    "title": title,
                    "slug": article_slug,
                    "article_type": ARTICLE_TYPE_SEQUENCE[article_index - 1],
                    "authors": authors_for(journal_index, article_index),
                    "ai_co_authors": AI_COAUTHORS[
                        (journal_index + article_index) % len(AI_COAUTHORS)
                    ],
                    "abstract": abstract,
                    "keywords": ", ".join(
                        [record["focus"], *record["concepts"], "reproducibility"]
                    ),
                    "publication_date": publication_date.isoformat(),
                    "status": ARTICLE_STATUS_SEQUENCE[article_index - 1],
                    "sort_order": article_index,
                    "is_pinned": article_index == 1,
                    "cover_image": banner_path,
                    "body_html": article_body(record, title, abstract, article_index),
                    "html_file": "",
                    "build_version": "curated-120-20260727",
                    "static_output_path": f"/articles/{article_slug}/index.html",
                    "primary_category_code": "",
                    "primary_category_path": "",
                    "related_category_codes": "",
                    "related_category_paths": "",
                    "main_site_slot": "",
                    "main_site_slot_name": "",
                    "main_site_slot_title": "",
                    "main_site_slot_summary": "",
                    "main_site_slot_order": "",
                    "main_site_slot_pinned": False,
                    "journal_slot": "",
                    "journal_slot_name": "",
                    "journal_slot_title": "",
                    "journal_slot_summary": "",
                    "journal_slot_order": "",
                    "journal_slot_pinned": False,
                    "notes": f"Concrete sample article {article_index}/5 for {record['name']}; imported content must still pass Wagtail review and placement.",
                }
            )

    journal_slugs = [row["slug"] for row in journal_rows]
    article_keys = [(row["journal_slug"], row["slug"]) for row in article_rows]
    if len(set(journal_slugs)) != 120:
        raise ValueError("Journal slugs are not unique")
    if len(set(article_keys)) != 600:
        raise ValueError("Article keys are not unique within journals")
    journal_bytes = make_workbook(JOURNAL_HEADERS, journal_rows, "journals")
    article_bytes = make_workbook(ARTICLE_HEADERS, article_rows, "articles")
    manifest = {
        "package": output.name,
        "schema": "ai-author-forum-journal-import/v1",
        "generated_at": generated_at.isoformat(),
        "journal_count": len(journal_rows),
        "article_count": len(article_rows),
        "articles_per_journal": 5,
        "asset_count": len(assets),
        "collections": [
            {"name": group[0], "name_cn": group[1], "journal_count": len(group[2])}
            for group in GROUPS
        ],
        "workflow_note": "Imported ArticlePage records enter draft moderation; approval and ArticlePlacement remain separate.",
    }
    readme = """AI Author Forum 120 个差异化子期刊一键导入测试包

内容：
- journals.xlsx：120 个子期刊，每刊具有独立中英文名称、研究范围、SEO、简介、封面和指标图。
- articles.xlsx：600 篇文章，每刊 5 篇，具有不同标题、类型、作者、摘要、关键词和正文。
- media/：每刊独立封面、指标图、文章横幅，共 360 个素材。
- manifest.json：数量和分组清单。

导入规则：
1. 在 Wagtail 后台“子期刊导入”上传整个 ZIP，不要只上传内部 Excel。
2. 先执行预检；预期结果为 120 条期刊、600 条文章、0 条失败。
3. 文章导入后仍进入草稿审核流程；审核通过也必须配置 ArticlePlacement 才能上前台。
4. 本包全部为差异化合成测试数据，不代表真实期刊统计或正式出版内容。
"""
    with zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as zf:
        zf.writestr("journals.xlsx", journal_bytes)
        zf.writestr("articles.xlsx", article_bytes)
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        zf.writestr("README.zh-CN.txt", readme)
        for path, content in assets.items():
            zf.writestr(path, content)

    if extracted_dir:
        extracted_dir.mkdir(parents=True, exist_ok=True)
        (extracted_dir / "journals-120-rich.xlsx").write_bytes(journal_bytes)
        (extracted_dir / "articles-600-rich.xlsx").write_bytes(article_bytes)
        (extracted_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (extracted_dir / "README.zh-CN.txt").write_text(readme, encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build the curated 120-journal one-click import package."
    )
    parser.add_argument(
        "--output",
        default="output/ai-author-forum-120-journals-rich-import-20260727.zip",
    )
    parser.add_argument(
        "--extracted-dir", default="output/ai-author-forum-120-journals-rich-preview"
    )
    args = parser.parse_args()
    manifest = generate(
        Path(args.output), Path(args.extracted_dir) if args.extracted_dir else None
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
