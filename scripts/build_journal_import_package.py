from __future__ import annotations

import argparse
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from string import Template

from openpyxl import Workbook

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:  # pragma: no cover - pillow is part of the project stack
    raise SystemExit("Pillow is required to build the sample package.") from exc


JOURNAL_HEADERS = [
    "journal_name",
    "journal_name_cn",
    "slug",
    "az_group",
    "status",
    "sort_order",
    "cover_image",
    "metrics_image",
    "seo_title",
    "seo_description",
    "homepage_intro",
    "static_site_path",
    "target_article_count",
    "notes",
]

ARTICLE_HEADERS = [
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "status",
    "sort_order",
    "is_pinned",
    "build_version",
    "notes",
    "static_output_path",
    "body_html",
    "cover_image",
    "main_site_slot",
    "main_site_slot_name",
    "main_site_slot_layout",
    "main_site_slot_order",
    "main_site_slot_pinned",
    "main_site_slot_title",
    "main_site_slot_summary",
    "journal_slot",
    "journal_slot_name",
    "journal_slot_layout",
    "journal_slot_order",
    "journal_slot_pinned",
    "journal_slot_title",
    "journal_slot_summary",
]


TOPICS = [
    "editorial AI systems",
    "model governance",
    "research productivity",
    "literature discovery",
    "agentic workflows",
    "review automation",
    "knowledge graphs",
    "multimodal reading",
    "content operations",
    "static publishing",
    "trust and safety",
    "scientific communication",
]


def slugify_like(text: str) -> str:
    return (
        text.lower()
        .replace("&", " and ")
        .replace("/", " ")
        .replace("'", "")
        .replace(",", "")
        .replace(".", "")
        .replace("  ", " ")
        .strip()
        .replace(" ", "-")
    )


def make_png_bytes(
    label: str, background: tuple[int, int, int], foreground: tuple[int, int, int]
) -> bytes:
    image = Image.new("RGB", (1280, 720), background)
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    draw.rectangle((56, 56, 1224, 664), outline=foreground, width=6)
    draw.text((96, 100), "AI Author Forum", fill=foreground, font=font)
    draw.text((96, 150), label, fill=foreground, font=font)
    draw.text((96, 200), "Static import package asset", fill=foreground, font=font)
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def build_article_html(
    journal_name: str,
    article_title: str,
    abstract: str,
    keywords: str,
    source_slot: str,
) -> str:
    chips = "".join(
        f'<span class="chip">{kw.strip()}</span>'
        for kw in keywords.split(",")
        if kw.strip()
    )
    template = Template("""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>$article_title</title>
  <style>
    :root { color-scheme: light; }
    body {
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      background: #f5f7fb;
      color: #111827;
    }
    .page {
      max-width: 1080px;
      margin: 0 auto;
      padding: 32px 20px 56px;
    }
    header {
      border-bottom: 1px solid #d8dee9;
      padding-bottom: 18px;
      margin-bottom: 28px;
    }
    .eyebrow {
      text-transform: uppercase;
      letter-spacing: .08em;
      font-size: 12px;
      color: #6b7280;
      margin-bottom: 10px;
    }
    h1 {
      font-size: 40px;
      line-height: 1.12;
      margin: 0 0 12px;
    }
    .meta {
      color: #4b5563;
      font-size: 14px;
      display: grid;
      gap: 6px;
    }
    article {
      background: #fff;
      border: 1px solid #e5e7eb;
      border-radius: 14px;
      padding: 24px;
      box-shadow: 0 8px 24px rgba(15, 23, 42, .06);
    }
    h2 {
      font-size: 22px;
      margin: 28px 0 10px;
    }
    p {
      line-height: 1.7;
      margin: 0 0 14px;
    }
    .chip {
      display: inline-block;
      padding: 6px 10px;
      margin: 4px 8px 0 0;
      background: #eef2ff;
      color: #3730a3;
      border-radius: 999px;
      font-size: 12px;
    }
    .slot {
      border-left: 4px solid #0f766e;
      padding-left: 14px;
      color: #0f172a;
    }
  </style>
</head>
<body>
  <div class="page">
    <header>
      <div class="eyebrow">$journal_name</div>
      <h1>$article_title</h1>
      <div class="meta">
        <div>Static article source for $journal_name</div>
        <div>Source slot: $source_slot</div>
      </div>
    </header>
    <article>
      <h2>Abstract</h2>
      <p>$abstract</p>
      <h2>Why it matters</h2>
      <p>This static HTML page is generated offline and served directly, keeping the publishing path lightweight.</p>
      <h2>Keywords</h2>
      <div>$chips</div>
      <h2>Notes</h2>
      <p>Featured layout is fixed at build time, so the front end stays deterministic.</p>
    </article>
  </div>
</body>
</html>
""")
    return template.substitute(
        journal_name=journal_name,
        article_title=article_title,
        abstract=abstract,
        keywords=keywords,
        source_slot=source_slot,
        chips=chips,
    )


def generate_package(output: Path, journals: int, articles_per_journal: int) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)

    base_time = datetime(2026, 7, 18, 9, 0)
    journal_colors = [
        ((15, 23, 42), (255, 255, 255)),
        ((30, 41, 59), (255, 255, 255)),
        ((15, 118, 110), (255, 255, 255)),
    ]
    article_colors = [
        ((248, 250, 252), (15, 23, 42)),
        ((239, 246, 255), (30, 41, 59)),
        ((240, 253, 250), (17, 94, 89)),
    ]

    journal_rows = []
    article_rows = []
    asset_blobs: dict[str, bytes] = {}

    for index in range(1, 4):
        bg, fg = journal_colors[(index - 1) % len(journal_colors)]
        asset_blobs[f"media/shared/journal-cover-{index}.png"] = make_png_bytes(
            f"Journal cover {index}",
            bg,
            fg,
        )
        asset_blobs[f"media/shared/journal-metrics-{index}.png"] = make_png_bytes(
            f"Journal metrics {index}",
            (232, 240, 254),
            (37, 99, 235),
        )
        bg2, fg2 = article_colors[(index - 1) % len(article_colors)]
        asset_blobs[f"media/shared/article-cover-{index}.png"] = make_png_bytes(
            f"Article cover {index}",
            bg2,
            fg2,
        )

    for journal_index in range(1, journals + 1):
        letter = chr(65 + ((journal_index - 1) % 26))
        journal_name = f"AI Journal {journal_index:03d}"
        journal_slug = f"ai-journal-{journal_index:03d}"
        journal_rows.append(
            {
                "journal_name": journal_name,
                "journal_name_cn": f"AI Journal {journal_index:03d} CN",
                "slug": journal_slug,
                "az_group": letter,
                "status": "active",
                "sort_order": journal_index,
                "cover_image": f"media/shared/journal-cover-{((journal_index - 1) % 3) + 1}.png",
                "metrics_image": f"media/shared/journal-metrics-{((journal_index - 1) % 3) + 1}.png",
                "seo_title": f"{journal_name} | AI Author Forum",
                "seo_description": f"Static journal homepage for {journal_name}.",
                "homepage_intro": f"{journal_name} hosts {articles_per_journal} static article pages.",
                "static_site_path": f"/journals/{journal_slug}/index.html",
                "target_article_count": articles_per_journal,
                "notes": "Generated import package for one-click backend and static frontend tests.",
            }
        )

        for article_index in range(1, articles_per_journal + 1):
            global_article_index = (
                journal_index - 1
            ) * articles_per_journal + article_index
            topic = TOPICS[(article_index - 1) % len(TOPICS)]
            article_title = f"{journal_name}: {topic.title()} {article_index:03d}"
            article_slug = f"{slugify_like(topic)}-{article_index:03d}"
            abstract = (
                f"Article {article_index:03d} in {journal_name} focuses on {topic} "
                f"and keeps the publication path fully static."
            )
            keywords = f"{topic}, static html, backend import, front-end publishing"
            source_slot = (
                "main"
                if article_index == 1
                else "journal" if article_index == 2 else "article"
            )
            article_html = build_article_html(
                journal_name,
                article_title,
                abstract,
                keywords,
                source_slot,
            )
            row = {
                "journal_slug": journal_slug,
                "title": article_title,
                "slug": article_slug,
                "article_type": "ai_article",
                "authors": f"Editorial Team {journal_index:03d}",
                "ai_co_authors": "Nature-style generator",
                "abstract": abstract,
                "keywords": keywords,
                "publication_date": (
                    base_time + timedelta(days=global_article_index)
                ).isoformat(),
                "status": "published",
                "sort_order": article_index,
                "is_pinned": article_index == 1,
                "build_version": "generated-2026-07-18",
                "notes": "Static HTML article built for import and export verification.",
                "static_output_path": f"/journals/{journal_slug}/articles/{article_slug}/index.html",
                "body_html": article_html,
                "cover_image": f"media/shared/article-cover-{((article_index - 1) % 3) + 1}.png",
            }
            if article_index == 1:
                row.update(
                    {
                        "main_site_slot": f"home-feature-{journal_index:03d}",
                        "main_site_slot_name": f"Home feature {journal_index:03d}",
                        "main_site_slot_layout": "hero",
                        "main_site_slot_order": 1,
                        "main_site_slot_pinned": True,
                        "main_site_slot_title": article_title,
                        "main_site_slot_summary": abstract,
                        "journal_slot": f"journal-feature-{journal_index:03d}",
                        "journal_slot_name": f"Journal feature {journal_index:03d}",
                        "journal_slot_layout": "grid",
                        "journal_slot_order": 1,
                        "journal_slot_pinned": True,
                        "journal_slot_title": article_title,
                        "journal_slot_summary": abstract,
                    }
                )
            elif article_index == 2:
                row.update(
                    {
                        "journal_slot": f"journal-feature-{journal_index:03d}-secondary",
                        "journal_slot_name": f"Journal secondary {journal_index:03d}",
                        "journal_slot_layout": "list",
                        "journal_slot_order": 2,
                        "journal_slot_pinned": False,
                        "journal_slot_title": article_title,
                        "journal_slot_summary": abstract,
                    }
                )
            article_rows.append(row)

    journal_wb = Workbook(write_only=True)
    journal_ws = journal_wb.create_sheet()
    journal_ws.append(JOURNAL_HEADERS)
    for row in journal_rows:
        journal_ws.append([row.get(column, "") for column in JOURNAL_HEADERS])
    journal_buffer = BytesIO()
    journal_wb.save(journal_buffer)

    article_wb = Workbook(write_only=True)
    article_ws = article_wb.create_sheet()
    article_ws.append(ARTICLE_HEADERS)
    for row in article_rows:
        article_ws.append([row.get(column, "") for column in ARTICLE_HEADERS])
    article_buffer = BytesIO()
    article_wb.save(article_buffer)

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("journals.xlsx", journal_buffer.getvalue())
        zf.writestr("articles.xlsx", article_buffer.getvalue())
        for path, blob in asset_blobs.items():
            zf.writestr(path, blob)

    return output


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a synthetic journal import package."
    )
    parser.add_argument(
        "--output",
        default="output/journal-import-package.zip",
        help="Path to the zip package to create.",
    )
    parser.add_argument(
        "--journals", type=int, default=120, help="Number of journals to generate."
    )
    parser.add_argument(
        "--articles-per-journal",
        type=int,
        default=100,
        help="Number of articles to generate per journal.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = Path(args.output)
    generate_package(output, args.journals, args.articles_per_journal)
    print(
        f"Created {output} with {args.journals} journals and "
        f"{args.journals * args.articles_per_journal} articles."
    )


if __name__ == "__main__":
    main()
