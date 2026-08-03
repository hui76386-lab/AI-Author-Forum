#!/usr/bin/env python
"""Import selected generated article-package parts through the official workflow.

Examples::

    python scripts/import_article_package_range.py --parts 3-5
    python scripts/import_article_package_range.py --parts 1,3-5,12

The script is designed for parallel workers that receive non-overlapping part ranges.
For each package it runs preview_article_import -> confirm_article_import ->
execute_confirmed_article_import. A package is skipped only when all of its expected
canonical ArticlePage records already exist. The default generated packages contain
100 articles each.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ai_author_forum.settings.dev")

import django  # noqa: E402

django.setup()

from django.contrib.auth import get_user_model  # noqa: E402
from django.db import connection  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from ai_author_forum.articles.import_services import (  # noqa: E402
    ArticleImportContext,
    confirm_article_import,
    execute_confirmed_article_import,
    preview_article_import,
)
from ai_author_forum.articles.models import ArticlePage  # noqa: E402
from ai_author_forum.journals.models import (  # noqa: E402
    ImportJobStatus,
    ImportRowStatus,
)

DEFAULT_BATCH_CODE = "aaf1200-20260730"
DEFAULT_PACKAGE_DIR = PROJECT_ROOT / "test-data" / DEFAULT_BATCH_CODE / "packages"
DEFAULT_EXPECTED_ARTICLES = 100


@dataclass(frozen=True)
class PackageContents:
    part: int
    path: Path
    article_slugs: tuple[str, ...]


def log(message: str) -> None:
    timestamp = datetime.now().isoformat(timespec="seconds")
    print(f"[{timestamp}] {message}", flush=True)


def parse_parts(value: str) -> list[int]:
    """Parse comma-separated part numbers and inclusive ranges."""

    parts: set[int] = set()
    for raw_token in value.split(","):
        token = raw_token.strip()
        if not token:
            raise argparse.ArgumentTypeError("Part list contains an empty item.")
        if "-" in token:
            bounds = token.split("-")
            if len(bounds) != 2 or not all(item.strip().isdigit() for item in bounds):
                raise argparse.ArgumentTypeError(
                    f"Invalid part range {token!r}; use a value such as 3-5."
                )
            start, end = (int(item.strip()) for item in bounds)
            if start < 1 or end < 1 or start > end:
                raise argparse.ArgumentTypeError(
                    f"Invalid part range {token!r}; ranges must be positive and ascending."
                )
            parts.update(range(start, end + 1))
        else:
            if not token.isdigit() or int(token) < 1:
                raise argparse.ArgumentTypeError(
                    f"Invalid part number {token!r}; part numbers start at 1."
                )
            parts.add(int(token))
    if not parts:
        raise argparse.ArgumentTypeError("At least one part is required.")
    return sorted(parts)


def package_path(package_dir: Path, batch_code: str, part: int) -> Path:
    return package_dir / f"{batch_code}-part-{part:02d}.zip"


def read_package_contents(
    path: Path, *, part: int, expected_articles: int
) -> PackageContents:
    """Read and validate the article identities carried by one generated package."""

    if not path.is_file():
        raise FileNotFoundError(f"Package does not exist: {path}")

    try:
        with ZipFile(path) as archive:
            names = set(archive.namelist())
            if "articles.xlsx" not in names:
                raise RuntimeError(f"{path.name} does not contain articles.xlsx")

            if "manifest.json" in names:
                manifest = json.loads(archive.read("manifest.json"))
                manifest_part = manifest.get("part")
                if manifest_part is not None and int(manifest_part) != part:
                    raise RuntimeError(
                        f"{path.name} manifest part is {manifest_part}, expected {part}"
                    )
                manifest_count = manifest.get("article_count")
                if (
                    manifest_count is not None
                    and int(manifest_count) != expected_articles
                ):
                    raise RuntimeError(
                        f"{path.name} manifest declares {manifest_count} articles; "
                        f"expected {expected_articles}"
                    )

            workbook_data = archive.read("articles.xlsx")
    except BadZipFile as exc:
        raise RuntimeError(f"Invalid ZIP package: {path}") from exc

    workbook = load_workbook(BytesIO(workbook_data), read_only=True, data_only=True)
    try:
        sheet = (
            workbook["articles"]
            if "articles" in workbook.sheetnames
            else workbook.active
        )
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise RuntimeError(f"{path.name} articles.xlsx is empty") from exc

        headers = [
            str(value).strip() if value is not None else "" for value in raw_headers
        ]
        try:
            slug_index = headers.index("slug")
        except ValueError as exc:
            raise RuntimeError(f"{path.name} articles.xlsx has no slug column") from exc

        slugs = []
        for row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue
            slug = str(row[slug_index] or "").strip()
            if not slug:
                raise RuntimeError(f"{path.name} row {row_number} has no slug")
            slugs.append(slug)
    finally:
        workbook.close()

    if len(slugs) != expected_articles:
        raise RuntimeError(
            f"{path.name} contains {len(slugs)} article rows; expected {expected_articles}"
        )
    if len(set(slugs)) != len(slugs):
        raise RuntimeError(f"{path.name} contains duplicate article slugs")

    return PackageContents(part=part, path=path, article_slugs=tuple(slugs))


def get_operator(username: str):
    try:
        operator = get_user_model().objects.get(username=username)
    except get_user_model().DoesNotExist as exc:
        raise RuntimeError(f"Import operator does not exist: {username}") from exc
    if not operator.is_active:
        raise RuntimeError(f"Import operator is inactive: {username}")
    return operator


def existing_article_count(contents: PackageContents) -> int:
    return ArticlePage.objects.filter(static_slug__in=contents.article_slugs).count()


def failed_rows(job) -> list[dict]:
    return list(
        job.rows.filter(status=ImportRowStatus.FAILED)
        .order_by("row_no")
        .values("row_no", "error_code", "error_message")[:10]
    )


def import_package(contents: PackageContents, operator) -> tuple[str, int | None]:
    """Import one part, returning (result, job_id)."""

    expected_count = len(contents.article_slugs)
    existing_before = existing_article_count(contents)
    if existing_before == expected_count:
        log(
            f"Part {contents.part:02d}: all {expected_count} canonical articles "
            "already exist; skipped."
        )
        return "skipped", None

    if existing_before:
        log(
            f"Part {contents.part:02d}: {existing_before}/{expected_count} canonical "
            "articles already exist; importing the complete package idempotently."
        )

    context = ArticleImportContext(scope="global")
    log(f"Part {contents.part:02d}: previewing {contents.path.name}")
    with contents.path.open("rb") as source:
        job = preview_article_import(source, context=context, operator=operator)
    job.refresh_from_db()
    if job.status != ImportJobStatus.READY or job.failed_rows:
        raise RuntimeError(
            f"Preview failed for {contents.path.name}: status={job.status}, "
            f"failed_rows={job.failed_rows}, failures={failed_rows(job)}"
        )

    log(f"Part {contents.part:02d}: confirming import job {job.pk}")
    job = confirm_article_import(job, operator=operator)
    if job.status != ImportJobStatus.PENDING:
        raise RuntimeError(
            f"Confirmation failed for {contents.path.name}: status={job.status}"
        )

    log(f"Part {contents.part:02d}: executing import job {job.pk}")
    job = execute_confirmed_article_import(job, operator=operator)
    job.refresh_from_db()
    if job.status != ImportJobStatus.COMPLETED or job.failed_rows:
        raise RuntimeError(
            f"Import failed for {contents.path.name}: status={job.status}, "
            f"failed_rows={job.failed_rows}, failures={failed_rows(job)}"
        )

    existing_after = existing_article_count(contents)
    if existing_after != expected_count:
        raise RuntimeError(
            f"Part {contents.part:02d} finished job {job.pk}, but only "
            f"{existing_after}/{expected_count} canonical articles exist."
        )

    log(
        f"Part {contents.part:02d}: completed job {job.pk}; "
        f"success_rows={job.success_rows}, canonical_articles={existing_after}."
    )
    return "imported", job.pk


def recycle_database_connection(part: int) -> None:
    """Drop and immediately re-establish the default DB connection."""

    connection.close()
    connection.ensure_connection()
    log(f"Part {part:02d}: database connection closed and reconnected.")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Import selected generated article package parts through the official "
            "preview -> confirm -> execute workflow."
        )
    )
    parser.add_argument(
        "--parts",
        required=True,
        type=parse_parts,
        help="Part numbers or inclusive ranges, for example 3-5 or 1,3-5,12.",
    )
    parser.add_argument(
        "--package-dir",
        type=Path,
        default=DEFAULT_PACKAGE_DIR,
        help=f"Package directory (default: {DEFAULT_PACKAGE_DIR}).",
    )
    parser.add_argument(
        "--batch-code",
        default=DEFAULT_BATCH_CODE,
        help=f"Package filename prefix (default: {DEFAULT_BATCH_CODE}).",
    )
    parser.add_argument(
        "--operator",
        default="project_lead_a",
        help="Username used for import permission checks and audit records.",
    )
    parser.add_argument(
        "--expected-articles",
        type=int,
        default=DEFAULT_EXPECTED_ARTICLES,
        help=(
            "Required article count in each package "
            f"(default: {DEFAULT_EXPECTED_ARTICLES})."
        ),
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if args.expected_articles < 1:
        raise SystemExit("--expected-articles must be a positive integer")

    package_dir = args.package_dir.resolve()
    contents_by_part = [
        read_package_contents(
            package_path(package_dir, args.batch_code, part),
            part=part,
            expected_articles=args.expected_articles,
        )
        for part in args.parts
    ]
    operator = get_operator(args.operator)

    imported = 0
    skipped = 0
    job_ids = []
    for contents in contents_by_part:
        try:
            result, job_id = import_package(contents, operator)
            if result == "skipped":
                skipped += 1
            else:
                imported += 1
                if job_id is not None:
                    job_ids.append(job_id)
        finally:
            recycle_database_connection(contents.part)

    log(
        f"Finished parts {','.join(str(part) for part in args.parts)}: "
        f"imported={imported}, skipped={skipped}, job_ids={job_ids}."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
