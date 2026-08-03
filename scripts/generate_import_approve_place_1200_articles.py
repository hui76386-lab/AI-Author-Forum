#!/usr/bin/env python
"""Generate, import, approve, and place 1,200 illustrated test articles.

The script is deterministic and resumable. It uses the project's official article
preview -> confirm -> execute import services, ArticlePage review methods, and the
manual journal placement service. Source photos are normalized from the supplied
material_images.zip and registered as Wagtail images with an audit record.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ai_author_forum.settings.dev")
# The requested operation stops at placement. Avoid an implicit static publish storm.
os.environ.setdefault("STATIC_PUBLISH_AUTO_ON_PLACEMENT_CHANGE", "false")

import django

django.setup()

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps

from ai_author_forum.articles.import_services import (
    ArticleImportContext,
    confirm_article_import,
    execute_confirmed_article_import,
    preview_article_import,
)
from ai_author_forum.articles.models import ArticlePage, ArticleReviewRecord
from ai_author_forum.images.models import CustomImage
from ai_author_forum.journals.category_services import create_category
from ai_author_forum.journals.models import (
    ArticleImportJob,
    ImportJobStatus,
    Journal,
    JournalCategory,
)
from ai_author_forum.placements.models import ArticlePlacement, LayoutSlot
from ai_author_forum.placements.services import bulk_place_articles_in_journal
from ai_author_forum.site_settings.models import (
    AuditAction,
    AuditLog,
    AuditStatus,
)
from ai_author_forum.site_settings.services import record_audit_event

BATCH_CODE = "aaf1200-20260730"
ARTICLES_PER_JOURNAL = 10
JOURNALS_PER_PACKAGE = 10
MATERIAL_IMAGE_COUNT = 20
HEADERS = [
    "journal_slug",
    "title",
    "slug",
    "article_type",
    "authors",
    "ai_co_authors",
    "abstract",
    "keywords",
    "publication_date",
    "body_html",
    "html_file",
    "docx_file",
    "markdown_file",
    "cover_image",
    "primary_category_code",
    "primary_category_path",
    "related_category_codes",
    "related_category_paths",
    "notes",
]
THEMES = [
    (
        "Research Foundations and Open Questions",
        "foundations, research agenda, open questions",
    ),
    (
        "Benchmark Design for Reliable Progress",
        "benchmarks, measurement, comparative evaluation",
    ),
    (
        "Robust Evaluation Under Distribution Shift",
        "robustness, evaluation, distribution shift",
    ),
    ("Dataset and Evidence Design", "datasets, evidence quality, data governance"),
    ("Scalable System Architecture", "systems, scalability, architecture"),
    (
        "Safety, Reliability, and Failure Analysis",
        "safety, reliability, failure analysis",
    ),
    (
        "Human-AI Collaboration in Practice",
        "human-ai collaboration, workflows, usability",
    ),
    (
        "Reproducibility and Responsible Governance",
        "reproducibility, governance, transparency",
    ),
    ("Real-World Deployment Lessons", "deployment, operations, monitoring"),
    (
        "Future Directions and Research Roadmap",
        "future directions, roadmap, emerging methods",
    ),
]


def log(message: str) -> None:
    print(f"[{datetime.now().isoformat(timespec='seconds')}] {message}", flush=True)


def slug_for(journal_slug: str, number: int) -> str:
    return f"{BATCH_CODE}-{journal_slug}-{number:02d}"


def image_title(index: int) -> str:
    return f"{BATCH_CODE}-image-{index:02d}"


def read_journal_source(path: Path) -> list[dict]:
    if path.suffix.lower() == ".zip":
        with ZipFile(path) as archive:
            payload = archive.read("journals.xlsx")
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    records = []
    for values in rows[1:]:
        record = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        if record.get("slug") and record.get("journal_name"):
            records.append(record)
    records.sort(key=lambda item: (int(item.get("sort_order") or 0), str(item["slug"])))
    if len(records) != 120:
        raise RuntimeError(f"Expected 120 journal source rows, found {len(records)}")
    return records


def normalized_materials(image_zip: Path, output_dir: Path) -> list[Path]:
    target = output_dir / "material-images"
    target.mkdir(parents=True, exist_ok=True)
    with ZipFile(image_zip) as archive:
        jpg_names = sorted(
            name
            for name in archive.namelist()
            if name.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
        )
        if len(jpg_names) < MATERIAL_IMAGE_COUNT:
            raise RuntimeError(
                f"Need at least {MATERIAL_IMAGE_COUNT} images; found {len(jpg_names)}"
            )
        # Sample across the whole archive instead of taking only adjacent files.
        selected = [
            jpg_names[(i * len(jpg_names)) // MATERIAL_IMAGE_COUNT]
            for i in range(MATERIAL_IMAGE_COUNT)
        ]
        outputs = []
        for idx, source_name in enumerate(selected, start=1):
            destination = target / f"{image_title(idx)}.jpg"
            if not destination.exists():
                with archive.open(source_name) as source:
                    with Image.open(source) as raw:
                        rgb = ImageOps.exif_transpose(raw).convert("RGB")
                        rgb.thumbnail((1200, 675), Image.Resampling.LANCZOS)
                        canvas = Image.new("RGB", (1200, 675), (245, 245, 245))
                        x = (1200 - rgb.width) // 2
                        y = (675 - rgb.height) // 2
                        canvas.paste(rgb, (x, y))
                        canvas.save(
                            destination,
                            "JPEG",
                            quality=80,
                            optimize=True,
                            progressive=True,
                        )
            outputs.append(destination)
    return outputs


def article_body(journal: dict, number: int, theme: str) -> str:
    name = str(journal["journal_name"]).strip()
    description = str(
        journal.get("seo_description")
        or journal.get("homepage_intro")
        or "advanced AI research"
    ).strip()
    return (
        f"<h2>Overview</h2>"
        f"<p>This test article examines <strong>{theme.lower()}</strong> for {name}. "
        f"The journal focuses on {description} The discussion is designed to exercise the complete "
        f"content-import, moderation, and placement workflow.</p>"
        f"<h2>Research Context</h2>"
        f"<p>Current work in this area must connect clear problem definitions with measurable evidence. "
        f"For {name}, that means documenting assumptions, selecting representative tasks, and explaining "
        f"how results may change across datasets, model families, and deployment environments.</p>"
        f"<h2>Methods and Evaluation</h2>"
        f"<p>A reliable study should combine controlled experiments, transparent baselines, ablation analysis, "
        f"and qualitative inspection. Evaluation should report uncertainty, known limitations, resource costs, "
        f"and the conditions under which the findings are expected to generalize.</p>"
        f"<h2>Operational Considerations</h2>"
        f"<p>Production use introduces monitoring, versioning, governance, and incident-response requirements. "
        f"Teams should preserve reproducible artifacts and maintain traceable decisions from data preparation "
        f"through model release and post-deployment review.</p>"
        f"<h2>Conclusion</h2>"
        f"<p>Chapter {number} provides deterministic, journal-specific test content. It is intentionally imported "
        f"as a draft before formal submission, approval, and controlled placement in the journal latest-articles slot.</p>"
    )


def build_packages(
    journals: list[dict], output_dir: Path, image_zip: Path
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    normalized_materials(image_zip, output_dir)
    package_dir = output_dir / "packages"
    package_dir.mkdir(exist_ok=True)
    packages = []
    base_time = datetime(2026, 7, 30, 9, 0, tzinfo=timezone(timedelta(hours=8)))
    for package_index in range(0, len(journals), JOURNALS_PER_PACKAGE):
        group = journals[package_index : package_index + JOURNALS_PER_PACKAGE]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "articles"
        sheet.append(HEADERS)
        for journal_offset, journal in enumerate(group):
            global_journal_index = package_index + journal_offset
            for number, (theme, keywords) in enumerate(THEMES, start=1):
                slug = slug_for(str(journal["slug"]), number)
                image_index = (
                    (global_journal_index * ARTICLES_PER_JOURNAL + number - 1)
                    % MATERIAL_IMAGE_COUNT
                ) + 1
                title = f"{theme}: A {journal['journal_name']} Test Study"
                abstract = (
                    f"A deterministic import test article for {journal['journal_name']} examining "
                    f"{theme.lower()} within the journal's research scope."
                )
                sheet.append(
                    [
                        journal["slug"],
                        title,
                        slug,
                        "ai_article",
                        f"{journal['journal_name']} Editorial Test Team",
                        "AI Author Forum Test Assistant",
                        abstract,
                        f"{keywords}, {journal['journal_name']}, import test",
                        (
                            base_time
                            - timedelta(days=global_journal_index % 30, hours=number)
                        ).isoformat(),
                        article_body(journal, number, theme),
                        "",
                        "",
                        "",
                        image_title(image_index),
                        "research",
                        "",
                        "",
                        "",
                        f"{BATCH_CODE}; package {package_index // JOURNALS_PER_PACKAGE + 1:02d}; "
                        f"journal {journal['slug']}; article {number:02d}",
                    ]
                )
        buffer = BytesIO()
        workbook.save(buffer)
        package_number = package_index // JOURNALS_PER_PACKAGE + 1
        package_path = package_dir / f"{BATCH_CODE}-part-{package_number:02d}.zip"
        with ZipFile(package_path, "w", ZIP_DEFLATED) as archive:
            archive.writestr("articles.xlsx", buffer.getvalue())
            archive.writestr(
                "manifest.json",
                json.dumps(
                    {
                        "batch": BATCH_CODE,
                        "part": package_number,
                        "journal_slugs": [row["slug"] for row in group],
                        "article_count": len(group) * ARTICLES_PER_JOURNAL,
                        "image_source": str(image_zip),
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
            )
        packages.append(package_path)
    manifest = {
        "batch": BATCH_CODE,
        "journals": len(journals),
        "articles_per_journal": ARTICLES_PER_JOURNAL,
        "articles": len(journals) * ARTICLES_PER_JOURNAL,
        "packages": [str(path) for path in packages],
        "material_images_used": MATERIAL_IMAGE_COUNT,
        "image_zip": str(image_zip),
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return packages


def get_operator(username: str):
    user = get_user_model().objects.get(username=username)
    if not user.is_active or not user.is_superuser:
        raise RuntimeError(f"Operator {username} must be an active superuser")
    return user


def bootstrap_images(paths: list[Path], operator) -> list[CustomImage]:
    images = []
    created = 0
    for idx, path in enumerate(paths, start=1):
        title = image_title(idx)
        image = CustomImage.objects.filter(title=title).order_by("pk").first()
        if image is None:
            raw = path.read_bytes()
            image = CustomImage(
                title=title, description=f"{BATCH_CODE} from material_images.zip"
            )
            image.file = ContentFile(raw, name=path.name)
            image.save()
            created += 1
        images.append(image)
        log(
            f"Material image {idx:02d}/{len(paths)} ready (id={image.pk}, created={image.title == title and created > 0})"
        )
    record_audit_event(
        action=AuditAction.IMPORT,
        status=AuditStatus.SUCCESS,
        actor=operator,
        target_type="MaterialImageBatch",
        target_id=BATCH_CODE,
        target_label=BATCH_CODE,
        message="Registered reusable article images from material_images.zip",
        metadata={
            "image_count": len(images),
            "created_count": created,
            "titles": [image.title for image in images],
        },
    )
    return images


def expected_slugs(journals: list[dict]) -> list[str]:
    return [
        slug_for(str(journal["slug"]), number)
        for journal in journals
        for number in range(1, 11)
    ]


def bootstrap_categories(journals: list[dict], operator) -> dict[str, JournalCategory]:
    db_journals = {
        journal.slug: journal
        for journal in Journal.objects.filter(
            slug__in=[row["slug"] for row in journals]
        )
    }
    categories = {}
    created = 0
    for index, source in enumerate(journals, start=1):
        journal = db_journals[str(source["slug"])]
        category = JournalCategory.objects.filter(
            journal=journal, code="research"
        ).first()
        if category is None:
            result = create_category(
                journal=journal,
                data={
                    "name": "Research",
                    "code": "research",
                    "slug": "research",
                    "description": f"Primary test category for {journal.name}",
                    "status": "active",
                    "show_in_navigation": False,
                    "generate_static_page": True,
                    "sort_order": 10,
                },
                actor=operator,
                request_id=BATCH_CODE,
            )
            category = result.category
            created += 1
        categories[journal.slug] = category
        if index % 20 == 0 or index == len(journals):
            log(f"Category bootstrap: {index}/{len(journals)} journals")
    log(
        f"Category bootstrap complete: {created} created, {len(categories) - created} reused"
    )
    return categories


def import_packages(journals: list[dict], packages: list[Path], operator) -> list[int]:
    job_ids = []
    context = ArticleImportContext(scope="global")
    for index, package in enumerate(packages, start=1):
        group = journals[
            (index - 1) * JOURNALS_PER_PACKAGE : index * JOURNALS_PER_PACKAGE
        ]
        slugs = expected_slugs(group)
        existing = ArticlePage.objects.filter(static_slug__in=slugs).count()
        if existing == len(slugs):
            log(
                f"Package {index:02d}: all {existing} canonical articles already exist; import skipped for resume"
            )
            continue
        log(f"Package {index:02d}: previewing {package.name}")
        with package.open("rb") as source:
            job = preview_article_import(source, context=context, operator=operator)
        job.refresh_from_db()
        if job.status != ImportJobStatus.READY or job.failed_rows:
            failures = list(
                job.rows.exclude(status__in=("success", "skipped")).values(
                    "row_no", "error_code", "error_message"
                )[:10]
            )
            raise RuntimeError(
                f"Preview failed for {package.name}: status={job.status}, failed={job.failed_rows}, rows={failures}"
            )
        log(f"Package {index:02d}: preview ready; confirming job {job.pk}")
        job = confirm_article_import(job, operator=operator)
        job = execute_confirmed_article_import(job, operator=operator)
        job.refresh_from_db()
        if job.status != ImportJobStatus.COMPLETED or job.failed_rows:
            failures = list(
                job.rows.exclude(status__in=("success", "skipped")).values(
                    "row_no", "error_code", "error_message"
                )[:10]
            )
            raise RuntimeError(
                f"Import failed for {package.name}: status={job.status}, failed={job.failed_rows}, rows={failures}"
            )
        count = ArticlePage.objects.filter(static_slug__in=slugs).count()
        if count != len(slugs):
            raise RuntimeError(
                f"Package {index:02d} created {count}/{len(slugs)} canonical pages"
            )
        job_ids.append(job.pk)
        log(
            f"Package {index:02d}: imported {job.success_rows} rows successfully as job {job.pk}"
        )
    return job_ids


def review_articles(journals: list[dict], operator) -> None:
    slugs = expected_slugs(journals)
    pages = ArticlePage.objects.filter(static_slug__in=slugs).order_by(
        "primary_journal_id", "static_slug"
    )
    total = pages.count()
    if total != len(slugs):
        raise RuntimeError(f"Cannot review: found {total}/{len(slugs)} articles")
    for index, page in enumerate(pages.iterator(chunk_size=50), start=1):
        if page.review_status not in (
            ArticlePage.ReviewStatus.SUBMITTED,
            ArticlePage.ReviewStatus.APPROVED,
        ):
            page.submit_for_review(
                operator, comment=f"{BATCH_CODE}: automated formal submission"
            )
            page.refresh_from_db()
        if page.review_status != ArticlePage.ReviewStatus.APPROVED:
            page.approve(
                operator,
                comment=f"{BATCH_CODE}: approved for controlled test placement",
            )
        if index % 50 == 0 or index == total:
            log(f"Review progress: {index}/{total}")


def place_articles(journals: list[dict], operator) -> None:
    slot = LayoutSlot.objects.get(code="journal_latest")
    if (
        not slot.is_active
        or slot.scope != LayoutSlot.Scope.JOURNAL
        or slot.max_items < ARTICLES_PER_JOURNAL
    ):
        raise RuntimeError(
            f"journal_latest slot is not usable: scope={slot.scope}, max_items={slot.max_items}, active={slot.is_active}"
        )
    db_journals = {
        journal.slug: journal
        for journal in Journal.objects.filter(
            slug__in=[row["slug"] for row in journals]
        )
    }
    for index, source in enumerate(journals, start=1):
        journal = db_journals[str(source["slug"])]
        slugs = [slug_for(journal.slug, number) for number in range(1, 11)]
        articles = list(
            ArticlePage.objects.filter(static_slug__in=slugs).order_by("static_slug")
        )
        if len(articles) != ARTICLES_PER_JOURNAL:
            raise RuntimeError(
                f"{journal.slug}: expected 10 articles, found {len(articles)}"
            )
        active_count = ArticlePlacement.objects.filter(
            article__in=articles,
            slot=slot,
            target_type=ArticlePlacement.TargetType.JOURNAL,
            target_slug=journal.slug,
            is_active=True,
        ).count()
        if active_count < ARTICLES_PER_JOURNAL:
            bulk_place_articles_in_journal(
                articles=articles, journal=journal, slot=slot, actor=operator
            )
        if index % 10 == 0 or index == len(journals):
            log(f"Placement progress: {index}/{len(journals)} journals")


def verify(journals: list[dict], output_dir: Path) -> dict:
    target_journal_slugs = [str(row["slug"]) for row in journals]
    slugs = expected_slugs(journals)
    pages = ArticlePage.objects.filter(static_slug__in=slugs)
    placements = ArticlePlacement.objects.filter(
        article__in=pages,
        is_active=True,
        slot__code="journal_latest",
        target_type=ArticlePlacement.TargetType.JOURNAL,
    )
    article_counts = Counter(pages.values_list("primary_journal__slug", flat=True))
    placement_counts = Counter(placements.values_list("target_slug", flat=True))
    image_ids = list(pages.values_list("featured_image_id", flat=True))
    images = CustomImage.objects.filter(pk__in={pk for pk in image_ids if pk})
    missing_files = []
    for image in images:
        if not image.file or not default_storage.exists(image.file.name):
            missing_files.append(
                {
                    "id": image.pk,
                    "title": image.title,
                    "file": image.file.name if image.file else "",
                }
            )
    submitted_records = ArticleReviewRecord.objects.filter(
        article__in=pages, action=ArticlePage.ReviewStatus.SUBMITTED
    ).count()
    approved_records = ArticleReviewRecord.objects.filter(
        article__in=pages, action=ArticlePage.ReviewStatus.APPROVED
    ).count()
    completed_jobs = ArticleImportJob.objects.filter(
        package_name__startswith=BATCH_CODE, status=ImportJobStatus.COMPLETED
    )
    failed_job_rows = sum(completed_jobs.values_list("failed_rows", flat=True))
    result = {
        "batch": BATCH_CODE,
        "target_active_journals": Journal.objects.filter(
            slug__in=target_journal_slugs, status="active"
        ).count(),
        "articles": pages.count(),
        "articles_with_images": pages.exclude(featured_image_id__isnull=True).count(),
        "approved_articles": pages.filter(
            review_status=ArticlePage.ReviewStatus.APPROVED
        ).count(),
        "placed_articles": placements.values("article_id").distinct().count(),
        "active_placements": placements.count(),
        "submit_review_records": submitted_records,
        "approval_review_records": approved_records,
        "completed_import_jobs": completed_jobs.count(),
        "failed_import_rows": failed_job_rows,
        "unique_images": images.count(),
        "missing_image_files": missing_files,
        "article_count_violations": {
            slug: article_counts.get(slug, 0)
            for slug in target_journal_slugs
            if article_counts.get(slug, 0) != 10
        },
        "placement_count_violations": {
            slug: placement_counts.get(slug, 0)
            for slug in target_journal_slugs
            if placement_counts.get(slug, 0) != 10
        },
        "wrong_target_placements": 0,
        "audit_import_events": AuditLog.objects.filter(
            action=AuditAction.IMPORT,
            target_id__in=[
                BATCH_CODE,
                *[str(pk) for pk in completed_jobs.values_list("pk", flat=True)],
            ],
        ).count(),
        "audit_article_approval_events": AuditLog.objects.filter(
            action=ArticlePage.ReviewStatus.APPROVED,
            target_id__in=[str(pk) for pk in pages.values_list("pk", flat=True)],
        ).count(),
        "audit_placement_events": AuditLog.objects.filter(
            action=AuditAction.CONFIGURE, metadata__target_slug__in=target_journal_slugs
        ).count(),
    }
    # Explicit target correctness check without relying on a cross-table expression in JSON output.
    result["wrong_target_placements"] = placements.exclude(
        target_slug__in=target_journal_slugs
    ).count() + sum(
        1
        for article_journal, target_slug in placements.values_list(
            "article__primary_journal__slug", "target_slug"
        )
        if article_journal != target_slug
    )
    required = {
        "target_active_journals": 120,
        "articles": 1200,
        "articles_with_images": 1200,
        "approved_articles": 1200,
        "placed_articles": 1200,
        "active_placements": 1200,
    }
    errors = []
    for key, expected in required.items():
        if result[key] != expected:
            errors.append(f"{key}={result[key]} expected {expected}")
    if result["submit_review_records"] < 1200:
        errors.append(
            f"submit_review_records={result['submit_review_records']} expected at least 1200"
        )
    if result["approval_review_records"] < 1200:
        errors.append(
            f"approval_review_records={result['approval_review_records']} expected at least 1200"
        )
    if result["completed_import_jobs"] < 12:
        errors.append(
            f"completed_import_jobs={result['completed_import_jobs']} expected at least 12"
        )
    if result["failed_import_rows"]:
        errors.append(f"failed_import_rows={result['failed_import_rows']}")
    if result["missing_image_files"]:
        errors.append(f"missing_image_files={len(result['missing_image_files'])}")
    if result["article_count_violations"]:
        errors.append(
            f"article_count_violations={len(result['article_count_violations'])}"
        )
    if result["placement_count_violations"]:
        errors.append(
            f"placement_count_violations={len(result['placement_count_violations'])}"
        )
    if result["wrong_target_placements"]:
        errors.append(f"wrong_target_placements={result['wrong_target_placements']}")
    result["errors"] = errors
    report_path = output_dir / "verification.json"
    report_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    log(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    if errors:
        raise RuntimeError("Verification failed: " + "; ".join(errors))
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "mode", choices=("generate", "apply", "verify", "all"), nargs="?", default="all"
    )
    parser.add_argument(
        "--journal-source",
        type=Path,
        default=Path("output/ai-author-forum-120-journals-rich-import-20260727.zip"),
    )
    parser.add_argument(
        "--image-zip",
        type=Path,
        default=Path(r"C:\Users\18700\Desktop\material_images.zip"),
    )
    parser.add_argument(
        "--output-dir", type=Path, default=Path("test-data") / BATCH_CODE
    )
    parser.add_argument("--operator", default="project_lead_a")
    args = parser.parse_args()

    journals = read_journal_source(args.journal_source)
    packages = [
        args.output_dir / "packages" / f"{BATCH_CODE}-part-{number:02d}.zip"
        for number in range(1, 13)
    ]
    if args.mode in ("generate", "all"):
        packages = build_packages(journals, args.output_dir, args.image_zip)
        log(f"Generated {len(packages)} packages containing 1,200 article rows")
    if args.mode in ("apply", "all"):
        if not all(path.exists() for path in packages):
            packages = build_packages(journals, args.output_dir, args.image_zip)
        active = Journal.objects.filter(
            slug__in=[row["slug"] for row in journals], status="active"
        ).count()
        if active != 120:
            raise RuntimeError(
                f"Authoritative database must contain 120 active target journals; found {active}"
            )
        operator = get_operator(args.operator)
        images = normalized_materials(args.image_zip, args.output_dir)
        bootstrap_images(images, operator)
        bootstrap_categories(journals, operator)
        import_packages(journals, packages, operator)
        review_articles(journals, operator)
        place_articles(journals, operator)
        verify(journals, args.output_dir)
    elif args.mode == "verify":
        verify(journals, args.output_dir)


if __name__ == "__main__":
    main()
